import asyncio
import io
from pathlib import Path
from types import SimpleNamespace
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from models import PersohubEventParticipantMode, PersohubEventRoundState
from routers.persohub_events_admin import event_leaderboard, export_leaderboard, export_participants


class FakeQuery:
    def __init__(self, rows):
        self._rows = rows

    def filter(self, *args, **kwargs):
        return self

    def order_by(self, *args, **kwargs):
        return self

    def all(self):
        return list(self._rows)


class FakeDb:
    def __init__(self, *, round_scope_rows, round_meta_rows=None, full_round_rows=None):
        self._round_scope_rows = list(round_scope_rows)
        self._round_meta_rows = list(round_meta_rows if round_meta_rows is not None else round_scope_rows)
        self._full_round_rows = list(full_round_rows if full_round_rows is not None else self._round_meta_rows)

    def query(self, *args):
        if len(args) == 1 and getattr(args[0], "__name__", None) == "PersohubEventRound":
            return FakeQuery(self._full_round_rows)

        column_names = [getattr(arg, "name", None) for arg in args]
        if column_names == ["id", "state", "is_frozen"]:
            rows = [
                SimpleNamespace(
                    id=round_row.id,
                    state=round_row.state,
                    is_frozen=round_row.is_frozen,
                )
                for round_row in self._round_scope_rows
            ]
            return FakeQuery(rows)
        if column_names == ["id", "state", "is_frozen", "round_no"]:
            rows = [
                SimpleNamespace(
                    id=round_row.id,
                    state=round_row.state,
                    is_frozen=round_row.is_frozen,
                    round_no=round_row.round_no,
                )
                for round_row in self._round_scope_rows
            ]
            return FakeQuery(rows)
        if column_names == ["id", "round_no"]:
            rows = [
                SimpleNamespace(
                    id=round_row.id,
                    round_no=round_row.round_no,
                )
                for round_row in self._round_meta_rows
            ]
            return FakeQuery(rows)
        if column_names == ["user_id", "round_id", "normalized_score"]:
            return FakeQuery([])
        if column_names == ["team_id", "round_id", "total_score"]:
            return FakeQuery([])

        raise AssertionError(f"Unexpected query columns: {column_names}")


async def _read_streaming_response(response) -> bytes:
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk)
    return b"".join(chunks)


def _make_round(round_id, round_no, name, description, *, frozen=True):
    return SimpleNamespace(
        id=round_id,
        round_no=round_no,
        name=name,
        description=description,
        state=PersohubEventRoundState.COMPLETED,
        is_frozen=frozen,
        evaluation_criteria=None,
    )


def test_export_leaderboard_xlsx_includes_minimal_round_details_sheet(monkeypatch):
    event = SimpleNamespace(
        id=99,
        participant_mode=PersohubEventParticipantMode.INDIVIDUAL,
        event_code="PHX",
    )
    rounds = [
        _make_round(1, 1, "Prelims", "Screening round"),
        _make_round(2, 2, "Finals", None),
    ]
    db = FakeDb(
        round_scope_rows=rounds,
        round_meta_rows=[rounds[1]],
        full_round_rows=[rounds[1]],
    )

    monkeypatch.setattr("routers.persohub_events_admin._get_event_or_404", lambda db, slug: event)
    monkeypatch.setattr(
        "routers.persohub_events_admin.event_leaderboard",
        lambda **kwargs: [
            {
                "entity_id": 7,
                "register_number": "RA2311007010001",
                "name": "Wildcard Entry",
                "department": "ECE",
                "gender": "Female",
                "batch": "2023",
                "status": "Active",
                "rounds_participated": 1,
                "attendance_count": 1,
                "referral_count": 0,
                "cumulative_score": 88.5,
                "rank": 1,
                "is_wildcard": True,
                "wildcard_seed_score": 41.5,
                "wildcard_start_round_no": 2,
            }
        ],
    )
    monkeypatch.setattr("routers.persohub_events_admin.round_participants", lambda **kwargs: [])

    response = export_leaderboard(
        slug="demo",
        format="xlsx",
        round_ids=[2],
        admin=SimpleNamespace(id=1),
        db=db,
    )
    content = asyncio.run(_read_streaming_response(response))
    workbook = load_workbook(io.BytesIO(content))

    assert workbook.sheetnames == ["Leaderboard", "Wildcards", "Round Details", "Round 2"]

    leaderboard_sheet = workbook["Leaderboard"]
    assert [cell.value for cell in leaderboard_sheet[1]] == [
        "Si.No",
        "Register Number",
        "Name",
        "Department",
        "Gender",
        "Batch",
        "Status",
        "Is Wildcard",
        "Rounds",
        "Attendance",
        "Referral Count",
        "R2 Score",
        "R2 Rank",
        "Overall Score",
        "Overall Rank",
    ]
    assert [cell.value for cell in leaderboard_sheet[2]] == [
        1,
        "RA2311007010001",
        "Wildcard Entry",
        "ECE",
        "Female",
        "2023",
        "Active",
        "Yes",
        1,
        1,
        0,
        0,
        1,
        88.5,
        1,
    ]

    wildcard_sheet = workbook["Wildcards"]
    assert [cell.value for cell in wildcard_sheet[1]] == [
        "Si.No",
        "Register Number",
        "Name",
        "Status",
        "Wildcard Seed Score",
        "Wildcard Start Round",
    ]
    assert [cell.value for cell in wildcard_sheet[2]] == [
        1,
        "RA2311007010001",
        "Wildcard Entry",
        "Active",
        41.5,
        2,
    ]

    details_sheet = workbook["Round Details"]
    assert [cell.value for cell in details_sheet[1]] == ["Round ID", "Round No", "Round Name", "Description"]
    assert [cell.value for cell in details_sheet[2]] == [2, 2, "Finals", None]
    assert details_sheet.max_row == 2


def test_export_leaderboard_csv_remains_flat(monkeypatch):
    event = SimpleNamespace(
        id=99,
        participant_mode=PersohubEventParticipantMode.INDIVIDUAL,
        event_code="PHX",
    )
    rounds = [_make_round(1, 1, "Prelims", "Screening round")]
    db = FakeDb(round_scope_rows=rounds)

    monkeypatch.setattr("routers.persohub_events_admin._get_event_or_404", lambda db, slug: event)
    monkeypatch.setattr(
        "routers.persohub_events_admin.event_leaderboard",
        lambda **kwargs: [
            {
                "entity_id": 9,
                "register_number": "RA2311007010009",
                "name": "CSV Wildcard",
                "department": "CT",
                "gender": "Male",
                "batch": "2023",
                "status": "Active",
                "rounds_participated": 0,
                "attendance_count": 0,
                "referral_count": 0,
                "cumulative_score": 44.0,
                "rank": 1,
                "is_wildcard": True,
                "wildcard_seed_score": 19.0,
                "wildcard_start_round_no": 1,
            }
        ],
    )

    response = export_leaderboard(
        slug="demo",
        format="csv",
        round_ids=None,
        admin=SimpleNamespace(id=1),
        db=db,
    )
    content = asyncio.run(_read_streaming_response(response)).decode("utf-8")

    assert "Round Details" not in content
    assert content.splitlines()[0].startswith("Si.No,Register Number,Name,Department,Gender,Batch,Status,Is Wildcard")
    assert "Wildcard Seed Score" in content.splitlines()[0]
    assert "Wildcard Start Round" in content.splitlines()[0]
    assert "CSV Wildcard" in content
    assert "Yes" in content
    assert "19.0" in content


def test_event_leaderboard_excludes_no_score_all_absent_rows(monkeypatch):
    event = SimpleNamespace(
        id=77,
        participant_mode=PersohubEventParticipantMode.INDIVIDUAL,
        event_code="PHX",
    )
    rounds = [_make_round(1, 1, "Prelims", "Screening round")]
    db = FakeDb(round_scope_rows=rounds)

    monkeypatch.setattr("routers.persohub_events_admin._get_event_or_404", lambda db, slug: event)
    monkeypatch.setattr(
        "routers.persohub_events_admin._registered_entities",
        lambda db, event: [
            {
                "entity_id": 1,
                "regno_or_code": "RA1",
                "name": "Absent Only",
                "email": "absent@example.com",
                "department": "IT",
                "gender": "Male",
                "batch": "2023",
                "status": "Active",
                "is_wildcard": False,
            },
            {
                "entity_id": 2,
                "regno_or_code": "RA2",
                "name": "Present Zero",
                "email": "present@example.com",
                "department": "IT",
                "gender": "Female",
                "batch": "2023",
                "status": "Active",
                "is_wildcard": False,
            },
            {
                "entity_id": 3,
                "regno_or_code": "RA3",
                "name": "Wildcard Seed",
                "email": "wild@example.com",
                "department": "ECE",
                "gender": "Female",
                "batch": "2023",
                "status": "Active",
                "is_wildcard": True,
                "wildcard_seed_score": 12.5,
                "wildcard_start_round_no": 2,
            },
        ],
    )
    monkeypatch.setattr(
        "routers.persohub_events_admin._effective_score_metrics",
        lambda db, event, **kwargs: {
            1: {"cumulative_score": 0.0, "rounds_participated": 0},
            2: {"cumulative_score": 0.0, "rounds_participated": 1},
            3: {"cumulative_score": 12.5, "rounds_participated": 0},
        },
    )

    rows = event_leaderboard(
        slug="demo",
        status_filter=None,
        wildcard_filter=None,
        round_ids=None,
        sort="rank",
        page=1,
        page_size=20,
        response=None,
        _=SimpleNamespace(id=1),
        db=db,
    )

    assert [row["name"] for row in rows] == ["Wildcard Seed", "Present Zero"]
    assert all(row["name"] != "Absent Only" for row in rows)


def test_export_leaderboard_xlsx_excludes_no_score_all_absent_rows(monkeypatch):
    event = SimpleNamespace(
        id=88,
        participant_mode=PersohubEventParticipantMode.INDIVIDUAL,
        event_code="PHX",
    )
    rounds = [_make_round(1, 1, "Prelims", "Screening round")]
    db = FakeDb(round_scope_rows=rounds, round_meta_rows=rounds, full_round_rows=rounds)

    monkeypatch.setattr("routers.persohub_events_admin._get_event_or_404", lambda db, slug: event)
    monkeypatch.setattr(
        "routers.persohub_events_admin._registered_entities",
        lambda db, event: [
            {
                "entity_id": 1,
                "regno_or_code": "RA1",
                "name": "Absent Only",
                "email": "absent@example.com",
                "department": "IT",
                "gender": "Male",
                "batch": "2023",
                "status": "Active",
                "referral_count": 0,
                "is_wildcard": False,
            },
            {
                "entity_id": 2,
                "regno_or_code": "RA2",
                "name": "Present Zero",
                "email": "present@example.com",
                "department": "IT",
                "gender": "Female",
                "batch": "2023",
                "status": "Active",
                "referral_count": 0,
                "is_wildcard": False,
            },
        ],
    )
    monkeypatch.setattr(
        "routers.persohub_events_admin._effective_score_metrics",
        lambda db, event, **kwargs: {
            1: {"cumulative_score": 0.0, "rounds_participated": 0},
            2: {"cumulative_score": 0.0, "rounds_participated": 1},
        },
    )
    monkeypatch.setattr("routers.persohub_events_admin.round_participants", lambda **kwargs: [])

    response = export_leaderboard(
        slug="demo",
        format="xlsx",
        status_filter=None,
        wildcard_filter=None,
        round_ids=None,
        sort="rank",
        admin=SimpleNamespace(id=1),
        db=db,
    )
    content = asyncio.run(_read_streaming_response(response))
    workbook = load_workbook(io.BytesIO(content))
    leaderboard_sheet = workbook["Leaderboard"]

    assert leaderboard_sheet.max_row == 2
    assert [cell.value for cell in leaderboard_sheet[2]][2] == "Present Zero"
    sheet_values = [
        value
        for row in leaderboard_sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    assert "Absent Only" not in sheet_values


def test_export_participants_xlsx_includes_wildcards_sheet(monkeypatch):
    event = SimpleNamespace(
        id=101,
        participant_mode=PersohubEventParticipantMode.INDIVIDUAL,
        event_code="PHX",
    )

    monkeypatch.setattr("routers.persohub_events_admin._get_event_or_404", lambda db, slug: event)
    monkeypatch.setattr(
        "routers.persohub_events_admin.event_participants",
        lambda **kwargs: [
            {
                "register_number": "RA2311007010003",
                "regno_or_code": "RA2311007010003",
                "name": "Participant Wildcard",
                "email": "wild@example.com",
                "college": "MIT",
                "department": "IT",
                "gender": "Female",
                "batch": "2023",
                "status": "Active",
                "referral_code": "REF123",
                "referred_by": "SENIOR",
                "referral_count": 2,
                "is_wildcard": True,
                "wildcard_seed_score": 33.0,
                "wildcard_start_round_no": 4,
            }
        ],
    )

    response = export_participants(
        slug="demo",
        format="xlsx",
        db=SimpleNamespace(),
    )
    content = asyncio.run(_read_streaming_response(response))
    workbook = load_workbook(io.BytesIO(content))

    assert workbook.sheetnames == ["Participants", "Wildcards"]

    wildcard_sheet = workbook["Wildcards"]
    assert [cell.value for cell in wildcard_sheet[1]] == [
        "Si.No",
        "Register Number",
        "Name",
        "Status",
        "Wildcard Seed Score",
        "Wildcard Start Round",
    ]
    assert [cell.value for cell in wildcard_sheet[2]] == [
        1,
        "RA2311007010003",
        "Participant Wildcard",
        "Active",
        33.0,
        4,
    ]
