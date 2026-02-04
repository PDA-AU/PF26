from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File, Query, Header
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, case, text
import os
import logging
from pathlib import Path
from typing import List, Optional
from datetime import datetime, timezone
import shutil
import uuid
import io
from openpyxl import Workbook
from fastapi.responses import StreamingResponse
import csv
import boto3

from database import engine, get_db, Base
from models import User, Round, Score, SystemConfig, Program, Event, UserRole, RoundState, ParticipantStatus, Department, YearOfStudy, Gender, RoundMode
from schemas import (
    UserRegister, UserLogin, TokenResponse, RefreshTokenRequest, UserResponse, UserUpdate,
    ParticipantListResponse, RoundCreate, RoundUpdate, RoundResponse, RoundPublicResponse,
    ScoreEntry, ScoreUpdate, ScoreResponse, ParticipantRoundStatus, AdminParticipantRoundStat, LeaderboardEntry,
    DashboardStats, TopReferrer, DepartmentEnum, YearOfStudyEnum, GenderEnum, 
    ParticipantStatusEnum, RoundStateEnum,
    ProgramCreate, ProgramUpdate, ProgramResponse,
    EventCreate, EventUpdate, EventResponse
)
from auth import (
    verify_password, get_password_hash, create_access_token, create_refresh_token,
    decode_token, get_current_user, get_current_admin, generate_referral_code
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Create tables
Base.metadata.create_all(bind=engine)

# Create upload directory
UPLOAD_DIR = Path(os.environ.get('UPLOAD_DIR', '/app/backend/uploads'))
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

S3_BUCKET_NAME = os.environ.get("S3_BUCKET_NAME")
AWS_REGION = os.environ.get("AWS_REGION")
AWS_ACCESS_KEY_ID = os.environ.get("AWS_ACCESS_KEY_ID")
AWS_SECRET_ACCESS_KEY = os.environ.get("AWS_SECRET_ACCESS_KEY")

S3_CLIENT = boto3.client(
    "s3",
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY
)

app = FastAPI(title="Persofest'26 API", version="1.0.0")
api_router = APIRouter(prefix="/api")

# Mount static files for uploads
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

PDA_ADMIN_PASSWORD = "pda@1984"


def require_pda_admin(x_pda_admin: str = Header(None, alias="X-PDA-ADMIN")):
    if x_pda_admin != PDA_ADMIN_PASSWORD:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid PDA admin credentials")
    return True


def _ensure_round_description_pdf_column():
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name = 'rounds' AND column_name = 'description_pdf'
                    """
                )
            ).fetchone()
            if not result:
                conn.execute(text("ALTER TABLE rounds ADD COLUMN description_pdf VARCHAR(500)"))
                logger.info("Added rounds.description_pdf column")
    except Exception as exc:
        logger.warning(f"Could not ensure description_pdf column: {exc}")


def _build_s3_url(key: str) -> str:
    if not S3_BUCKET_NAME or not AWS_REGION:
        raise RuntimeError("S3 configuration missing")
    return f"https://{S3_BUCKET_NAME}.s3.{AWS_REGION}.amazonaws.com/{key}"


def _upload_to_s3(file: UploadFile, key_prefix: str, allowed_types: Optional[List[str]] = None) -> str:
    if not S3_BUCKET_NAME or not AWS_REGION:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="S3 not configured")
    if not file.content_type:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Missing file content type")
    if allowed_types and file.content_type not in allowed_types:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid file type")

    extension = Path(file.filename or "").suffix.lower()
    unique_name = f"{uuid.uuid4().hex}{extension}"
    key = f"{key_prefix.rstrip('/')}/{unique_name}"

    try:
        S3_CLIENT.upload_fileobj(
            file.file,
            S3_BUCKET_NAME,
            key,
            ExtraArgs={"ContentType": file.content_type}
        )
    except Exception as exc:
        logger.error(f"S3 upload failed: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Upload failed")

    return _build_s3_url(key)


# ==================== STARTUP ====================
@app.on_event("startup")
async def startup_event():
    db = next(get_db())
    try:
        _ensure_round_description_pdf_column()
        # Initialize system config
        reg_config = db.query(SystemConfig).filter(SystemConfig.key == "registration_open").first()
        if not reg_config:
            db.add(SystemConfig(key="registration_open", value="true"))
            db.commit()
        
        # Create default admin if not exists
        admin = db.query(User).filter(User.role == UserRole.ADMIN).first()
        if not admin:
            admin_user = User(
                register_number="0000000000",
                email="admin@persofest.com",
                hashed_password=get_password_hash("admin123"),
                name="Admin",
                phone="0000000000",
                gender=Gender.MALE,
                department=Department.AI_DS,
                year_of_study=YearOfStudy.FIRST,
                role=UserRole.ADMIN,
                referral_code=generate_referral_code(),
                status=ParticipantStatus.ACTIVE
            )
            db.add(admin_user)
            db.commit()
            logger.info("Default admin created: register_number=0000000000, password=admin123")
    finally:
        db.close()


# ==================== PUBLIC ROUTES ====================
@api_router.get("/")
async def root():
    return {"message": "Persofest'26 API is running"}


@api_router.get("/health")
async def health_check():
    return {"status": "healthy"}


# ==================== AUTH ROUTES ====================
@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserRegister, db: Session = Depends(get_db)):
    # Check registration status
    reg_config = db.query(SystemConfig).filter(SystemConfig.key == "registration_open").first()
    if reg_config and reg_config.value == "false":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Registrations are currently closed")
    
    # Check existing user
    if db.query(User).filter(User.register_number == user_data.register_number).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Register number already exists")
    if db.query(User).filter(User.email == user_data.email).first():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already registered")
    
    # Handle referral
    referred_by_code = None
    if user_data.referral_code:
        referrer = db.query(User).filter(User.referral_code == user_data.referral_code).first()
        if referrer:
            # Prevent self-referral
            if referrer.register_number != user_data.register_number:
                referrer.referral_count += 1
                referred_by_code = user_data.referral_code
    
    # Create user
    new_user = User(
        register_number=user_data.register_number,
        email=user_data.email,
        hashed_password=get_password_hash(user_data.password),
        name=user_data.name,
        phone=user_data.phone,
        gender=Gender[user_data.gender.name],
        department=Department[user_data.department.name],
        year_of_study=YearOfStudy[user_data.year_of_study.name],
        role=UserRole.PARTICIPANT,
        referral_code=generate_referral_code(),
        referred_by=referred_by_code,
        status=ParticipantStatus.ACTIVE
    )
    
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    # Generate tokens
    access_token = create_access_token(data={"sub": new_user.register_number})
    refresh_token = create_refresh_token(data={"sub": new_user.register_number})
    
    return TokenResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user=UserResponse.model_validate(new_user)
    )


@api_router.post("/auth/login", response_model=TokenResponse)
async def login(login_data: UserLogin, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.register_number == login_data.register_number).first()
    if not user or not verify_password(login_data.password, user.hashed_password):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid credentials")
    
    access_token = create_access_token(data={"sub": user.register_number})
    refresh_token = create_refresh_token(data={"sub": user.register_number})
    
    return TokenResponse(
        access_token=access_token,
        refresh_token=refresh_token,
        user=UserResponse.model_validate(user)
    )


@api_router.post("/auth/refresh", response_model=TokenResponse)
async def refresh_token(token_data: RefreshTokenRequest, db: Session = Depends(get_db)):
    payload = decode_token(token_data.refresh_token)
    if payload.get("type") != "refresh":
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token type")
    
    register_number = payload.get("sub")
    user = db.query(User).filter(User.register_number == register_number).first()
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="User not found")
    
    access_token = create_access_token(data={"sub": user.register_number})
    new_refresh_token = create_refresh_token(data={"sub": user.register_number})
    
    return TokenResponse(
        access_token=access_token,
        refresh_token=new_refresh_token,
        user=UserResponse.model_validate(user)
    )


# ==================== USER ROUTES ====================
@api_router.get("/me", response_model=UserResponse)
async def get_me(current_user: User = Depends(get_current_user)):
    return UserResponse.model_validate(current_user)


@api_router.put("/me", response_model=UserResponse)
async def update_me(update_data: UserUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if update_data.name:
        current_user.name = update_data.name
    if update_data.phone:
        current_user.phone = update_data.phone
    if update_data.email:
        existing = db.query(User).filter(User.email == update_data.email, User.id != current_user.id).first()
        if existing:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already in use")
        current_user.email = update_data.email
    
    db.commit()
    db.refresh(current_user)
    return UserResponse.model_validate(current_user)


@api_router.post("/me/profile-picture", response_model=UserResponse)
async def upload_profile_picture(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Validate file type
    if not file.content_type or not file.content_type.startswith("image/png"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only PNG images are allowed")

    # Validate file size (max 2MB)
    contents = await file.read()
    if len(contents) > 2 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File size exceeds 2MB limit")
    file.file = io.BytesIO(contents)

    s3_url = _upload_to_s3(file, "persofest/profiles", allowed_types=["image/png"])
    current_user.profile_picture = s3_url
    db.commit()
    db.refresh(current_user)
    
    return UserResponse.model_validate(current_user)


@api_router.get("/me/rounds", response_model=List[ParticipantRoundStatus])
async def get_my_round_status(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # Get all completed or active rounds
    rounds = db.query(Round).filter(Round.state.in_([RoundState.ACTIVE, RoundState.COMPLETED])).order_by(Round.id).all()
    
    statuses = []
    for round in rounds:
        score = db.query(Score).filter(Score.participant_id == current_user.id, Score.round_id == round.id).first()
        
        if score:
            if not score.is_present:
                status = "Absent"
            elif current_user.status == ParticipantStatus.ELIMINATED:
                status = "Eliminated"
            else:
                status = "Active"
            is_present = score.is_present
        else:
            status = "Pending"
            is_present = None
        
        statuses.append(ParticipantRoundStatus(
            round_no=round.round_no,
            round_name=round.name,
            status=status,
            is_present=is_present
        ))
    
    return statuses


# ==================== PUBLIC ROUNDS ====================
@api_router.get("/rounds/public", response_model=List[RoundPublicResponse])
async def get_public_rounds(db: Session = Depends(get_db)):
    rounds = db.query(Round).filter(Round.state != RoundState.DRAFT).order_by(Round.id).all()
    return [RoundPublicResponse.model_validate(r) for r in rounds]


@api_router.get("/top-referrers", response_model=List[TopReferrer])
async def get_top_referrers(db: Session = Depends(get_db)):
    top_users = db.query(User).filter(
        User.role == UserRole.PARTICIPANT,
        User.referral_count > 0
    ).order_by(User.referral_count.desc()).limit(3).all()
    
    return [TopReferrer(
        name=u.name,
        department=DepartmentEnum[u.department.name],
        referral_count=u.referral_count
    ) for u in top_users]


@api_router.get("/registration-status")
async def get_registration_status(db: Session = Depends(get_db)):
    config = db.query(SystemConfig).filter(SystemConfig.key == "registration_open").first()
    return {"registration_open": config.value == "true" if config else True}


# ==================== PDA PUBLIC ROUTES ====================
@api_router.get("/pda/programs", response_model=List[ProgramResponse])
async def get_pda_programs(db: Session = Depends(get_db)):
    programs = db.query(Program).order_by(Program.created_at.desc()).all()
    return [ProgramResponse.model_validate(p) for p in programs]


@api_router.get("/pda/events", response_model=List[EventResponse])
async def get_pda_events(db: Session = Depends(get_db)):
    events = db.query(Event).order_by(Event.start_date.is_(None), Event.start_date.asc(), Event.created_at.desc()).all()
    return [EventResponse.model_validate(e) for e in events]


@api_router.get("/pda/featured-event", response_model=EventResponse)
async def get_featured_event(db: Session = Depends(get_db)):
    event = db.query(Event).filter(Event.is_featured == True).order_by(Event.updated_at.desc()).first()
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No featured event")
    return EventResponse.model_validate(event)


# ==================== PDA ADMIN ROUTES ====================
@api_router.post("/pda-admin/programs", response_model=ProgramResponse)
async def create_pda_program(
    program_data: ProgramCreate,
    _: bool = Depends(require_pda_admin),
    db: Session = Depends(get_db)
):
    new_program = Program(
        title=program_data.title,
        description=program_data.description,
        tag=program_data.tag,
        poster_url=program_data.poster_url
    )
    db.add(new_program)
    db.commit()
    db.refresh(new_program)
    return ProgramResponse.model_validate(new_program)


@api_router.put("/pda-admin/programs/{program_id}", response_model=ProgramResponse)
async def update_pda_program(
    program_id: int,
    program_data: ProgramUpdate,
    _: bool = Depends(require_pda_admin),
    db: Session = Depends(get_db)
):
    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found")

    if program_data.title is not None:
        program.title = program_data.title
    if program_data.description is not None:
        program.description = program_data.description
    if program_data.tag is not None:
        program.tag = program_data.tag
    if program_data.poster_url is not None:
        program.poster_url = program_data.poster_url

    db.commit()
    db.refresh(program)
    return ProgramResponse.model_validate(program)


@api_router.delete("/pda-admin/programs/{program_id}")
async def delete_pda_program(
    program_id: int,
    _: bool = Depends(require_pda_admin),
    db: Session = Depends(get_db)
):
    program = db.query(Program).filter(Program.id == program_id).first()
    if not program:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Program not found")
    db.delete(program)
    db.commit()
    return {"message": "Program deleted successfully"}


@api_router.post("/pda-admin/events", response_model=EventResponse)
async def create_pda_event(
    event_data: EventCreate,
    _: bool = Depends(require_pda_admin),
    db: Session = Depends(get_db)
):
    if event_data.is_featured:
        db.query(Event).filter(Event.is_featured == True).update({"is_featured": False}, synchronize_session=False)

    new_event = Event(
        title=event_data.title,
        start_date=event_data.start_date,
        end_date=event_data.end_date,
        format=event_data.format,
        description=event_data.description,
        poster_url=event_data.poster_url,
        hero_caption=event_data.hero_caption,
        hero_url=event_data.hero_url,
        is_featured=event_data.is_featured
    )
    db.add(new_event)
    db.commit()
    db.refresh(new_event)
    return EventResponse.model_validate(new_event)


@api_router.put("/pda-admin/events/{event_id}", response_model=EventResponse)
async def update_pda_event(
    event_id: int,
    event_data: EventUpdate,
    _: bool = Depends(require_pda_admin),
    db: Session = Depends(get_db)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Event not found")

    if event_data.is_featured is True:
        db.query(Event).filter(Event.is_featured == True, Event.id != event_id).update({"is_featured": False}, synchronize_session=False)

    if event_data.title is not None:
        event.title = event_data.title
    if event_data.start_date is not None:
        event.start_date = event_data.start_date
    if event_data.end_date is not None:
        event.end_date = event_data.end_date
    if event_data.format is not None:
        event.format = event_data.format
    if event_data.description is not None:
        event.description = event_data.description
    if event_data.poster_url is not None:
        event.poster_url = event_data.poster_url
    if event_data.hero_caption is not None:
        event.hero_caption = event_data.hero_caption
    if event_data.hero_url is not None:
        event.hero_url = event_data.hero_url
    if event_data.is_featured is not None:
        event.is_featured = event_data.is_featured

    db.commit()
    db.refresh(event)
    return EventResponse.model_validate(event)


@api_router.post("/pda-admin/events/{event_id}/feature", response_model=EventResponse)
async def feature_pda_event(
    event_id: int,
    _: bool = Depends(require_pda_admin),
    db: Session = Depends(get_db)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Event not found")

    db.query(Event).filter(Event.is_featured == True, Event.id != event_id).update({"is_featured": False}, synchronize_session=False)
    event.is_featured = True
    db.commit()
    db.refresh(event)
    return EventResponse.model_validate(event)


@api_router.delete("/pda-admin/events/{event_id}")
async def delete_pda_event(
    event_id: int,
    _: bool = Depends(require_pda_admin),
    db: Session = Depends(get_db)
):
    event = db.query(Event).filter(Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Event not found")
    db.delete(event)
    db.commit()
    return {"message": "Event deleted successfully"}


@api_router.post("/pda-admin/posters")
async def upload_pda_poster(
    file: UploadFile = File(...),
    _: bool = Depends(require_pda_admin)
):
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Missing filename")
    allowed_types = ["image/png", "image/jpeg", "image/webp"]
    url = _upload_to_s3(file, "posters", allowed_types=allowed_types)
    return {"url": url}


# ==================== ADMIN ROUTES ====================
@api_router.get("/admin/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    total = db.query(User).filter(User.role == UserRole.PARTICIPANT).count()
    active = db.query(User).filter(User.role == UserRole.PARTICIPANT, User.status == ParticipantStatus.ACTIVE).count()
    eliminated = total - active
    
    reg_config = db.query(SystemConfig).filter(SystemConfig.key == "registration_open").first()
    registration_open = reg_config.value == "true" if reg_config else True
    
    rounds_completed = db.query(Round).filter(Round.state == RoundState.COMPLETED).count()
    active_round = db.query(Round).filter(Round.state == RoundState.ACTIVE).first()
    
    # Gender distribution
    gender_dist = db.query(User.gender, func.count(User.id)).filter(
        User.role == UserRole.PARTICIPANT
    ).group_by(User.gender).all()
    gender_distribution = {g.value: c for g, c in gender_dist}
    
    # Department distribution
    dept_dist = db.query(User.department, func.count(User.id)).filter(
        User.role == UserRole.PARTICIPANT
    ).group_by(User.department).all()
    department_distribution = {d.value: c for d, c in dept_dist}
    
    # Year distribution
    year_dist = db.query(User.year_of_study, func.count(User.id)).filter(
        User.role == UserRole.PARTICIPANT
    ).group_by(User.year_of_study).all()
    year_distribution = {y.value: c for y, c in year_dist}
    
    return DashboardStats(
        total_participants=total,
        registration_open=registration_open,
        rounds_completed=rounds_completed,
        current_active_round=active_round.round_no if active_round else None,
        active_count=active,
        eliminated_count=eliminated,
        gender_distribution=gender_distribution,
        department_distribution=department_distribution,
        year_distribution=year_distribution
    )


@api_router.post("/admin/toggle-registration")
async def toggle_registration(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    config = db.query(SystemConfig).filter(SystemConfig.key == "registration_open").first()
    if config:
        config.value = "false" if config.value == "true" else "true"
    else:
        db.add(SystemConfig(key="registration_open", value="false"))
    db.commit()
    
    return {"registration_open": config.value == "true" if config else False}


# ==================== ADMIN PARTICIPANT MANAGEMENT ====================
@api_router.get("/admin/participants", response_model=List[ParticipantListResponse])
async def get_participants(
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    gender: Optional[GenderEnum] = None,
    status: Optional[ParticipantStatusEnum] = None,
    search: Optional[str] = None,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    query = db.query(User).filter(User.role == UserRole.PARTICIPANT)
    
    if department:
        query = query.filter(User.department == Department[department.name])
    if year:
        query = query.filter(User.year_of_study == YearOfStudy[year.name])
    if gender:
        query = query.filter(User.gender == Gender[gender.name])
    if status:
        query = query.filter(User.status == ParticipantStatus[status.name])
    if search:
        query = query.filter(
            (User.name.ilike(f"%{search}%")) | 
            (User.register_number.ilike(f"%{search}%")) |
            (User.email.ilike(f"%{search}%"))
        )
    
    participants = query.order_by(User.name).all()
    return [ParticipantListResponse.model_validate(p) for p in participants]


@api_router.put("/admin/participants/{participant_id}/status")
async def update_participant_status(
    participant_id: int,
    new_status: ParticipantStatusEnum,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    participant = db.query(User).filter(User.id == participant_id, User.role == UserRole.PARTICIPANT).first()
    if not participant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Participant not found")
    
    participant.status = ParticipantStatus[new_status.name]
    db.commit()
    
    return {"message": "Status updated successfully"}


@api_router.get("/admin/participants/{participant_id}/rounds", response_model=List[AdminParticipantRoundStat])
async def get_participant_round_stats(
    participant_id: int,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    participant = db.query(User).filter(User.id == participant_id, User.role == UserRole.PARTICIPANT).first()
    if not participant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Participant not found")

    rounds = db.query(Round).order_by(Round.id).all()
    stats = []
    for round in rounds:
        score = db.query(Score).filter(Score.participant_id == participant_id, Score.round_id == round.id).first()
        if score:
            if not score.is_present:
                status_label = "Absent"
            elif participant.status == ParticipantStatus.ELIMINATED:
                status_label = "Eliminated"
            else:
                status_label = "Active"
            is_present = score.is_present
            total_score = score.total_score
            normalized_score = score.normalized_score
        else:
            status_label = "Pending"
            is_present = None
            total_score = None
            normalized_score = None

        stats.append(AdminParticipantRoundStat(
            round_id=round.id,
            round_no=round.round_no,
            round_name=round.name,
            round_state=RoundStateEnum[round.state.name],
            status=status_label,
            is_present=is_present,
            total_score=total_score,
            normalized_score=normalized_score
        ))

    return stats


# ==================== ADMIN ROUND MANAGEMENT ====================
@api_router.get("/admin/rounds", response_model=List[RoundResponse])
async def get_all_rounds(admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    rounds = db.query(Round).order_by(Round.id).all()
    return [RoundResponse.model_validate(r) for r in rounds]


@api_router.post("/admin/rounds", response_model=RoundResponse)
async def create_round(round_data: RoundCreate, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    # Generate round number
    last_round = db.query(Round).order_by(Round.id.desc()).first()
    round_num = 1 if not last_round else int(last_round.round_no[2:]) + 1
    round_no = f"PF{round_num:02d}"
    
    criteria = [c.model_dump() for c in round_data.evaluation_criteria] if round_data.evaluation_criteria else None
    if not criteria:
        criteria = [{"name": "Overall", "max_marks": 100}]

    new_round = Round(
        round_no=round_no,
        name=round_data.name,
        description=round_data.description,
        tags=round_data.tags,
        date=round_data.date,
        mode=RoundMode[round_data.mode.name],
        conducted_by=round_data.conducted_by,
        state=RoundState.DRAFT,
        evaluation_criteria=criteria
    )
    
    db.add(new_round)
    db.commit()
    db.refresh(new_round)
    
    return RoundResponse.model_validate(new_round)


@api_router.put("/admin/rounds/{round_id}", response_model=RoundResponse)
async def update_round(
    round_id: int,
    round_data: RoundUpdate,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    if round.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is frozen and cannot be edited")
    
    if round_data.name is not None:
        round.name = round_data.name
    if round_data.description is not None:
        round.description = round_data.description
    if round_data.tags is not None:
        round.tags = round_data.tags
    if round_data.date is not None:
        round.date = round_data.date
    if round_data.mode is not None:
        round.mode = RoundMode[round_data.mode.name]
    if round_data.conducted_by is not None:
        round.conducted_by = round_data.conducted_by
    if round_data.state is not None:
        round.state = RoundState[round_data.state.name]
    if round_data.evaluation_criteria is not None:
        criteria = [c.model_dump() for c in round_data.evaluation_criteria] if round_data.evaluation_criteria else []
        if not criteria:
            criteria = [{"name": "Overall", "max_marks": 100}]
        round.evaluation_criteria = criteria
    if round_data.elimination_type is not None:
        round.elimination_type = round_data.elimination_type
    if round_data.elimination_value is not None:
        round.elimination_value = round_data.elimination_value
    
    db.commit()
    db.refresh(round)
    
    return RoundResponse.model_validate(round)

@api_router.post("/admin/rounds/{round_id}/description-pdf", response_model=RoundResponse)
async def upload_round_description_pdf(
    round_id: int,
    file: UploadFile = File(...),
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only PDF files are allowed")

    s3_url = _upload_to_s3(file, "persofest/rounds", allowed_types=["application/pdf"])
    round.description_pdf = s3_url
    db.commit()
    db.refresh(round)

    return RoundResponse.model_validate(round)


@api_router.delete("/admin/rounds/{round_id}")
async def delete_round(round_id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    if round.state != RoundState.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Can only delete draft rounds")
    
    db.delete(round)
    db.commit()
    
    return {"message": "Round deleted successfully"}


# ==================== ADMIN SCORE MANAGEMENT ====================
@api_router.get("/admin/rounds/{round_id}/participants")
async def get_round_participants(
    round_id: int,
    search: Optional[str] = None,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    # Get active participants (not eliminated before this round)
    query = db.query(User).filter(User.role == UserRole.PARTICIPANT)
    
    if search:
        query = query.filter(
            (User.name.ilike(f"%{search}%")) | 
            (User.register_number.ilike(f"%{search}%"))
        )
    
    participants = query.order_by(User.name).all()
    
    result = []
    for p in participants:
        score = db.query(Score).filter(Score.participant_id == p.id, Score.round_id == round_id).first()
        result.append({
            "id": p.id,
            "register_number": p.register_number,
            "name": p.name,
            "department": p.department.value,
            "status": p.status.value,
            "is_present": score.is_present if score else False,
            "criteria_scores": score.criteria_scores if score else None,
            "total_score": score.total_score if score else 0,
            "normalized_score": score.normalized_score if score else 0
        })
    
    return result


@api_router.post("/admin/rounds/{round_id}/scores")
async def enter_scores(
    round_id: int,
    scores: List[ScoreEntry],
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    if round.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is frozen")
    
    # Calculate max possible score
    max_score = sum(c["max_marks"] for c in round.evaluation_criteria) if round.evaluation_criteria else 100
    
    for entry in scores:
        existing = db.query(Score).filter(
            Score.participant_id == entry.participant_id,
            Score.round_id == round_id
        ).first()
        
        total = sum(entry.criteria_scores.values()) if entry.criteria_scores else 0
        normalized = (total / max_score * 100) if max_score > 0 else 0
        
        if existing:
            existing.criteria_scores = entry.criteria_scores
            existing.total_score = total
            existing.normalized_score = normalized
            existing.is_present = entry.is_present
        else:
            new_score = Score(
                participant_id=entry.participant_id,
                round_id=round_id,
                criteria_scores=entry.criteria_scores,
                total_score=total,
                normalized_score=normalized,
                is_present=entry.is_present
            )
            db.add(new_score)
    
    db.commit()
    return {"message": "Scores saved successfully"}


@api_router.post("/admin/rounds/{round_id}/import-scores")
async def import_scores_from_excel(
    round_id: int,
    file: UploadFile = File(...),
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """
    Import scores from Excel file.
    Expected columns: Register Number, Present (Yes/No), [Criteria columns...]
    """
    from openpyxl import load_workbook
    
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    if round.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is frozen")
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only Excel files (.xlsx, .xls) are allowed")
    
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        if not headers or 'Register Number' not in headers:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel must have 'Register Number' column")
        
        # Get criteria names from round
        criteria_names = [c["name"] for c in round.evaluation_criteria] if round.evaluation_criteria else []
        max_score = sum(c["max_marks"] for c in round.evaluation_criteria) if round.evaluation_criteria else 100
        
        imported_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            
            row_dict = dict(zip(headers, row))
            register_number = str(row_dict.get('Register Number', '')).strip()
            
            if not register_number:
                continue
            
            # Find participant
            participant = db.query(User).filter(
                User.register_number == register_number,
                User.role == UserRole.PARTICIPANT
            ).first()
            
            if not participant:
                errors.append(f"Row {row_idx}: Register number {register_number} not found")
                continue
            
            # Parse presence
            present_val = str(row_dict.get('Present', 'Yes')).strip().lower()
            is_present = present_val in ('yes', 'y', '1', 'true', 'present')
            
            # Parse criteria scores
            criteria_scores = {}
            for cname in criteria_names:
                if cname in row_dict and row_dict[cname] is not None:
                    try:
                        criteria_scores[cname] = float(row_dict[cname])
                    except (ValueError, TypeError):
                        criteria_scores[cname] = 0
                else:
                    criteria_scores[cname] = 0
            
            # Calculate totals
            total = sum(criteria_scores.values())
            normalized = (total / max_score * 100) if max_score > 0 else 0
            
            # Update or create score
            existing = db.query(Score).filter(
                Score.participant_id == participant.id,
                Score.round_id == round_id
            ).first()
            
            if existing:
                existing.criteria_scores = criteria_scores
                existing.total_score = total
                existing.normalized_score = normalized
                existing.is_present = is_present
            else:
                new_score = Score(
                    participant_id=participant.id,
                    round_id=round_id,
                    criteria_scores=criteria_scores,
                    total_score=total,
                    normalized_score=normalized,
                    is_present=is_present
                )
                db.add(new_score)
            
            imported_count += 1
        
        db.commit()
        
        return {
            "message": f"Successfully imported {imported_count} scores",
            "imported": imported_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        logger.error(f"Excel import error: {str(e)}")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse Excel: {str(e)}")


@api_router.get("/admin/rounds/{round_id}/score-template")
async def download_score_template(
    round_id: int,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Download Excel template for bulk score import"""
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    # Get active participants
    participants = db.query(User).filter(
        User.role == UserRole.PARTICIPANT,
        User.status == ParticipantStatus.ACTIVE
    ).order_by(User.name).all()
    
    criteria_names = [c["name"] for c in round.evaluation_criteria] if round.evaluation_criteria else ["Score"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{round.round_no} Scores"
    
    # Headers
    headers = ["Register Number", "Name", "Present"] + criteria_names
    ws.append(headers)
    
    # Add participant rows
    for p in participants:
        row = [p.register_number, p.name, "Yes"] + [0] * len(criteria_names)
        ws.append(row)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={round.round_no}_score_template.xlsx"}
    )


@api_router.post("/admin/rounds/{round_id}/freeze")
async def freeze_round(round_id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    if not round.elimination_type or round.elimination_value is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Set elimination rules before freezing")
    
    # Apply elimination
    scores = db.query(Score).filter(Score.round_id == round_id).all()
    
    if round.elimination_type == "top_k":
        # Sort by score and keep top K
        sorted_scores = sorted(scores, key=lambda s: s.normalized_score, reverse=True)
        k = int(round.elimination_value)
        
        for i, score in enumerate(sorted_scores):
            participant = db.query(User).filter(User.id == score.participant_id).first()
            if not score.is_present or i >= k:
                participant.status = ParticipantStatus.ELIMINATED
    else:  # min_score
        min_score = round.elimination_value
        for score in scores:
            participant = db.query(User).filter(User.id == score.participant_id).first()
            if not score.is_present or score.normalized_score < min_score:
                participant.status = ParticipantStatus.ELIMINATED
    
    # Also eliminate participants who don't have scores (absent)
    scored_ids = [s.participant_id for s in scores]
    absent_participants = db.query(User).filter(
        User.role == UserRole.PARTICIPANT,
        User.status == ParticipantStatus.ACTIVE,
        ~User.id.in_(scored_ids)
    ).all()
    
    for p in absent_participants:
        p.status = ParticipantStatus.ELIMINATED
    
    round.is_frozen = True
    round.state = RoundState.COMPLETED
    db.commit()
    
    return {"message": "Round frozen and eliminations applied"}

@api_router.post("/admin/rounds/{round_id}/unfreeze")
async def unfreeze_round(round_id: int, admin: User = Depends(get_current_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    if not round.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is not frozen")

    round.is_frozen = False
    round.state = RoundState.ACTIVE
    db.commit()
    return {"message": "Round unfrozen"}


# ==================== ADMIN LEADERBOARD ====================
@api_router.get("/admin/leaderboard", response_model=List[LeaderboardEntry])
async def get_leaderboard(
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    search: Optional[str] = None,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    # Get all participants
    query = db.query(User).filter(User.role == UserRole.PARTICIPANT)
    
    if department:
        query = query.filter(User.department == Department[department.name])
    if year:
        query = query.filter(User.year_of_study == YearOfStudy[year.name])
    if search:
        query = query.filter(
            (User.name.ilike(f"%{search}%")) | 
            (User.register_number.ilike(f"%{search}%"))
        )
    
    participants = query.all()
    
    # Calculate cumulative scores from frozen rounds only
    frozen_rounds = db.query(Round).filter(Round.is_frozen == True).all()
    frozen_round_ids = [r.id for r in frozen_rounds]
    
    leaderboard = []
    for p in participants:
        scores = db.query(Score).filter(
            Score.participant_id == p.id,
            Score.round_id.in_(frozen_round_ids),
            Score.is_present == True
        ).all()
        
        cumulative = sum(s.normalized_score for s in scores)
        rounds_participated = len(scores)
        
        leaderboard.append({
            "participant_id": p.id,
            "register_number": p.register_number,
            "name": p.name,
            "department": p.department,
            "year_of_study": p.year_of_study,
            "cumulative_score": cumulative,
            "rounds_participated": rounds_participated,
            "status": p.status,
            "profile_picture": p.profile_picture
        })
    
    # Sort by cumulative score
    leaderboard.sort(key=lambda x: x["cumulative_score"], reverse=True)
    
    # Add rank
    result = []
    for i, entry in enumerate(leaderboard):
        result.append(LeaderboardEntry(
            rank=i + 1,
            participant_id=entry["participant_id"],
            register_number=entry["register_number"],
            name=entry["name"],
            department=DepartmentEnum[entry["department"].name],
            year_of_study=YearOfStudyEnum[entry["year_of_study"].name],
            cumulative_score=entry["cumulative_score"],
            rounds_participated=entry["rounds_participated"],
            status=ParticipantStatusEnum[entry["status"].name],
            profile_picture=entry.get("profile_picture")
        ))
    
    return result


# ==================== ADMIN EXPORT ====================
@api_router.get("/admin/export/participants")
async def export_participants(
    format: str = Query("csv", enum=["csv", "xlsx"]),
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    status: Optional[ParticipantStatusEnum] = None,
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    query = db.query(User).filter(User.role == UserRole.PARTICIPANT)
    
    if department:
        query = query.filter(User.department == Department[department.name])
    if year:
        query = query.filter(User.year_of_study == YearOfStudy[year.name])
    if status:
        query = query.filter(User.status == ParticipantStatus[status.name])
    
    participants = query.order_by(User.name).all()
    
    headers = ["Register Number", "Name", "Email", "Phone", "Gender", "Department", "Year", "Status", "Referral Code", "Referrals"]
    rows = [[
        p.register_number, p.name, p.email, p.phone, p.gender.value,
        p.department.value, p.year_of_study.value, p.status.value,
        p.referral_code, p.referral_count
    ] for p in participants]
    
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=participants.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=participants.csv"}
        )


@api_router.get("/admin/export/round/{round_id}")
async def export_round_results(
    round_id: int,
    format: str = Query("csv", enum=["csv", "xlsx"]),
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    
    scores = db.query(Score).filter(Score.round_id == round_id).all()
    
    criteria_names = [c["name"] for c in round.evaluation_criteria] if round.evaluation_criteria else []
    headers = ["Register Number", "Name", "Present"] + criteria_names + ["Total Score", "Normalized Score"]
    
    rows = []
    for s in scores:
        participant = db.query(User).filter(User.id == s.participant_id).first()
        row = [participant.register_number, participant.name, "Yes" if s.is_present else "No"]
        for cn in criteria_names:
            row.append(s.criteria_scores.get(cn, 0) if s.criteria_scores else 0)
        row.extend([s.total_score, s.normalized_score])
        rows.append(row)
    
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = round.round_no
        ws.append(headers)
        for row in rows:
            ws.append(row)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={round.round_no}_results.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={round.round_no}_results.csv"}
        )


@api_router.get("/admin/export/leaderboard")
async def export_leaderboard(
    format: str = Query("csv", enum=["csv", "xlsx"]),
    admin: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    # Get leaderboard data
    participants = db.query(User).filter(User.role == UserRole.PARTICIPANT).all()
    frozen_rounds = db.query(Round).filter(Round.is_frozen == True).all()
    frozen_round_ids = [r.id for r in frozen_rounds]
    
    leaderboard = []
    for p in participants:
        scores = db.query(Score).filter(
            Score.participant_id == p.id,
            Score.round_id.in_(frozen_round_ids),
            Score.is_present == True
        ).all()
        
        cumulative = sum(s.normalized_score for s in scores)
        rounds_participated = len(scores)
        
        leaderboard.append({
            "register_number": p.register_number,
            "name": p.name,
            "department": p.department.value,
            "year": p.year_of_study.value,
            "status": p.status.value,
            "cumulative_score": cumulative,
            "rounds_participated": rounds_participated
        })
    
    leaderboard.sort(key=lambda x: x["cumulative_score"], reverse=True)
    
    headers = ["Rank", "Register Number", "Name", "Department", "Year", "Status", "Cumulative Score", "Rounds Participated"]
    rows = [[i+1, e["register_number"], e["name"], e["department"], e["year"], e["status"], e["cumulative_score"], e["rounds_participated"]] for i, e in enumerate(leaderboard)]
    
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Leaderboard"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=leaderboard.xlsx"}
        )
    else:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=leaderboard.csv"}
        )


# Include router and add middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
