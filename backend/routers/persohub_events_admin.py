import csv
import io
import math
import re
import base64
import hashlib
import ssl
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.request import Request as UrlRequest, urlopen

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Query, Request, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from auth import decode_token
from database import get_db, SessionLocal
from models import (
    PdaAdmin,
    PdaUser,
    PdaItem,
    PersohubEvent,
    PersohubEventType,
    PersohubEventFormat,
    PersohubEventTemplate,
    PersohubEventStatus,
    PersohubEventParticipantMode,
    PersohubEventRoundMode,
    PersohubEventEntityType,
    PersohubEventRegistration,
    PersohubEventRegistrationStatus,
    PersohubEventTeam,
    PersohubEventTeamMember,
    PdaTeam,
    PersohubEventRound,
    PersohubEventRoundState,
    PersohubEventRoundPanel,
    PersohubEventRoundPanelMember,
    PersohubEventRoundPanelAssignment,
    PersohubEventAttendance,
    PersohubEventScore,
    PersohubEventRoundSubmission,
    PersohubEventBadge,
    PersohubEventBadgePlace,
    PersohubEventInvite,
    PersohubEventLog,
    PersohubClub,
    PersohubCommunity,
)
from schemas import (
    PersohubManagedAttendanceMarkRequest,
    PersohubManagedAttendanceScanRequest,
    PersohubManagedBadgeCreate,
    PersohubManagedBadgeResponse,
    PersohubManagedEntityTypeEnum,
    PersohubManagedEventCreate,
    PersohubManagedEventRegistrationUpdate,
    PersohubManagedEventResponse,
    PersohubManagedEventStatusUpdate,
    PersohubManagedEventVisibilityUpdate,
    PersohubManagedEventUpdate,
    PersohubManagedRoundCreate,
    PersohubEventLogResponse,
    PersohubManagedRoundResponse,
    PersohubManagedRoundUpdate,
    PersohubManagedRegistrationStatusBulkRequest,
    PersohubManagedRegistrationStatusBulkResponse,
    PersohubRoundSubmissionAdminListItem,
    PersohubRoundSubmissionAdminUpdate,
    PersohubManagedScoreEntry,
    PersohubManagedTeamResponse,
    EventBulkEmailRequest,
    PersohubRoundPanelListResponse,
    PersohubRoundPanelsUpdateRequest,
    PersohubRoundPanelsAutoAssignRequest,
    PersohubRoundPanelAssignmentsUpdateRequest,
    PersohubRoundPanelEmailRequest,
    PersohubRoundPanelResponse,
    PersohubRoundPanelMemberResponse,
    PersohubRoundPanelAdminOption,
)
from emailer import send_bulk_email
from email_bulk import render_email_template, derive_text_from_html, extract_batch
from security import (
    get_persohub_admin_context,
    require_persohub_event_admin,
    require_persohub_events_parity_enabled,
    require_persohub_root_community_admin,
    require_persohub_community,
)
from utils import log_admin_action, log_persohub_event_action, _upload_bytes_to_s3

router = APIRouter(dependencies=[Depends(require_persohub_events_parity_enabled)])
OFFICIAL_LETTERHEAD_LEFT_LOGO_URL = ""


def _slugify(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower()).strip("-")
    return cleaned[:110] if cleaned else "event"


@router.get("/persohub/admin/persohub-events/parity-enabled")
def get_persohub_events_parity_enabled(
    _: PersohubCommunity = Depends(require_persohub_community),
):
    return {"enabled": True}


def _next_slug(db: Session, title: str) -> str:
    base = _slugify(title)
    slug = base
    counter = 2
    while db.query(PersohubEvent).filter(PersohubEvent.slug == slug).first():
        slug = f"{base}-{counter}"
        counter += 1
    return slug


def _next_event_code(db: Session) -> str:
    latest = db.query(PersohubEvent).order_by(PersohubEvent.id.desc()).first()
    next_id = (latest.id + 1) if latest else 1
    return f"EVT{next_id:03d}"


def _get_event_or_404(db: Session, slug: str) -> PersohubEvent:
    event = db.query(PersohubEvent).filter(PersohubEvent.slug == slug).first()
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Event not found")
    return event


def _get_event_round_or_404(db: Session, event_id: int, round_id: int) -> PersohubEventRound:
    round_row = db.query(PersohubEventRound).filter(
        PersohubEventRound.id == round_id,
        PersohubEventRound.event_id == event_id,
    ).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    return round_row


def _ensure_events_policy_shape(policy: Optional[dict]) -> dict:
    safe = dict(policy or {})
    if not isinstance(safe.get("events"), dict):
        safe["events"] = {}
    return safe


def _criteria_def(round_obj: PersohubEventRound) -> List[dict]:
    return round_obj.evaluation_criteria or [{"name": "Score", "max_marks": 100}]


def _default_round_allowed_mime_types() -> List[str]:
    return [
        "application/pdf",
        "application/vnd.ms-powerpoint",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        "image/png",
        "image/jpeg",
        "image/webp",
        "video/mp4",
        "video/quicktime",
        "application/zip",
    ]


PANEL_TEAM_DISTRIBUTION_MODES = {"team_count", "member_count_weighted"}


def _normalize_panel_distribution_mode(value) -> str:
    raw = str(value or "team_count").strip().lower()
    return raw if raw in PANEL_TEAM_DISTRIBUTION_MODES else "team_count"


def _is_superadmin_user(db: Session, user: PdaUser) -> bool:
    admin_row = db.query(PdaAdmin).filter(PdaAdmin.user_id == user.id).first()
    policy = admin_row.policy if admin_row and isinstance(admin_row.policy, dict) else {}
    return bool(admin_row and policy.get("superAdmin"))


def _panel_entity_key(entity_type: str, entity_id: int) -> Tuple[str, int]:
    return (str(entity_type or "").strip().lower(), int(entity_id))


def _round_panel_maps(db: Session, round_row: PersohubEventRound) -> Tuple[Dict[int, PersohubEventRoundPanel], Dict[Tuple[str, int], PersohubEventRoundPanelAssignment]]:
    panels = db.query(PersohubEventRoundPanel).filter(
        PersohubEventRoundPanel.event_id == round_row.event_id,
        PersohubEventRoundPanel.round_id == round_row.id,
    ).all()
    panel_map = {int(panel.id): panel for panel in panels}
    assignments = db.query(PersohubEventRoundPanelAssignment).filter(
        PersohubEventRoundPanelAssignment.event_id == round_row.event_id,
        PersohubEventRoundPanelAssignment.round_id == round_row.id,
    ).all()
    assignment_map: Dict[Tuple[str, int], PersohubEventRoundPanelAssignment] = {}
    for assignment in assignments:
        if assignment.entity_type == PersohubEventEntityType.USER and assignment.user_id is not None:
            assignment_map[_panel_entity_key("user", int(assignment.user_id))] = assignment
        if assignment.entity_type == PersohubEventEntityType.TEAM and assignment.team_id is not None:
            assignment_map[_panel_entity_key("team", int(assignment.team_id))] = assignment
    return panel_map, assignment_map


def _round_admin_panel_scope(
    db: Session,
    round_row: PersohubEventRound,
    admin: PdaUser,
) -> Dict[str, object]:
    panel_mode_enabled = bool(round_row.panel_mode_enabled)
    is_superadmin = _is_superadmin_user(db, admin)
    panel_ids: Set[int] = set()
    if panel_mode_enabled and not is_superadmin:
        panel_ids = {
            int(row.panel_id)
            for row in db.query(PersohubEventRoundPanelMember.panel_id).filter(
                PersohubEventRoundPanelMember.event_id == round_row.event_id,
                PersohubEventRoundPanelMember.round_id == round_row.id,
                PersohubEventRoundPanelMember.admin_user_id == admin.id,
            ).all()
            if row.panel_id is not None
        }
    allowed_entities: Optional[Set[Tuple[str, int]]] = None
    if panel_mode_enabled and not is_superadmin:
        assignments = db.query(PersohubEventRoundPanelAssignment).filter(
            PersohubEventRoundPanelAssignment.event_id == round_row.event_id,
            PersohubEventRoundPanelAssignment.round_id == round_row.id,
        )
        if panel_ids:
            assignments = assignments.filter(PersohubEventRoundPanelAssignment.panel_id.in_(panel_ids))
        else:
            assignments = assignments.filter(text("1=0"))
        allowed_entities = set()
        for assignment in assignments.all():
            if assignment.entity_type == PersohubEventEntityType.USER and assignment.user_id is not None:
                allowed_entities.add(_panel_entity_key("user", int(assignment.user_id)))
            elif assignment.entity_type == PersohubEventEntityType.TEAM and assignment.team_id is not None:
                allowed_entities.add(_panel_entity_key("team", int(assignment.team_id)))
    return {
        "panel_mode_enabled": panel_mode_enabled,
        "is_superadmin": is_superadmin,
        "panel_ids": panel_ids,
        "allowed_entities": allowed_entities,
    }


def _is_entity_editable_by_admin(scope: Dict[str, object], entity_type: str, entity_id: int, panel_id: Optional[int]) -> bool:
    return True


def _legacy_normalized(total_score: float, max_total: float, is_present: bool) -> float:
    if not bool(is_present):
        return 0.0
    max_total_value = float(max_total or 0.0)
    if max_total_value <= 0.0:
        return 0.0
    normalized = (float(total_score or 0.0) / max_total_value) * 100.0
    if normalized < 0.0:
        return 0.0
    if normalized > 100.0:
        return 100.0
    return float(normalized)


def _recompute_round_normalized_scores(
    db: Session,
    event: PersohubEvent,
    round_row: PersohubEventRound,
) -> None:
    # SessionLocal uses autoflush=False; flush pending score/attendance/assignment writes
    # so recompute always reads the latest in-transaction state.
    db.flush()
    criteria = _criteria_def(round_row)
    max_total = float(sum(float(item.get("max_marks", 0) or 0.0) for item in criteria) or 0.0)
    score_rows = (
        db.query(PersohubEventScore)
        .filter(
            PersohubEventScore.event_id == event.id,
            PersohubEventScore.round_id == round_row.id,
        )
        .all()
    )
    if not score_rows:
        return

    for score_row in score_rows:
        score_row.normalized_score = _legacy_normalized(
            total_score=float(score_row.total_score or 0.0),
            max_total=max_total,
            is_present=bool(score_row.is_present),
        )


def _round_submission_deadline_has_passed(round_row: PersohubEventRound) -> bool:
    if not round_row.submission_deadline:
        return False
    now = datetime.now(timezone.utc)
    deadline = round_row.submission_deadline
    if deadline.tzinfo is None:
        deadline = deadline.replace(tzinfo=timezone.utc)
    return now >= deadline


def _round_submission_payload_for_admin(
    round_row: PersohubEventRound,
    event: PersohubEvent,
    entity: dict,
    submission: Optional[PersohubEventRoundSubmission],
) -> PersohubRoundSubmissionAdminListItem:
    deadline_passed = _round_submission_deadline_has_passed(round_row)
    lock_reason = None
    if submission and submission.is_locked:
        lock_reason = "Submission is locked by admin"
    elif deadline_passed:
        lock_reason = "Submission deadline has passed"
    elif round_row.state in {PersohubEventRoundState.COMPLETED, PersohubEventRoundState.REVEAL}:
        lock_reason = "Round is finalized"
    elif round_row.is_frozen:
        lock_reason = "Round is frozen"

    return PersohubRoundSubmissionAdminListItem(
        id=submission.id if submission else None,
        event_id=event.id,
        round_id=round_row.id,
        entity_type=entity["entity_type"],
        user_id=entity["entity_id"] if str(entity["entity_type"]) == "user" else None,
        team_id=entity["entity_id"] if str(entity["entity_type"]) == "team" else None,
        participant_name=str(entity.get("name") or ""),
        participant_register_number=str(entity.get("regno_or_code") or ""),
        participant_status=str(entity.get("status") or "Active"),
        submission_type=(submission.submission_type if submission else None),
        file_url=(submission.file_url if submission else None),
        file_name=(submission.file_name if submission else None),
        file_size_bytes=(submission.file_size_bytes if submission else None),
        mime_type=(submission.mime_type if submission else None),
        link_url=(submission.link_url if submission else None),
        notes=(submission.notes if submission else None),
        version=(int(submission.version or 0) if submission else 0),
        is_locked=(bool(submission.is_locked) if submission else False),
        submitted_at=(submission.submitted_at if submission else None),
        updated_at=(submission.updated_at if submission else None),
        updated_by_user_id=(submission.updated_by_user_id if submission else None),
        is_editable=bool(submission and not submission.is_locked),
        lock_reason=lock_reason,
        deadline_at=round_row.submission_deadline,
    )


SCORE_RATIO_RE = re.compile(r"^\s*([+-]?\d+(?:\.\d+)?)\s*/\s*([+-]?\d+(?:\.\d+)?)\s*$")


def _parse_import_score_value(raw_value, max_marks: float) -> float:
    if raw_value is None:
        return 0.0
    if isinstance(raw_value, (int, float)):
        value = float(raw_value)
    else:
        text = str(raw_value).strip()
        if not text:
            return 0.0
        ratio_match = SCORE_RATIO_RE.match(text)
        if ratio_match:
            numerator = float(ratio_match.group(1))
            denominator = float(ratio_match.group(2))
            if denominator <= 0:
                raise ValueError("invalid_denominator")
            value = (numerator / denominator) * float(max_marks)
        else:
            value = float(text)
    if not math.isfinite(value):
        raise ValueError("invalid_number")
    return float(value)


def _normalize_compare_text(value) -> str:
    return str(value or "").strip().lower()


def _to_event_type(value) -> PersohubEventType:
    return PersohubEventType[value.name] if hasattr(value, "name") else PersohubEventType(value)


def _to_event_format(value) -> PersohubEventFormat:
    return PersohubEventFormat[value.name] if hasattr(value, "name") else PersohubEventFormat(value)


def _to_event_template(value) -> PersohubEventTemplate:
    return PersohubEventTemplate[value.name] if hasattr(value, "name") else PersohubEventTemplate(value)


def _to_participant_mode(value) -> PersohubEventParticipantMode:
    return PersohubEventParticipantMode[value.name] if hasattr(value, "name") else PersohubEventParticipantMode(value)


def _to_round_mode(value) -> PersohubEventRoundMode:
    return PersohubEventRoundMode[value.name] if hasattr(value, "name") else PersohubEventRoundMode(value)


def _to_event_status(value) -> PersohubEventStatus:
    return PersohubEventStatus[value.name] if hasattr(value, "name") else PersohubEventStatus(value)


def _to_event_open_for(value) -> str:
    raw = str(value.value if hasattr(value, "value") else value or "").strip().upper()
    return "ALL" if raw == "ALL" else "MIT"


def _to_round_state(value) -> PersohubEventRoundState:
    return PersohubEventRoundState[value.name] if hasattr(value, "name") else PersohubEventRoundState(value)


def _validate_event_dates(start_date, end_date) -> None:
    if start_date and end_date and start_date > end_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="start_date cannot be after end_date")


def _managed_event_home_link(slug: str) -> str:
    return f"/events/{slug}"


def _find_managed_home_item(db: Session, slug: str) -> Optional[PdaItem]:
    link = _managed_event_home_link(slug)
    return (
        db.query(PdaItem)
        .filter(
            PdaItem.type == "event",
            PdaItem.hero_url == link,
        )
        .first()
    )


def _sync_managed_event_to_home_item(db: Session, event: PersohubEvent) -> None:
    item = _find_managed_home_item(db, event.slug)
    if not item:
        item = PdaItem(type="event")
        db.add(item)

    item.title = event.title
    item.description = event.description
    item.poster_url = event.poster_url
    item.start_date = event.start_date
    item.end_date = event.end_date
    item.format = event.format.value if hasattr(event.format, "value") else str(event.format)
    item.hero_url = _managed_event_home_link(event.slug)
    item.tag = item.tag or "managed-event"


def _delete_managed_home_item(db: Session, slug: str) -> None:
    item = _find_managed_home_item(db, slug)
    if item:
        db.delete(item)


def _entity_from_payload(event: PersohubEvent, row: dict) -> Tuple[PersohubEventEntityType, Optional[int], Optional[int]]:
    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        user_id = row.get("user_id")
        if not user_id:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_id is required for individual event")
        return PersohubEventEntityType.USER, int(user_id), None
    team_id = row.get("team_id")
    if not team_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="team_id is required for team event")
    return PersohubEventEntityType.TEAM, None, int(team_id)


def _batch_from_regno(regno: str) -> Optional[str]:
    value = str(regno or "").strip()
    if len(value) < 4 or not value[:4].isdigit():
        return None
    return value[:4]


def _registration_status_label(value) -> str:
    if hasattr(value, "value"):
        return str(value.value)
    raw = str(value or "").strip().upper()
    return "Eliminated" if "ELIMINATED" in raw else "Active"


def _status_is_active(value) -> bool:
    return str(value or "").strip().lower() == "active"


def _sync_event_round_count(db: Session, event_id: int) -> None:
    round_count = (
        db.query(PersohubEventRound)
        .filter(PersohubEventRound.event_id == event_id)
        .count()
    )
    db.query(PersohubEvent).filter(PersohubEvent.id == event_id).update(
        {PersohubEvent.round_count: int(round_count)},
        synchronize_session=False,
    )


def _log_event_admin_action(
    db: Session,
    admin: PdaUser,
    event: PersohubEvent,
    action: str,
    method: str,
    path: str,
    meta: Optional[dict] = None,
):
    log_admin_action(db, admin, action, method=method, path=path, meta=meta)
    log_persohub_event_action(
        db=db,
        event_slug=event.slug,
        admin=admin,
        action=action,
        event_id=event.id,
        method=method,
        path=path,
        meta=meta,
    )


def _audit_fragment(value: object, fallback: str = "na") -> str:
    raw = str(value or "").strip().lower()
    cleaned = re.sub(r"[^a-z0-9]+", "-", raw).strip("-")
    return (cleaned[:48] if cleaned else fallback)


def _build_round_audit_csv(
    db: Session,
    event: PersohubEvent,
    round_row: PersohubEventRound,
    admin: PdaUser,
    audit_type: str,
    generated_at: datetime,
) -> Tuple[List[str], List[List[object]]]:
    result = round_participants(
        slug=event.slug,
        round_id=int(round_row.id),
        search=None,
        admin=admin,
        db=db,
    )
    criteria = _criteria_def(round_row)
    criteria_names = [str(item.get("name") or "").strip() for item in criteria if str(item.get("name") or "").strip()]
    id_header = "Register Number" if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else "Team Code"
    name_header = "Name" if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else "Team Name"
    headers = [
        "audit_type",
        "event_slug",
        "event_code",
        "round_id",
        "round_no",
        "round_name",
        "round_state",
        "is_frozen",
        "generated_at_utc",
        "entity_type",
        "entity_id",
        name_header,
        id_header,
        "participant_status",
        "is_present",
        "total_score",
        "normalized_score",
        "panel_no",
        "panel_name",
        "submission_type",
        "submission_locked",
    ] + criteria_names
    rows: List[List[object]] = []
    for row in result:
        criteria_scores = row.get("criteria_scores") if isinstance(row.get("criteria_scores"), dict) else {}
        entity_id = row.get("entity_id")
        if entity_id is None:
            entity_id = row.get("participant_id")
        line = [
            audit_type,
            event.slug,
            event.event_code,
            int(round_row.id),
            int(round_row.round_no),
            str(round_row.name or ""),
            str(round_row.state.value if hasattr(round_row.state, "value") else round_row.state or ""),
            bool(round_row.is_frozen),
            generated_at.isoformat(),
            row.get("entity_type"),
            entity_id,
            row.get("name") or row.get("participant_name"),
            row.get("regno_or_code") or row.get("participant_register_number"),
            row.get("status") or row.get("participant_status"),
            bool(row.get("is_present")),
            float(row.get("total_score") or 0.0),
            float(row.get("normalized_score") or 0.0),
            row.get("panel_no"),
            row.get("panel_name"),
            row.get("submission_type"),
            bool(row.get("submission_is_locked")) if row.get("submission_is_locked") is not None else False,
        ]
        for criteria_name in criteria_names:
            raw = criteria_scores.get(criteria_name)
            if raw is None:
                line.append(0)
                continue
            try:
                line.append(float(raw))
            except Exception:
                line.append(raw)
        rows.append(line)
    return headers, rows


def _upload_round_audit_snapshot(
    db: Session,
    event: PersohubEvent,
    round_row: PersohubEventRound,
    admin: PdaUser,
    *,
    audit_type: str,
    folder: str,
    extra_meta: Optional[dict] = None,
) -> dict:
    generated_at = datetime.now(timezone.utc)
    base_meta = {
        "audit_type": audit_type,
        "audit_folder": folder,
        "audit_generated_at": generated_at.isoformat(),
        "audit_csv_uploaded": False,
        "audit_csv_url": None,
        "audit_csv_error": None,
    }
    if extra_meta:
        for key, value in extra_meta.items():
            if key not in base_meta:
                base_meta[key] = value
    try:
        headers, rows = _build_round_audit_csv(db, event, round_row, admin, audit_type, generated_at)
        content = _export_to_csv(headers, rows)
        timestamp_text = generated_at.strftime("%Y%m%dT%H%M%SZ")
        filename = (
            f"{_audit_fragment(event.event_code, 'evt')}"
            f"_round-{int(round_row.round_no)}"
            f"_{_audit_fragment(audit_type, 'audit')}"
            f"_{timestamp_text}"
            f"_by-{_audit_fragment(getattr(admin, 'regno', None) or getattr(admin, 'id', None), 'admin')}.csv"
        )
        key_prefix = f"pda-events/{event.slug}/audits/{folder}/round-{int(round_row.round_no)}"
        url = _upload_bytes_to_s3(content, key_prefix, filename, content_type="text/csv")
        base_meta["audit_csv_uploaded"] = True
        base_meta["audit_csv_url"] = url
        base_meta["audit_csv_rows"] = len(rows)
        base_meta["audit_csv_filename"] = filename
        base_meta["audit_csv_key_prefix"] = key_prefix
    except Exception as exc:
        base_meta["audit_csv_error"] = str(exc)
    return base_meta


def _send_bulk_event_email_background(
    recipients: List[Tuple[str, Dict[str, object]]],
    subject: str,
    html: str,
    text: Optional[str],
    admin_id: int,
    event_id: int,
    event_slug: str,
    mode: str,
    request_method: Optional[str],
    request_path: Optional[str],
    skipped_no_email: int,
    skipped_duplicate: int,
) -> None:
    db = SessionLocal()
    sent = 0
    failed = 0
    errors: List[Dict[str, str]] = []
    try:
        for email_value, context in recipients:
            try:
                rendered_html = render_email_template(html, context, html_mode=True)
                rendered_text = render_email_template(text, context, html_mode=False) if text else None
                text_content = rendered_text if rendered_text is not None else derive_text_from_html(rendered_html)
                send_bulk_email(email_value, subject, rendered_html, text_content)
                sent += 1
            except Exception as exc:
                failed += 1
                if len(errors) < 10:
                    errors.append({"email": email_value, "error": str(exc)})
        admin = db.query(PdaUser).filter(PdaUser.id == admin_id).first()
        event = db.query(PersohubEvent).filter(PersohubEvent.id == event_id).first()
        if admin and event:
            _log_event_admin_action(
                db,
                admin,
                event,
                "send_persohub_event_bulk_email",
                method=request_method or "POST",
                path=request_path or f"/persohub/admin/persohub-events/{event_slug}/email/bulk",
                meta={
                    "mode": mode,
                    "queued": len(recipients),
                    "sent": sent,
                    "failed": failed,
                    "skipped_no_email": skipped_no_email,
                    "skipped_duplicate": skipped_duplicate,
                    "errors": errors,
                },
            )
    finally:
        db.close()


def _registered_entities(db: Session, event: PersohubEvent):
    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        query = (
            db.query(PersohubEventRegistration, PdaUser)
            .join(PdaUser, PersohubEventRegistration.user_id == PdaUser.id)
            .filter(
                PersohubEventRegistration.event_id == event.id,
                PersohubEventRegistration.entity_type == PersohubEventEntityType.USER,
                PersohubEventRegistration.user_id.isnot(None),
            )
        )
        rows = query.all()
        payload = []
        for reg, user in rows:
            payload.append(
                {
                    "entity_type": "user",
                    "entity_id": user.id,
                    "participant_id": user.id,
                    "name": user.name,
                    "participant_name": user.name,
                    "regno_or_code": user.regno,
                    "register_number": user.regno,
                    "participant_register_number": user.regno,
                    "email": user.email,
                    "department": user.dept,
                    "gender": user.gender,
                    "batch": _batch_from_regno(user.regno),
                    "profile_picture": user.image_url,
                    "status": _registration_status_label(reg.status),
                    "participant_status": _registration_status_label(reg.status),
                    "referral_code": reg.referral_code,
                    "referred_by": reg.referred_by,
                    "referral_count": int(reg.referral_count or 0),
                }
            )
        return payload
    rows = (
        db.query(PersohubEventRegistration, PersohubEventTeam)
        .join(PersohubEventTeam, PersohubEventRegistration.team_id == PersohubEventTeam.id)
        .filter(PersohubEventRegistration.event_id == event.id, PersohubEventRegistration.team_id.isnot(None))
        .all()
    )
    team_ids = [int(team.id) for _, team in rows]
    member_count_rows = (
        db.query(PersohubEventTeamMember.team_id, func.count(PersohubEventTeamMember.id))
        .filter(PersohubEventTeamMember.team_id.in_(team_ids))
        .group_by(PersohubEventTeamMember.team_id)
        .all()
        if team_ids
        else []
    )
    member_count_map = {int(team_id): int(count or 0) for team_id, count in member_count_rows}
    payload = []
    for reg, team in rows:
        payload.append(
            {
                "entity_type": "team",
                "entity_id": team.id,
                "name": team.team_name,
                "regno_or_code": team.team_code,
                "members_count": int(member_count_map.get(int(team.id), 0)),
                "status": _registration_status_label(reg.status),
                "participant_status": _registration_status_label(reg.status),
            }
        )
    return payload


def _round_scoring_entities(db: Session, event: PersohubEvent, round_row: PersohubEventRound):
    entities = _registered_entities(db, event)
    if not (round_row.is_frozen or round_row.state in {PersohubEventRoundState.COMPLETED, PersohubEventRoundState.REVEAL}):
        entities = [item for item in entities if _status_is_active(item.get("status"))]
    return entities


def _unregistered_entities(db: Session, event: PersohubEvent):
    query = db.query(PdaUser).filter(PdaUser.regno != "0000000000")
    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        registered_subq = (
            db.query(PersohubEventRegistration.user_id)
            .filter(
                PersohubEventRegistration.event_id == event.id,
                PersohubEventRegistration.entity_type == PersohubEventEntityType.USER,
                PersohubEventRegistration.user_id.isnot(None),
            )
            .subquery()
        )
        query = query.filter(~PdaUser.id.in_(registered_subq))
    else:
        team_ids = db.query(PersohubEventTeam.id).filter(PersohubEventTeam.event_id == event.id).subquery()
        member_ids = (
            db.query(PersohubEventTeamMember.user_id)
            .filter(PersohubEventTeamMember.team_id.in_(team_ids))
            .subquery()
        )
        lead_ids = (
            db.query(PersohubEventTeam.team_lead_user_id)
            .filter(PersohubEventTeam.event_id == event.id)
            .subquery()
        )
        query = query.filter(~PdaUser.id.in_(member_ids)).filter(~PdaUser.id.in_(lead_ids))

    users = query.order_by(PdaUser.name.asc().nullslast(), PdaUser.regno.asc()).all()
    payload = []
    for user in users:
        payload.append(
            {
                "entity_type": "user",
                "entity_id": user.id,
                "participant_id": user.id,
                "name": user.name,
                "participant_name": user.name,
                "regno_or_code": user.regno,
                "register_number": user.regno,
                "participant_register_number": user.regno,
                "email": user.email,
                "department": user.dept,
                "gender": user.gender,
                "batch": _batch_from_regno(user.regno),
                "profile_picture": user.image_url,
                "status": "Unregistered",
                "participant_status": "Unregistered",
                "referral_code": None,
                "referred_by": None,
                "referral_count": 0,
            }
        )
    return payload


@router.get("/persohub/admin/persohub-events/_parity_legacy", response_model=List[PersohubManagedEventResponse])
def list_managed_events(
    admin_ctx=Depends(get_persohub_admin_context),
    db: Session = Depends(get_db),
):
    admin_row = admin_ctx.get("admin_row")
    policy = admin_ctx.get("policy") if isinstance(admin_ctx.get("policy"), dict) else {}
    is_superadmin = bool(admin_ctx.get("is_superadmin"))
    if not is_superadmin and not admin_row:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin access required")

    query = db.query(PersohubEvent)
    if not is_superadmin:
        events = policy.get("events") if isinstance(policy.get("events"), dict) else {}
        allowed_slugs = [slug for slug, allowed in events.items() if allowed]
        if not allowed_slugs:
            return []
        query = query.filter(PersohubEvent.slug.in_(allowed_slugs))
    events = query.order_by(PersohubEvent.created_at.desc()).all()
    return [PersohubManagedEventResponse.model_validate(event) for event in events]


@router.post("/persohub/admin/persohub-events/_parity_legacy", response_model=PersohubManagedEventResponse)
def create_managed_event(
    payload: PersohubManagedEventCreate,
    admin: PdaUser = Depends(require_persohub_root_community_admin),
    community: PersohubCommunity = Depends(require_persohub_community),
    db: Session = Depends(get_db),
):
    _validate_event_dates(payload.start_date, payload.end_date)

    if payload.participant_mode.value == "team":
        if payload.team_min_size is None or payload.team_max_size is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="team_min_size and team_max_size are required for team events")
        if payload.team_min_size > payload.team_max_size:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="team_min_size cannot exceed team_max_size")
    else:
        payload.team_min_size = None
        payload.team_max_size = None

    round_count = payload.round_count
    if payload.round_mode == PersohubEventRoundMode.SINGLE:
        round_count = 1

    new_event = PersohubEvent(
        slug=_next_slug(db, payload.title),
        event_code=_next_event_code(db),
        community_id=int(community.id),
        title=payload.title.strip(),
        description=payload.description,
        start_date=payload.start_date,
        end_date=payload.end_date,
        poster_url=payload.poster_url,
        whatsapp_url=payload.whatsapp_url,
        external_url_name=(str(payload.external_url_name or "").strip() or "Join whatsapp channel"),
        event_type=_to_event_type(payload.event_type),
        format=_to_event_format(payload.format),
        template_option=_to_event_template(payload.template_option),
        participant_mode=_to_participant_mode(payload.participant_mode),
        round_mode=_to_round_mode(payload.round_mode),
        round_count=round_count,
        team_min_size=payload.team_min_size,
        team_max_size=payload.team_max_size,
        is_visible=True,
        registration_open=True,
        open_for=_to_event_open_for(payload.open_for),
        status=PersohubEventStatus.CLOSED,
    )
    db.add(new_event)
    db.flush()

    # Auto-provision rounds for single/multi round events.
    for round_no in range(1, round_count + 1):
        db.add(
            PersohubEventRound(
                event_id=new_event.id,
                round_no=round_no,
                name=f"Round {round_no}",
                mode=new_event.format,
                state=PersohubEventRoundState.DRAFT,
                evaluation_criteria=[{"name": "Score", "max_marks": 100}],
                requires_submission=False,
                submission_mode="file_or_link",
                allowed_mime_types=_default_round_allowed_mime_types(),
                max_file_size_mb=25,
            )
        )

    _sync_managed_event_to_home_item(db, new_event)

    # Add dynamic event policy key for all admins.
    admin_rows = db.query(PdaAdmin).all()
    for row in admin_rows:
        policy = _ensure_events_policy_shape(row.policy)
        if row.policy and row.policy.get("superAdmin"):
            policy["events"][new_event.slug] = True
        else:
            policy["events"][new_event.slug] = bool(policy["events"].get(new_event.slug, False))
        row.policy = policy

    db.commit()
    db.refresh(new_event)
    _log_event_admin_action(
        db,
        admin,
        new_event,
        "create_persohub_managed_event",
        method="POST",
        path="/persohub/admin/persohub-events",
        meta={"slug": new_event.slug, "event_id": new_event.id},
    )
    return PersohubManagedEventResponse.model_validate(new_event)


@router.put("/persohub/admin/persohub-events/{slug}/_parity_legacy", response_model=PersohubManagedEventResponse)
def update_managed_event(
    slug: str,
    payload: PersohubManagedEventUpdate,
    admin: PdaUser = Depends(require_persohub_root_community_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    updates = payload.model_dump(exclude_unset=True)
    next_start_date = updates.get("start_date", event.start_date)
    next_end_date = updates.get("end_date", event.end_date)
    _validate_event_dates(next_start_date, next_end_date)
    if "participant_mode" in updates and updates["participant_mode"] == PersohubEventParticipantMode.TEAM:
        min_size = updates.get("team_min_size", event.team_min_size)
        max_size = updates.get("team_max_size", event.team_max_size)
        if min_size is None or max_size is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="team_min_size and team_max_size are required for team events")
        if min_size > max_size:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="team_min_size cannot exceed team_max_size")

    if "round_mode" in updates and updates["round_mode"] == PersohubEventRoundMode.SINGLE:
        updates["round_count"] = 1

    if "event_type" in updates:
        updates["event_type"] = _to_event_type(payload.event_type)
    if "format" in updates:
        updates["format"] = _to_event_format(payload.format)
    if "template_option" in updates:
        updates["template_option"] = _to_event_template(payload.template_option)
    if "participant_mode" in updates:
        updates["participant_mode"] = _to_participant_mode(payload.participant_mode)
    if "round_mode" in updates:
        updates["round_mode"] = _to_round_mode(payload.round_mode)
    if "status" in updates:
        updates["status"] = _to_event_status(payload.status)
    if "external_url_name" in updates:
        updates["external_url_name"] = str(updates.get("external_url_name") or "").strip() or "Join whatsapp channel"
    if "open_for" in updates:
        updates["open_for"] = _to_event_open_for(payload.open_for)

    for field, value in updates.items():
        setattr(event, field, value)

    _sync_managed_event_to_home_item(db, event)

    db.commit()
    db.refresh(event)
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_managed_event",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}",
        meta={"slug": slug},
    )
    return PersohubManagedEventResponse.model_validate(event)


@router.get("/persohub/admin/persohub-events/{slug}", response_model=PersohubManagedEventResponse)
def get_managed_event(
    slug: str,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    return PersohubManagedEventResponse.model_validate(event)


@router.delete("/persohub/admin/persohub-events/{slug}/_parity_legacy")
def delete_managed_event(
    slug: str,
    admin: PdaUser = Depends(require_persohub_root_community_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    event_id = int(event.id)
    event_slug = str(event.slug)

    team_ids = [
        int(row[0])
        for row in db.query(PersohubEventTeam.id).filter(PersohubEventTeam.event_id == event_id).all()
    ]
    round_ids = [
        int(row[0])
        for row in db.query(PersohubEventRound.id).filter(PersohubEventRound.event_id == event_id).all()
    ]

    if team_ids:
        db.query(PersohubEventRoundPanelAssignment).filter(PersohubEventRoundPanelAssignment.team_id.in_(team_ids)).delete(synchronize_session=False)
        db.query(PersohubEventInvite).filter(PersohubEventInvite.team_id.in_(team_ids)).delete(synchronize_session=False)
        db.query(PersohubEventBadge).filter(PersohubEventBadge.team_id.in_(team_ids)).delete(synchronize_session=False)
        db.query(PersohubEventScore).filter(PersohubEventScore.team_id.in_(team_ids)).delete(synchronize_session=False)
        db.query(PersohubEventRoundSubmission).filter(PersohubEventRoundSubmission.team_id.in_(team_ids)).delete(synchronize_session=False)
        db.query(PersohubEventAttendance).filter(PersohubEventAttendance.team_id.in_(team_ids)).delete(synchronize_session=False)
        db.query(PersohubEventRegistration).filter(PersohubEventRegistration.team_id.in_(team_ids)).delete(synchronize_session=False)
        db.query(PersohubEventTeamMember).filter(PersohubEventTeamMember.team_id.in_(team_ids)).delete(synchronize_session=False)

    if round_ids:
        db.query(PersohubEventRoundPanelAssignment).filter(PersohubEventRoundPanelAssignment.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(PersohubEventRoundPanelMember).filter(PersohubEventRoundPanelMember.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(PersohubEventRoundPanel).filter(PersohubEventRoundPanel.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(PersohubEventScore).filter(PersohubEventScore.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(PersohubEventRoundSubmission).filter(PersohubEventRoundSubmission.round_id.in_(round_ids)).delete(synchronize_session=False)
        db.query(PersohubEventAttendance).filter(PersohubEventAttendance.round_id.in_(round_ids)).delete(synchronize_session=False)

    db.query(PersohubEventInvite).filter(PersohubEventInvite.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventBadge).filter(PersohubEventBadge.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventRoundPanelAssignment).filter(PersohubEventRoundPanelAssignment.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventRoundPanelMember).filter(PersohubEventRoundPanelMember.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventRoundPanel).filter(PersohubEventRoundPanel.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventScore).filter(PersohubEventScore.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventRoundSubmission).filter(PersohubEventRoundSubmission.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventAttendance).filter(PersohubEventAttendance.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventRegistration).filter(PersohubEventRegistration.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventTeamMember).filter(
        PersohubEventTeamMember.team_id.in_(
            db.query(PersohubEventTeam.id).filter(PersohubEventTeam.event_id == event_id)
        )
    ).delete(synchronize_session=False)
    db.query(PersohubEventTeam).filter(PersohubEventTeam.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventRound).filter(PersohubEventRound.event_id == event_id).delete(synchronize_session=False)
    db.query(PersohubEventLog).filter(
        (PersohubEventLog.event_id == event_id) | (PersohubEventLog.event_slug == event_slug)
    ).delete(synchronize_session=False)

    admin_rows = db.query(PdaAdmin).all()
    for row in admin_rows:
        policy = _ensure_events_policy_shape(row.policy)
        if event_slug in policy["events"]:
            del policy["events"][event_slug]
            row.policy = policy

    _delete_managed_home_item(db, event_slug)
    db.delete(event)
    db.commit()

    log_admin_action(
        db,
        admin,
        "delete_persohub_managed_event",
        method="DELETE",
        path=f"/persohub/admin/persohub-events/{event_slug}",
        meta={"slug": event_slug, "event_id": event_id},
    )
    return {"message": "Event deleted"}


@router.put("/persohub/admin/persohub-events/{slug}/status", response_model=PersohubManagedEventResponse)
def update_managed_event_status(
    slug: str,
    payload: PersohubManagedEventStatusUpdate,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    event.status = _to_event_status(payload.status)
    db.commit()
    db.refresh(event)
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_managed_event_status",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/status",
        meta={"slug": slug, "status": payload.status.value},
    )
    return PersohubManagedEventResponse.model_validate(event)


@router.put("/persohub/admin/persohub-events/{slug}/registration", response_model=PersohubManagedEventResponse)
def update_managed_event_registration(
    slug: str,
    payload: PersohubManagedEventRegistrationUpdate,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    event.registration_open = bool(payload.registration_open)
    db.commit()
    db.refresh(event)
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_managed_event_registration",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/registration",
        meta={"slug": slug, "registration_open": bool(payload.registration_open)},
    )
    return PersohubManagedEventResponse.model_validate(event)


@router.put("/persohub/admin/persohub-events/{slug}/visibility", response_model=PersohubManagedEventResponse)
def update_managed_event_visibility(
    slug: str,
    payload: PersohubManagedEventVisibilityUpdate,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    event.is_visible = bool(payload.is_visible)
    db.commit()
    db.refresh(event)
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_managed_event_visibility",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/visibility",
        meta={"slug": slug, "is_visible": bool(payload.is_visible)},
    )
    return PersohubManagedEventResponse.model_validate(event)


@router.get("/persohub/admin/persohub-events/{slug}/dashboard")
def event_dashboard(
    slug: str,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    registrations = db.query(PersohubEventRegistration).filter(PersohubEventRegistration.event_id == event.id).count()
    round_rows = db.query(PersohubEventRound).filter(PersohubEventRound.event_id == event.id).order_by(PersohubEventRound.round_no.asc()).all()
    rounds = len(round_rows)
    attendance_present = db.query(PersohubEventAttendance).filter(
        PersohubEventAttendance.event_id == event.id,
        PersohubEventAttendance.is_present == True,  # noqa: E712
    ).count()
    scores = db.query(PersohubEventScore).filter(PersohubEventScore.event_id == event.id).count()
    badges = db.query(PersohubEventBadge).filter(PersohubEventBadge.event_id == event.id).count()
    active_count = db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.status == PersohubEventRegistrationStatus.ACTIVE,
    ).count()
    eliminated_count = db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.status == PersohubEventRegistrationStatus.ELIMINATED,
    ).count()
    rounds_completed = sum(1 for row in round_rows if row.state in {PersohubEventRoundState.COMPLETED, PersohubEventRoundState.REVEAL})
    current_active = next((row for row in round_rows if row.state == PersohubEventRoundState.ACTIVE), None)

    department_distribution: Dict[str, int] = {}
    gender_distribution: Dict[str, int] = {}
    batch_distribution: Dict[str, int] = {}
    leaderboard_scores: List[float] = []
    entities = _registered_entities(db, event)
    active_entities = [item for item in entities if _status_is_active(item.get("status"))]

    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        dept_counter = Counter(str(item.get("department") or "").strip() for item in entities if str(item.get("department") or "").strip())
        gender_counter = Counter(str(item.get("gender") or "").strip() for item in entities if str(item.get("gender") or "").strip())
        batch_counter = Counter(str(item.get("batch") or "").strip() for item in entities if str(item.get("batch") or "").strip())
        department_distribution = dict(dept_counter)
        gender_distribution = dict(gender_counter)
        batch_distribution = dict(batch_counter)

        score_rows = db.execute(
            text(
                """
                SELECT user_id, COALESCE(SUM(normalized_score), 0) AS total
                FROM persohub_event_scores
                WHERE event_id = :event_id
                  AND entity_type = 'USER'
                  AND user_id IS NOT NULL
                GROUP BY user_id
                """
            ),
            {"event_id": event.id},
        ).mappings().all()
        score_map = {int(row["user_id"]): float(row["total"] or 0.0) for row in score_rows}
        leaderboard_scores = [float(score_map.get(int(item["entity_id"]), 0.0)) for item in active_entities]
    else:
        score_rows = db.execute(
            text(
                """
                SELECT team_id, COALESCE(SUM(total_score), 0) AS total
                FROM persohub_event_scores
                WHERE event_id = :event_id
                  AND entity_type = 'TEAM'
                  AND team_id IS NOT NULL
                GROUP BY team_id
                """
            ),
            {"event_id": event.id},
        ).mappings().all()
        score_map = {int(row["team_id"]): float(row["total"] or 0.0) for row in score_rows}
        leaderboard_scores = [float(score_map.get(int(item["entity_id"]), 0.0)) for item in active_entities]

    leaderboard_min_score = min(leaderboard_scores) if leaderboard_scores else None
    leaderboard_max_score = max(leaderboard_scores) if leaderboard_scores else None
    leaderboard_avg_score = (sum(leaderboard_scores) / len(leaderboard_scores)) if leaderboard_scores else None

    return {
        "event": PersohubManagedEventResponse.model_validate(event),
        "registrations": registrations,
        "rounds": rounds,
        "attendance_present": attendance_present,
        "score_rows": scores,
        "badges": badges,
        "active_count": active_count,
        "eliminated_count": eliminated_count,
        "rounds_completed": rounds_completed,
        "current_active_round": (
            {
                "round_id": current_active.id,
                "round_no": int(current_active.round_no),
                "name": current_active.name,
            }
            if current_active
            else None
        ),
        "department_distribution": department_distribution,
        "gender_distribution": gender_distribution,
        "batch_distribution": batch_distribution,
        "leaderboard_min_score": leaderboard_min_score,
        "leaderboard_max_score": leaderboard_max_score,
        "leaderboard_avg_score": leaderboard_avg_score,
    }


@router.get("/persohub/admin/persohub-events/{slug}/participants")
def event_participants(
    slug: str,
    search: Optional[str] = None,
    department: Optional[str] = None,
    gender: Optional[str] = None,
    batch: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    response: Response = None,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    items = _registered_entities(db, event)
    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        if department:
            items = [item for item in items if str(item.get("department") or "") == str(department)]
        if gender:
            items = [item for item in items if str(item.get("gender") or "") == str(gender)]
        if batch:
            items = [item for item in items if str(item.get("batch") or "") == str(batch)]
    if status_filter:
        normalized = str(status_filter or "").strip().lower()
        items = [item for item in items if str(item.get("status") or "").strip().lower() == normalized]
    if search:
        needle = search.lower()
        filtered = []
        for item in items:
            haystack = " ".join(
                [
                    str(item.get("name") or ""),
                    str(item.get("regno_or_code") or ""),
                    str(item.get("email") or ""),
                    str(item.get("department") or ""),
                    str(item.get("gender") or ""),
                    str(item.get("batch") or ""),
                ]
            ).lower()
            if needle in haystack:
                filtered.append(item)
        items = filtered

    total = len(items)
    start = (page - 1) * page_size
    end = start + page_size
    paged = items[start:end]
    if response is not None:
        response.headers["X-Total-Count"] = str(total)
        response.headers["X-Page"] = str(page)
        response.headers["X-Page-Size"] = str(page_size)
    return paged


@router.get("/persohub/admin/persohub-events/{slug}/unregistered-users")
def event_unregistered_users(
    slug: str,
    search: Optional[str] = None,
    department: Optional[str] = None,
    gender: Optional[str] = None,
    batch: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    response: Response = None,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    items = _unregistered_entities(db, event)
    if department:
        items = [item for item in items if str(item.get("department") or "") == str(department)]
    if gender:
        items = [item for item in items if str(item.get("gender") or "") == str(gender)]
    if batch:
        items = [item for item in items if str(item.get("batch") or "") == str(batch)]
    if search:
        needle = search.lower()
        filtered = []
        for item in items:
            haystack = " ".join(
                [
                    str(item.get("name") or ""),
                    str(item.get("regno_or_code") or ""),
                    str(item.get("email") or ""),
                    str(item.get("department") or ""),
                    str(item.get("gender") or ""),
                    str(item.get("batch") or ""),
                ]
            ).lower()
            if needle in haystack:
                filtered.append(item)
        items = filtered

    total = len(items)
    start = (page - 1) * page_size
    end = start + page_size
    paged = items[start:end]
    if response is not None:
        response.headers["X-Total-Count"] = str(total)
        response.headers["X-Page"] = str(page)
        response.headers["X-Page-Size"] = str(page_size)
    return paged


@router.put("/persohub/admin/persohub-events/{slug}/participants/{user_id}/status")
def update_participant_status(
    slug: str,
    user_id: int,
    status_value: str = Query(..., alias="status"),
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    if event.participant_mode != PersohubEventParticipantMode.INDIVIDUAL:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Status is only available for individual events")
    normalized = str(status_value or "").strip().lower()
    if normalized not in {"active", "eliminated"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="status must be Active or Eliminated")
    row = db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.user_id == user_id,
        PersohubEventRegistration.entity_type == PersohubEventEntityType.USER,
    ).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Participant not found")
    row.status = PersohubEventRegistrationStatus.ACTIVE if normalized == "active" else PersohubEventRegistrationStatus.ELIMINATED
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_event_participant_status",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/participants/{user_id}/status",
        meta={"user_id": user_id, "status": normalized},
    )
    return {"message": "Status updated"}


@router.put(
    "/persohub/admin/persohub-events/{slug}/registrations/status-bulk",
    response_model=PersohubManagedRegistrationStatusBulkResponse,
)
def update_registration_status_bulk(
    slug: str,
    payload: PersohubManagedRegistrationStatusBulkRequest,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    raw_updates = list(payload.updates or [])
    if not raw_updates:
        return PersohubManagedRegistrationStatusBulkResponse(updated_count=0)

    is_team_mode = event.participant_mode == PersohubEventParticipantMode.TEAM
    expected_entity_type = "team" if is_team_mode else "user"
    expected_entity_enum = PersohubEventEntityType.TEAM if is_team_mode else PersohubEventEntityType.USER

    deduped_updates: Dict[int, str] = {}
    for item in raw_updates:
        item_entity_type = str(item.entity_type.value if hasattr(item.entity_type, "value") else item.entity_type).strip().lower()
        if item_entity_type != expected_entity_type:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Entity type must be {expected_entity_type} for this event",
            )
        status_text = str(item.status.value if hasattr(item.status, "value") else item.status).strip()
        if status_text not in {"Active", "Eliminated"}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="status must be Active or Eliminated")
        deduped_updates[int(item.entity_id)] = status_text

    if not deduped_updates:
        return PersohubManagedRegistrationStatusBulkResponse(updated_count=0)

    entity_ids = list(deduped_updates.keys())
    query = db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.entity_type == expected_entity_enum,
    )
    if is_team_mode:
        query = query.filter(PersohubEventRegistration.team_id.in_(entity_ids))
    else:
        query = query.filter(PersohubEventRegistration.user_id.in_(entity_ids))
    rows = query.all()

    updated_count = 0
    for row in rows:
        target_id = int(row.team_id) if is_team_mode else int(row.user_id)
        target_status = deduped_updates.get(target_id)
        if target_status is None:
            continue
        next_status = (
            PersohubEventRegistrationStatus.ACTIVE
            if target_status == "Active"
            else PersohubEventRegistrationStatus.ELIMINATED
        )
        if row.status != next_status:
            row.status = next_status
            updated_count += 1

    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "bulk_update_persohub_event_registration_status",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/registrations/status-bulk",
        meta={
            "entity_type": expected_entity_type,
            "requested_updates": len(deduped_updates),
            "updated_count": updated_count,
        },
    )
    return PersohubManagedRegistrationStatusBulkResponse(updated_count=updated_count)


@router.delete("/persohub/admin/persohub-events/{slug}/participants/{user_id}")
def delete_participant_with_cascade(
    slug: str,
    user_id: int,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    if event.participant_mode != PersohubEventParticipantMode.INDIVIDUAL:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Participant deletion is only available for individual events")

    registration = db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.entity_type == PersohubEventEntityType.USER,
        PersohubEventRegistration.user_id == user_id,
    ).first()
    if not registration:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Participant not found")

    participant = db.query(PdaUser).filter(PdaUser.id == user_id).first()
    participant_name = participant.name if participant else None
    participant_regno = participant.regno if participant else None

    db.query(PersohubEventInvite).filter(
        PersohubEventInvite.event_id == event.id,
        PersohubEventInvite.invited_user_id == user_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventInvite).filter(
        PersohubEventInvite.event_id == event.id,
        PersohubEventInvite.invited_by_user_id == user_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventBadge).filter(
        PersohubEventBadge.event_id == event.id,
        PersohubEventBadge.user_id == user_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventScore).filter(
        PersohubEventScore.event_id == event.id,
        PersohubEventScore.user_id == user_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventAttendance).filter(
        PersohubEventAttendance.event_id == event.id,
        PersohubEventAttendance.user_id == user_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.entity_type == PersohubEventEntityType.USER,
        PersohubEventRegistration.user_id == user_id,
    ).delete(synchronize_session=False)
    db.commit()

    _log_event_admin_action(
        db,
        admin,
        event,
        "delete_persohub_event_participant",
        method="DELETE",
        path=f"/persohub/admin/persohub-events/{slug}/participants/{user_id}",
        meta={"user_id": user_id, "regno": participant_regno, "name": participant_name},
    )
    return {"message": "Participant deleted", "participant_id": user_id}


@router.get("/persohub/admin/persohub-events/{slug}/participants/{user_id}/rounds")
def participant_rounds(
    slug: str,
    user_id: int,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    entity_type = (
        PersohubEventEntityType.USER
        if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL
        else PersohubEventEntityType.TEAM
    )
    registration_filters = [
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.entity_type == entity_type,
    ]
    if entity_type == PersohubEventEntityType.USER:
        registration_filters.append(PersohubEventRegistration.user_id == user_id)
    else:
        registration_filters.append(PersohubEventRegistration.team_id == user_id)
    registration = db.query(PersohubEventRegistration).filter(*registration_filters).first()
    if not registration:
        missing_label = "Participant" if entity_type == PersohubEventEntityType.USER else "Team"
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"{missing_label} not found")
    params = {"event_id": event.id, "entity_id": user_id}
    if entity_type == PersohubEventEntityType.USER:
        rows = db.execute(
            text(
                """
                WITH rounds AS (
                    SELECT id, round_no, name, state
                    FROM persohub_event_rounds
                    WHERE event_id = :event_id
                ),
                target_scores AS (
                    SELECT
                        round_id,
                        is_present,
                        total_score,
                        normalized_score
                    FROM persohub_event_scores
                    WHERE event_id = :event_id
                      AND entity_type = 'USER'
                      AND user_id = :entity_id
                ),
                ranked_present AS (
                    SELECT
                        round_id,
                        user_id AS entity_id,
                        DENSE_RANK() OVER (
                            PARTITION BY round_id
                            ORDER BY normalized_score DESC
                        ) AS round_rank
                    FROM persohub_event_scores
                    WHERE event_id = :event_id
                      AND entity_type = 'USER'
                      AND is_present = TRUE
                )
                SELECT
                    r.id AS round_id,
                    r.round_no,
                    r.name AS round_name,
                    r.state AS round_state,
                    ts.is_present,
                    ts.total_score,
                    ts.normalized_score,
                    rp.round_rank
                FROM rounds r
                LEFT JOIN target_scores ts ON ts.round_id = r.id
                LEFT JOIN ranked_present rp
                    ON rp.round_id = r.id
                    AND rp.entity_id = :entity_id
                ORDER BY r.round_no ASC
                """
            ),
            params,
        ).mappings().all()
    else:
        rows = db.execute(
            text(
                """
                WITH rounds AS (
                    SELECT id, round_no, name, state
                    FROM persohub_event_rounds
                    WHERE event_id = :event_id
                ),
                target_scores AS (
                    SELECT
                        round_id,
                        is_present,
                        total_score,
                        normalized_score
                    FROM persohub_event_scores
                    WHERE event_id = :event_id
                      AND entity_type = 'TEAM'
                      AND team_id = :entity_id
                ),
                ranked_present AS (
                    SELECT
                        round_id,
                        team_id AS entity_id,
                        DENSE_RANK() OVER (
                            PARTITION BY round_id
                            ORDER BY normalized_score DESC
                        ) AS round_rank
                    FROM persohub_event_scores
                    WHERE event_id = :event_id
                      AND entity_type = 'TEAM'
                      AND is_present = TRUE
                )
                SELECT
                    r.id AS round_id,
                    r.round_no,
                    r.name AS round_name,
                    r.state AS round_state,
                    ts.is_present,
                    ts.total_score,
                    ts.normalized_score,
                    rp.round_rank
                FROM rounds r
                LEFT JOIN target_scores ts ON ts.round_id = r.id
                LEFT JOIN ranked_present rp
                    ON rp.round_id = r.id
                    AND rp.entity_id = :entity_id
                ORDER BY r.round_no ASC
                """
            ),
            params,
        ).mappings().all()

    items = []
    for row in rows:
        is_present = row["is_present"]
        if registration.status == PersohubEventRegistrationStatus.ELIMINATED:
            status_label = "Eliminated"
        elif is_present is None:
            status_label = "Pending"
        else:
            status_label = "Active" if bool(is_present) else "Absent"

        round_state = row["round_state"]
        round_state_text = str(round_state)
        if round_state_text.isupper():
            round_state_text = round_state_text.title()

        items.append(
            {
                "round_id": int(row["round_id"]),
                "round_no": f"PF{int(row['round_no']):02d}",
                "round_name": row["round_name"],
                "round_state": round_state_text,
                "status": status_label,
                "is_present": bool(is_present) if is_present is not None else None,
                "total_score": float(row["total_score"]) if row["total_score"] is not None else None,
                "normalized_score": float(row["normalized_score"]) if row["normalized_score"] is not None else None,
                "round_rank": int(row["round_rank"]) if row["round_rank"] is not None else None,
            }
        )
    return items


@router.get("/persohub/admin/persohub-events/{slug}/participants/{user_id}/summary")
def participant_summary(
    slug: str,
    user_id: int,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    entity_type = (
        PersohubEventEntityType.USER
        if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL
        else PersohubEventEntityType.TEAM
    )
    registration_filters = [
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.entity_type == entity_type,
    ]
    if entity_type == PersohubEventEntityType.USER:
        registration_filters.append(PersohubEventRegistration.user_id == user_id)
    else:
        registration_filters.append(PersohubEventRegistration.team_id == user_id)
    registration = db.query(PersohubEventRegistration).filter(*registration_filters).first()
    if not registration:
        missing_label = "Participant" if entity_type == PersohubEventEntityType.USER else "Team"
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"{missing_label} not found")
    params = {"event_id": event.id, "entity_id": user_id}
    if entity_type == PersohubEventEntityType.USER:
        summary_row = db.execute(
            text(
                """
                WITH active_entities AS (
                    SELECT user_id AS entity_id
                    FROM persohub_event_registrations
                    WHERE event_id = :event_id
                      AND entity_type = 'USER'
                      AND status = 'ACTIVE'
                      AND user_id IS NOT NULL
                ),
                active_totals AS (
                    SELECT
                        ae.entity_id,
                        COALESCE(SUM(s.normalized_score), 0) AS total_points
                    FROM active_entities ae
                    LEFT JOIN persohub_event_scores s
                        ON s.event_id = :event_id
                       AND s.entity_type = 'USER'
                       AND s.user_id = ae.entity_id
                    GROUP BY ae.entity_id
                ),
                ranked AS (
                    SELECT
                        entity_id,
                        DENSE_RANK() OVER (ORDER BY total_points DESC) AS overall_rank
                    FROM active_totals
                ),
                target_total AS (
                    SELECT COALESCE(SUM(normalized_score), 0) AS overall_points
                    FROM persohub_event_scores
                    WHERE event_id = :event_id
                      AND entity_type = 'USER'
                      AND user_id = :entity_id
                )
                SELECT
                    tt.overall_points,
                    r.overall_rank
                FROM target_total tt
                LEFT JOIN ranked r ON r.entity_id = :entity_id
                """
            ),
            params,
        ).mappings().first()
    else:
        summary_row = db.execute(
            text(
                """
                WITH active_entities AS (
                    SELECT team_id AS entity_id
                    FROM persohub_event_registrations
                    WHERE event_id = :event_id
                      AND entity_type = 'TEAM'
                      AND status = 'ACTIVE'
                      AND team_id IS NOT NULL
                ),
                active_totals AS (
                    SELECT
                        ae.entity_id,
                        COALESCE(SUM(s.total_score), 0) AS total_points
                    FROM active_entities ae
                    LEFT JOIN persohub_event_scores s
                        ON s.event_id = :event_id
                       AND s.entity_type = 'TEAM'
                       AND s.team_id = ae.entity_id
                    GROUP BY ae.entity_id
                ),
                ranked AS (
                    SELECT
                        entity_id,
                        DENSE_RANK() OVER (ORDER BY total_points DESC) AS overall_rank
                    FROM active_totals
                ),
                target_total AS (
                    SELECT COALESCE(SUM(total_score), 0) AS overall_points
                    FROM persohub_event_scores
                    WHERE event_id = :event_id
                      AND entity_type = 'TEAM'
                      AND team_id = :entity_id
                )
                SELECT
                    tt.overall_points,
                    r.overall_rank
                FROM target_total tt
                LEFT JOIN ranked r ON r.entity_id = :entity_id
                """
            ),
            params,
        ).mappings().first()
    target_total = float((summary_row or {}).get("overall_points") or 0.0)
    rank_value = (summary_row or {}).get("overall_rank")
    rank = int(rank_value) if rank_value is not None else None

    return {
        "participant_id": user_id,
        "overall_rank": rank if registration.status == PersohubEventRegistrationStatus.ACTIVE else None,
        "overall_points": float(target_total),
    }


@router.get("/persohub/admin/persohub-events/{slug}/teams/{team_id}", response_model=PersohubManagedTeamResponse)
def team_details(
    slug: str,
    team_id: int,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    if event.participant_mode != PersohubEventParticipantMode.TEAM:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Team details are only available for team events")
    team = db.query(PersohubEventTeam).filter(PersohubEventTeam.event_id == event.id, PersohubEventTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Team not found")
    members = (
        db.query(PersohubEventTeamMember, PdaUser)
        .join(PdaUser, PersohubEventTeamMember.user_id == PdaUser.id)
        .filter(PersohubEventTeamMember.team_id == team.id)
        .order_by(PersohubEventTeamMember.role.desc(), PdaUser.regno.asc())
        .all()
    )
    return PersohubManagedTeamResponse(
        id=team.id,
        event_id=team.event_id,
        team_code=team.team_code,
        team_name=team.team_name,
        team_lead_user_id=team.team_lead_user_id,
        members=[
            {
                "user_id": user.id,
                "regno": user.regno,
                "name": user.name,
                "role": member.role,
            }
            for member, user in members
        ],
    )


@router.delete("/persohub/admin/persohub-events/{slug}/teams/{team_id}")
def delete_team_with_cascade(
    slug: str,
    team_id: int,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    if event.participant_mode != PersohubEventParticipantMode.TEAM:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Team deletion is only available for team events")

    team = db.query(PersohubEventTeam).filter(PersohubEventTeam.event_id == event.id, PersohubEventTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Team not found")

    team_name = str(team.team_name)
    team_code = str(team.team_code)

    db.query(PersohubEventInvite).filter(
        PersohubEventInvite.event_id == event.id,
        PersohubEventInvite.team_id == team_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventBadge).filter(
        PersohubEventBadge.event_id == event.id,
        PersohubEventBadge.team_id == team_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventScore).filter(
        PersohubEventScore.event_id == event.id,
        PersohubEventScore.team_id == team_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventAttendance).filter(
        PersohubEventAttendance.event_id == event.id,
        PersohubEventAttendance.team_id == team_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.team_id == team_id,
    ).delete(synchronize_session=False)
    db.query(PersohubEventTeamMember).filter(
        PersohubEventTeamMember.team_id == team_id
    ).delete(synchronize_session=False)
    db.query(PersohubEventTeam).filter(
        PersohubEventTeam.event_id == event.id,
        PersohubEventTeam.id == team_id,
    ).delete(synchronize_session=False)
    db.commit()

    _log_event_admin_action(
        db,
        admin,
        event,
        "delete_persohub_event_team",
        method="DELETE",
        path=f"/persohub/admin/persohub-events/{slug}/teams/{team_id}",
        meta={"team_id": team_id, "team_code": team_code, "team_name": team_name},
    )
    return {"message": "Team deleted", "team_id": team_id}


@router.get("/persohub/admin/persohub-events/{slug}/attendance")
def event_attendance(
    slug: str,
    round_id: int = Query(...),
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    _get_event_round_or_404(db, event.id, round_id)
    entities = _registered_entities(db, event)
    rows = db.query(PersohubEventAttendance).filter(
        PersohubEventAttendance.event_id == event.id,
        PersohubEventAttendance.round_id == round_id,
    ).all()
    row_map = {}
    for row in rows:
        key = ("user", row.user_id) if row.user_id else ("team", row.team_id)
        row_map[key] = row
    results = []
    for entity in entities:
        key = (entity["entity_type"], entity["entity_id"])
        row = row_map.get(key)
        results.append(
            {
                **entity,
                "attendance_id": row.id if row else None,
                "round_id": row.round_id if row else round_id,
                "is_present": bool(row.is_present) if row else False,
                "marked_at": row.marked_at.isoformat() if row and row.marked_at else None,
            }
        )
    return results


@router.post("/persohub/admin/persohub-events/{slug}/attendance/mark")
def mark_attendance(
    slug: str,
    payload: PersohubManagedAttendanceMarkRequest,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    if payload.round_id is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="round_id is required")
    _get_event_round_or_404(db, event.id, int(payload.round_id))
    entity_type = PersohubEventEntityType.USER if payload.entity_type.value == "user" else PersohubEventEntityType.TEAM
    if entity_type == PersohubEventEntityType.USER and not payload.user_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="user_id required for user attendance")
    if entity_type == PersohubEventEntityType.TEAM and not payload.team_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="team_id required for team attendance")

    row = db.query(PersohubEventAttendance).filter(
        PersohubEventAttendance.event_id == event.id,
        PersohubEventAttendance.round_id == payload.round_id,
        PersohubEventAttendance.entity_type == entity_type,
        PersohubEventAttendance.user_id == payload.user_id,
        PersohubEventAttendance.team_id == payload.team_id,
    ).first()
    if row:
        row.is_present = payload.is_present
        row.marked_by_user_id = admin.id
    else:
        row = PersohubEventAttendance(
            event_id=event.id,
            round_id=payload.round_id,
            entity_type=entity_type,
            user_id=payload.user_id,
            team_id=payload.team_id,
            is_present=payload.is_present,
            marked_by_user_id=admin.id,
        )
        db.add(row)
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "mark_persohub_event_attendance",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/attendance/mark",
        meta={
            "entity_type": payload.entity_type.value,
            "user_id": payload.user_id,
            "team_id": payload.team_id,
            "round_id": payload.round_id,
            "is_present": bool(payload.is_present),
        },
    )
    return {"message": "Attendance updated"}


@router.post("/persohub/admin/persohub-events/{slug}/attendance/scan")
def scan_attendance(
    slug: str,
    payload: PersohubManagedAttendanceScanRequest,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    if payload.round_id is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="round_id is required")
    _get_event_round_or_404(db, event.id, int(payload.round_id))
    decoded = decode_token(payload.token)
    if decoded.get("qr") != "persohub_event_attendance" or decoded.get("event_slug") != event.slug:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid QR token")
    entity_type = decoded.get("entity_type")
    entity_id = int(decoded.get("entity_id"))
    mark_payload = PersohubManagedAttendanceMarkRequest(
        entity_type=PersohubManagedEntityTypeEnum.USER if entity_type == "user" else PersohubManagedEntityTypeEnum.TEAM,
        user_id=entity_id if entity_type == "user" else None,
        team_id=entity_id if entity_type == "team" else None,
        round_id=payload.round_id,
        is_present=True,
    )
    response = mark_attendance(slug=slug, payload=mark_payload, admin=admin, db=db)
    _log_event_admin_action(
        db,
        admin,
        event,
        "scan_persohub_event_attendance",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/attendance/scan",
        meta={"round_id": payload.round_id, "entity_type": entity_type, "entity_id": entity_id},
    )
    return response


@router.get("/persohub/admin/persohub-events/{slug}/rounds", response_model=List[PersohubManagedRoundResponse])
def list_rounds(
    slug: str,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    rounds = db.query(PersohubEventRound).filter(PersohubEventRound.event_id == event.id).order_by(PersohubEventRound.round_no.asc()).all()
    return [PersohubManagedRoundResponse.model_validate(row) for row in rounds]


@router.post("/persohub/admin/persohub-events/{slug}/rounds", response_model=PersohubManagedRoundResponse)
def create_round(
    slug: str,
    payload: PersohubManagedRoundCreate,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    existing = db.query(PersohubEventRound).filter(PersohubEventRound.event_id == event.id, PersohubEventRound.round_no == payload.round_no).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Round already exists")
    round_row = PersohubEventRound(
        event_id=event.id,
        round_no=payload.round_no,
        name=payload.name,
        description=payload.description,
        round_poster=payload.round_poster,
        external_url=payload.external_url,
        external_url_name=(str(payload.external_url_name or "").strip() or "Explore Round"),
        date=payload.date,
        mode=_to_event_format(payload.mode),
        evaluation_criteria=[c.model_dump() for c in payload.evaluation_criteria] if payload.evaluation_criteria else [{"name": "Score", "max_marks": 100}],
        requires_submission=bool(payload.requires_submission),
        submission_mode=(payload.submission_mode.value if hasattr(payload.submission_mode, "value") else str(payload.submission_mode or "file_or_link")),
        submission_deadline=payload.submission_deadline,
        allowed_mime_types=list(payload.allowed_mime_types or _default_round_allowed_mime_types()),
        max_file_size_mb=int(payload.max_file_size_mb or 25),
        panel_mode_enabled=bool(payload.panel_mode_enabled),
        panel_team_distribution_mode=_normalize_panel_distribution_mode(
            payload.panel_team_distribution_mode.value
            if hasattr(payload.panel_team_distribution_mode, "value")
            else payload.panel_team_distribution_mode
        ),
        panel_structure_locked=bool(payload.panel_structure_locked),
    )
    db.add(round_row)
    _sync_event_round_count(db, event.id)
    db.commit()
    db.refresh(round_row)
    _log_event_admin_action(
        db,
        admin,
        event,
        "create_persohub_event_round",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/rounds",
        meta={"round_id": round_row.id},
    )
    return PersohubManagedRoundResponse.model_validate(round_row)


@router.put("/persohub/admin/persohub-events/{slug}/rounds/{round_id}", response_model=PersohubManagedRoundResponse)
def update_round(
    slug: str,
    round_id: int,
    payload: PersohubManagedRoundUpdate,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    panel_mode_before = bool(round_row.panel_mode_enabled)
    updates = payload.model_dump(exclude_unset=True)
    if "external_url" not in updates and updates.get("whatsapp_url") is not None:
        updates["external_url"] = updates.get("whatsapp_url")
    updates.pop("whatsapp_url", None)
    if "external_url_name" in updates:
        updates["external_url_name"] = str(updates.get("external_url_name") or "").strip() or "Explore Round"
    eliminate_absent = bool(updates.pop("eliminate_absent", False))
    requested_round_no = updates.pop("round_no", None)
    if "mode" in updates:
        updates["mode"] = _to_event_format(payload.mode)
    if "state" in updates:
        updates["state"] = _to_round_state(payload.state)
    if "evaluation_criteria" in updates and payload.evaluation_criteria is not None:
        updates["evaluation_criteria"] = [c.model_dump() for c in payload.evaluation_criteria]
    if "submission_mode" in updates:
        updates["submission_mode"] = (
            payload.submission_mode.value
            if hasattr(payload.submission_mode, "value")
            else str(payload.submission_mode or "file_or_link")
        )
    if "allowed_mime_types" in updates:
        updates["allowed_mime_types"] = list(payload.allowed_mime_types or _default_round_allowed_mime_types())
    if "panel_team_distribution_mode" in updates:
        updates["panel_team_distribution_mode"] = _normalize_panel_distribution_mode(
            payload.panel_team_distribution_mode.value
            if hasattr(payload.panel_team_distribution_mode, "value")
            else payload.panel_team_distribution_mode
        )

    if requested_round_no is not None:
        next_round_no = int(requested_round_no)
        current_round_no = int(round_row.round_no)
        if next_round_no != current_round_no:
            conflict_round = (
                db.query(PersohubEventRound)
                .filter(
                    PersohubEventRound.event_id == event.id,
                    PersohubEventRound.round_no == next_round_no,
                    PersohubEventRound.id != round_row.id,
                )
                .first()
            )
            if conflict_round:
                # Force a two-step swap to avoid uniqueness conflicts during flush.
                conflict_round.round_no = -1
                db.flush()
                round_row.round_no = next_round_no
                db.flush()
                conflict_round.round_no = current_round_no
            else:
                round_row.round_no = next_round_no

    for field, value in updates.items():
        setattr(round_row, field, value)

    panel_mode_toggled = (
        "panel_mode_enabled" in updates
        and bool(round_row.panel_mode_enabled) != panel_mode_before
    )
    if panel_mode_toggled:
        _recompute_round_normalized_scores(db, event, round_row)

    should_apply_shortlisting = (
        round_row.is_frozen
        and round_row.elimination_type
        and round_row.elimination_value is not None
        and (
            "elimination_type" in updates
            or "elimination_value" in updates
            or eliminate_absent
        )
    )
    if should_apply_shortlisting:
        entity_type = PersohubEventEntityType.USER if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else PersohubEventEntityType.TEAM
        active_regs = (
            db.query(PersohubEventRegistration)
            .filter(
                PersohubEventRegistration.event_id == event.id,
                PersohubEventRegistration.entity_type == entity_type,
                PersohubEventRegistration.status == PersohubEventRegistrationStatus.ACTIVE,
            )
            .all()
        )

        score_rows = (
            db.query(PersohubEventScore)
            .filter(
                PersohubEventScore.event_id == event.id,
                PersohubEventScore.round_id == round_row.id,
                PersohubEventScore.entity_type == entity_type,
            )
            .all()
        )
        score_map = {}
        for score_row in score_rows:
            score_key = int(score_row.user_id) if entity_type == PersohubEventEntityType.USER else int(score_row.team_id)
            score_map[score_key] = score_row

        shortlist_regs = []
        for reg in active_regs:
            reg_entity_id = int(reg.user_id) if entity_type == PersohubEventEntityType.USER else int(reg.team_id)
            target_score = score_map.get(reg_entity_id)
            absent = (target_score is None) or (not bool(target_score.is_present))
            if eliminate_absent and absent:
                reg.status = PersohubEventRegistrationStatus.ELIMINATED
                continue
            shortlist_regs.append(reg)

        totals = []
        for reg in shortlist_regs:
            entity_id = int(reg.user_id) if entity_type == PersohubEventEntityType.USER else int(reg.team_id)
            if entity_type == PersohubEventEntityType.USER:
                total = db.execute(
                    text(
                        """
                        SELECT COALESCE(SUM(normalized_score), 0)
                        FROM persohub_event_scores
                        WHERE event_id = :event_id AND entity_type = 'USER' AND user_id = :entity_id
                        """
                    ),
                    {"event_id": event.id, "entity_id": entity_id},
                ).scalar() or 0.0
            else:
                total = db.execute(
                    text(
                        """
                        SELECT COALESCE(SUM(total_score), 0)
                        FROM persohub_event_scores
                        WHERE event_id = :event_id
                          AND entity_type = 'TEAM'
                          AND team_id = :entity_id
                        """
                    ),
                    {"event_id": event.id, "entity_id": entity_id},
                ).scalar() or 0.0
            totals.append((reg, float(total), entity_id))

        totals.sort(key=lambda item: (-item[1], item[2]))
        elimination_type = str(round_row.elimination_type).strip().lower()
        if elimination_type == "top_k":
            cutoff = max(0, int(round_row.elimination_value))
            for idx, (reg, _, _) in enumerate(totals):
                reg.status = PersohubEventRegistrationStatus.ACTIVE if idx < cutoff else PersohubEventRegistrationStatus.ELIMINATED
        elif elimination_type == "min_score":
            threshold = float(round_row.elimination_value)
            for reg, score, _ in totals:
                reg.status = PersohubEventRegistrationStatus.ACTIVE if score >= threshold else PersohubEventRegistrationStatus.ELIMINATED
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid elimination type")

        round_row.state = PersohubEventRoundState.COMPLETED
    db.commit()
    db.refresh(round_row)
    shortlist_audit_meta = {}
    if should_apply_shortlisting:
        shortlist_audit_meta = _upload_round_audit_snapshot(
            db=db,
            event=event,
            round_row=round_row,
            admin=admin,
            audit_type="shortlisting_snapshot",
            folder="shortlisting",
            extra_meta={
                "shortlist_elimination_type": str(round_row.elimination_type or ""),
                "shortlist_elimination_value": float(round_row.elimination_value or 0.0),
                "shortlist_eliminate_absent": bool(eliminate_absent),
            },
        )
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_event_round",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}",
        meta={
            "round_id": round_id,
            "elimination_type": round_row.elimination_type,
            "elimination_value": round_row.elimination_value,
            "eliminate_absent": eliminate_absent,
            **shortlist_audit_meta,
        },
    )
    return PersohubManagedRoundResponse.model_validate(round_row)


@router.delete("/persohub/admin/persohub-events/{slug}/rounds/{round_id}")
def delete_round(
    slug: str,
    round_id: int,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    if round_row.state != PersohubEventRoundState.DRAFT:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only draft rounds can be deleted")
    db.query(PersohubEventRoundSubmission).filter(
        PersohubEventRoundSubmission.event_id == event.id,
        PersohubEventRoundSubmission.round_id == round_row.id,
    ).delete(synchronize_session=False)
    db.delete(round_row)
    _sync_event_round_count(db, event.id)
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "delete_persohub_event_round",
        method="DELETE",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}",
        meta={"round_id": round_id},
    )
    return {"message": "Round deleted"}


def _event_admin_options(db: Session, event: PersohubEvent) -> List[PersohubRoundPanelAdminOption]:
    joined = (
        db.query(PdaAdmin, PdaUser)
        .join(PdaUser, PdaAdmin.user_id == PdaUser.id)
        .order_by(PdaUser.name.asc(), PdaUser.regno.asc())
        .all()
    )
    options: List[PersohubRoundPanelAdminOption] = []
    for admin_row, user in joined:
        user_regno = str(user.regno or "").strip()
        if user_regno == "0000000000":
            continue
        policy = admin_row.policy if isinstance(admin_row.policy, dict) else {}
        options.append(
            PersohubRoundPanelAdminOption(
                admin_user_id=int(user.id),
                regno=user_regno,
                name=str(user.name or user.regno or f"User {user.id}"),
                email=str(user.email or "").strip() or None,
            )
        )
    return options


def _build_round_panel_list_response(
    db: Session,
    event: PersohubEvent,
    round_row: PersohubEventRound,
    admin: PdaUser,
) -> PersohubRoundPanelListResponse:
    is_superadmin = _is_superadmin_user(db, admin)
    panels = (
        db.query(PersohubEventRoundPanel)
        .filter(
            PersohubEventRoundPanel.event_id == event.id,
            PersohubEventRoundPanel.round_id == round_row.id,
        )
        .order_by(PersohubEventRoundPanel.panel_no.asc(), PersohubEventRoundPanel.id.asc())
        .all()
    )
    panel_ids = [int(panel.id) for panel in panels]
    members = (
        db.query(PersohubEventRoundPanelMember, PdaUser, PdaAdmin)
        .join(PdaUser, PersohubEventRoundPanelMember.admin_user_id == PdaUser.id)
        .outerjoin(PdaAdmin, PdaAdmin.user_id == PdaUser.id)
        .filter(
            PersohubEventRoundPanelMember.event_id == event.id,
            PersohubEventRoundPanelMember.round_id == round_row.id,
            PersohubEventRoundPanelMember.panel_id.in_(panel_ids),
        )
        .all()
        if panel_ids
        else []
    )
    member_map: Dict[int, List[PersohubRoundPanelMemberResponse]] = {}
    for member_row, user, admin_row in members:
        user_regno = str(user.regno or "").strip()
        if user_regno == "0000000000":
            continue
        panel_id = int(member_row.panel_id)
        member_map.setdefault(panel_id, []).append(
            PersohubRoundPanelMemberResponse(
                admin_user_id=int(user.id),
                regno=user_regno,
                name=str(user.name or user.regno or f"User {user.id}"),
                email=str(user.email or "").strip() or None,
            )
        )
    for panel_id, member_rows in member_map.items():
        member_rows.sort(key=lambda row: (row.name.lower(), row.regno))

    assignment_count_rows = (
        db.query(
            PersohubEventRoundPanelAssignment.panel_id.label("panel_id"),
            func.count(PersohubEventRoundPanelAssignment.id).label("count"),
        )
        .filter(
            PersohubEventRoundPanelAssignment.event_id == event.id,
            PersohubEventRoundPanelAssignment.round_id == round_row.id,
            PersohubEventRoundPanelAssignment.panel_id.in_(panel_ids),
        )
        .group_by(PersohubEventRoundPanelAssignment.panel_id)
        .all()
        if panel_ids
        else []
    )
    assignment_count_map = {
        int(row.panel_id): int(row.count or 0)
        for row in assignment_count_rows
        if row.panel_id is not None
    }
    my_panel_ids = {
        int(value.panel_id)
        for value in db.query(PersohubEventRoundPanelMember.panel_id).filter(
            PersohubEventRoundPanelMember.event_id == event.id,
            PersohubEventRoundPanelMember.round_id == round_row.id,
            PersohubEventRoundPanelMember.admin_user_id == admin.id,
        ).all()
        if value.panel_id is not None
    }
    panel_rows = [
        PersohubRoundPanelResponse(
            id=int(panel.id),
            event_id=int(panel.event_id),
            round_id=int(panel.round_id),
            panel_no=int(panel.panel_no),
            panel_name=str(panel.name or "").strip() or None,
            panel_link=str(panel.panel_link or "").strip() or None,
            panel_time=panel.panel_time,
            instructions=str(panel.instructions or "").strip() or None,
            members=member_map.get(int(panel.id), []),
            assignment_count=int(assignment_count_map.get(int(panel.id), 0)),
        )
        for panel in panels
    ]
    return PersohubRoundPanelListResponse(
        panel_mode_enabled=bool(round_row.panel_mode_enabled),
        panel_team_distribution_mode=_normalize_panel_distribution_mode(round_row.panel_team_distribution_mode),
        panel_structure_locked=bool(round_row.panel_structure_locked),
        current_admin_is_superadmin=is_superadmin,
        my_panel_ids=sorted(my_panel_ids),
        available_admins=_event_admin_options(db, event),
        panels=panel_rows,
    )


def _send_round_panel_email_background(
    recipients: List[Tuple[str, Dict[str, object]]],
    subject: str,
    html: str,
    text: Optional[str],
    admin_id: int,
    event_id: int,
    event_slug: str,
    round_id: int,
    request_method: Optional[str],
    request_path: Optional[str],
) -> None:
    db = SessionLocal()
    sent = 0
    failed = 0
    errors: List[Dict[str, str]] = []
    try:
        for email_value, context in recipients:
            try:
                rendered_html = render_email_template(html, context, html_mode=True)
                rendered_text = render_email_template(text, context, html_mode=False) if text else None
                text_content = rendered_text if rendered_text is not None else derive_text_from_html(rendered_html)
                send_bulk_email(email_value, subject, rendered_html, text_content)
                sent += 1
            except Exception as exc:
                failed += 1
                if len(errors) < 10:
                    errors.append({"email": email_value, "error": str(exc)})
        admin = db.query(PdaUser).filter(PdaUser.id == admin_id).first()
        event = db.query(PersohubEvent).filter(PersohubEvent.id == event_id).first()
        if admin and event:
            _log_event_admin_action(
                db,
                admin,
                event,
                "send_persohub_event_round_panel_email",
                method=request_method or "POST",
                path=request_path or f"/persohub/admin/persohub-events/{event_slug}/rounds/{round_id}/panels/email",
                meta={
                    "round_id": round_id,
                    "queued": len(recipients),
                    "sent": sent,
                    "failed": failed,
                    "errors": errors,
                },
            )
    finally:
        db.close()


@router.get("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels", response_model=PersohubRoundPanelListResponse)
def get_round_panels(
    slug: str,
    round_id: int,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    return _build_round_panel_list_response(db, event, round_row, admin)


@router.put("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels", response_model=PersohubRoundPanelListResponse)
def update_round_panels(
    slug: str,
    round_id: int,
    payload: PersohubRoundPanelsUpdateRequest,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    panel_defs = list(payload.panels or [])
    panel_nos = [int(item.panel_no) for item in panel_defs]
    if len(panel_nos) != len(set(panel_nos)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="panel_no must be unique")
    panel_structure_locked = bool(round_row.panel_structure_locked)

    existing_panels = (
        db.query(PersohubEventRoundPanel)
        .filter(
            PersohubEventRoundPanel.event_id == event.id,
            PersohubEventRoundPanel.round_id == round_row.id,
        )
        .all()
    )
    existing_by_id = {int(panel.id): panel for panel in existing_panels}
    existing_panel_ids = set(existing_by_id.keys())
    seen_existing_panel_ids: Set[int] = set()
    kept_existing_panel_ids: Set[int] = set()
    panel_member_targets: Dict[int, Set[int]] = {}
    pending_new_panels: List[Dict[str, object]] = []
    target_admin_ids: Set[int] = set()

    for panel_def in panel_defs:
        try:
            target_panel_no = int(panel_def.panel_no)
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="panel_no is required and must be a positive integer")
        if target_panel_no < 1:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="panel_no must be >= 1")

        panel_name = str(panel_def.panel_name or "").strip() or None
        panel_link = str(panel_def.panel_link or "").strip() or None
        panel_time = panel_def.panel_time
        instructions = str(panel_def.instructions or "").strip() or None
        member_admin_user_ids = {int(user_id) for user_id in (panel_def.member_admin_user_ids or [])}
        target_admin_ids.update(member_admin_user_ids)

        if panel_def.id is not None:
            panel_id = int(panel_def.id)
            if panel_id in seen_existing_panel_ids:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Duplicate panel id in payload: {panel_id}")
            seen_existing_panel_ids.add(panel_id)
            panel_row = existing_by_id.get(panel_id)
            if not panel_row:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Panel {panel_def.id} not found")
            if target_panel_no != int(panel_row.panel_no):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="panel_no cannot be changed")
            panel_row.name = panel_name
            panel_row.panel_link = panel_link
            panel_row.panel_time = panel_time
            panel_row.instructions = instructions
            kept_existing_panel_ids.add(int(panel_row.id))
            panel_member_targets[int(panel_row.id)] = member_admin_user_ids
            continue
        if panel_structure_locked:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Panel structure is locked; cannot add new panels")

        pending_new_panels.append(
            {
                "panel_no": target_panel_no,
                "name": panel_name,
                "panel_link": panel_link,
                "panel_time": panel_time,
                "instructions": instructions,
                "member_admin_user_ids": member_admin_user_ids,
            }
        )

    if panel_structure_locked:
        missing_panel_ids = sorted(existing_panel_ids - seen_existing_panel_ids)
        if missing_panel_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Panel structure is locked; missing panel ids in payload: {missing_panel_ids}",
            )

    allowed_admin_ids = {
        int(option.admin_user_id)
        for option in _event_admin_options(db, event)
    }
    invalid_admin_ids = sorted([value for value in target_admin_ids if value not in allowed_admin_ids])
    if invalid_admin_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid admin_user_id for this event: {invalid_admin_ids}",
        )

    removable_panel_ids = [panel_id for panel_id in existing_by_id.keys() if panel_id not in kept_existing_panel_ids]
    if panel_structure_locked and removable_panel_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Panel structure is locked; cannot delete panels")
    if removable_panel_ids:
        db.query(PersohubEventRoundPanelAssignment).filter(
            PersohubEventRoundPanelAssignment.event_id == event.id,
            PersohubEventRoundPanelAssignment.round_id == round_row.id,
            PersohubEventRoundPanelAssignment.panel_id.in_(removable_panel_ids),
        ).delete(synchronize_session=False)
        db.query(PersohubEventRoundPanelMember).filter(
            PersohubEventRoundPanelMember.event_id == event.id,
            PersohubEventRoundPanelMember.round_id == round_row.id,
            PersohubEventRoundPanelMember.panel_id.in_(removable_panel_ids),
        ).delete(synchronize_session=False)
        db.query(PersohubEventRoundPanel).filter(
            PersohubEventRoundPanel.event_id == event.id,
            PersohubEventRoundPanel.round_id == round_row.id,
            PersohubEventRoundPanel.id.in_(removable_panel_ids),
        ).delete(synchronize_session=False)

    for pending in pending_new_panels:
        panel_row = PersohubEventRoundPanel(
            event_id=event.id,
            round_id=round_row.id,
            panel_no=int(pending.get("panel_no") or 0),
            name=pending.get("name"),
            panel_link=pending.get("panel_link"),
            panel_time=pending.get("panel_time"),
            instructions=pending.get("instructions"),
        )
        db.add(panel_row)
        db.flush()
        panel_member_targets[int(panel_row.id)] = set(pending.get("member_admin_user_ids") or [])

    for panel_id, member_ids in panel_member_targets.items():
        db.query(PersohubEventRoundPanelMember).filter(
            PersohubEventRoundPanelMember.event_id == event.id,
            PersohubEventRoundPanelMember.round_id == round_row.id,
            PersohubEventRoundPanelMember.panel_id == panel_id,
            ~PersohubEventRoundPanelMember.admin_user_id.in_(member_ids or {-1}),
        ).delete(synchronize_session=False)
        existing_members = {
            int(row.admin_user_id)
            for row in db.query(PersohubEventRoundPanelMember.admin_user_id).filter(
                PersohubEventRoundPanelMember.event_id == event.id,
                PersohubEventRoundPanelMember.round_id == round_row.id,
                PersohubEventRoundPanelMember.panel_id == panel_id,
            ).all()
            if row.admin_user_id is not None
        }
        for admin_user_id in sorted(member_ids):
            if admin_user_id in existing_members:
                continue
            db.add(
                PersohubEventRoundPanelMember(
                    event_id=event.id,
                    round_id=round_row.id,
                    panel_id=panel_id,
                    admin_user_id=admin_user_id,
                )
            )

    if bool(round_row.panel_mode_enabled):
        _recompute_round_normalized_scores(db, event, round_row)
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_event_round_panels",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels",
        meta={"round_id": round_id, "panel_count": len(panel_defs)},
    )
    return _build_round_panel_list_response(db, event, round_row, admin)


@router.post("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels/auto-assign")
def auto_assign_round_panels(
    slug: str,
    round_id: int,
    payload: PersohubRoundPanelsAutoAssignRequest,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    import random

    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    if not bool(round_row.panel_mode_enabled):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Enable panel mode for this round first")

    panels = (
        db.query(PersohubEventRoundPanel)
        .filter(
            PersohubEventRoundPanel.event_id == event.id,
            PersohubEventRoundPanel.round_id == round_row.id,
        )
        .order_by(PersohubEventRoundPanel.panel_no.asc(), PersohubEventRoundPanel.id.asc())
        .all()
    )
    if not panels:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Configure at least one panel before auto-assign")

    entities = [item for item in _round_scoring_entities(db, event, round_row) if _status_is_active(item.get("status"))]
    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        entity_type_key = "user"
        entity_ids = [int(item.get("entity_id")) for item in entities if item.get("entity_id") is not None]
        score_rows = (
            db.query(
                PersohubEventScore.user_id.label("entity_id"),
                func.coalesce(func.sum(PersohubEventScore.normalized_score), 0.0).label("score"),
            )
            .filter(
                PersohubEventScore.event_id == event.id,
                PersohubEventScore.entity_type == PersohubEventEntityType.USER,
                PersohubEventScore.user_id.in_(entity_ids),
            )
            .group_by(PersohubEventScore.user_id)
            .all()
            if entity_ids
            else []
        )
        score_map = {int(row.entity_id): float(row.score or 0.0) for row in score_rows if row.entity_id is not None}
        weighted_member_map = {int(entity_id): 1 for entity_id in entity_ids}
    else:
        entity_type_key = "team"
        entity_ids = [int(item.get("entity_id")) for item in entities if item.get("entity_id") is not None]
        score_rows = (
            db.query(
                PersohubEventScore.team_id.label("entity_id"),
                func.coalesce(func.sum(PersohubEventScore.total_score), 0.0).label("score"),
            )
            .filter(
                PersohubEventScore.event_id == event.id,
                PersohubEventScore.entity_type == PersohubEventEntityType.TEAM,
                PersohubEventScore.team_id.in_(entity_ids),
            )
            .group_by(PersohubEventScore.team_id)
            .all()
            if entity_ids
            else []
        )
        score_map = {int(row.entity_id): float(row.score or 0.0) for row in score_rows if row.entity_id is not None}
        member_count_rows = (
            db.query(PersohubEventTeamMember.team_id, func.count(PersohubEventTeamMember.id))
            .filter(PersohubEventTeamMember.team_id.in_(entity_ids))
            .group_by(PersohubEventTeamMember.team_id)
            .all()
            if entity_ids
            else []
        )
        weighted_member_map = {
            int(team_id): max(1, int(count or 0))
            for team_id, count in member_count_rows
        }
        for entity_id in entity_ids:
            weighted_member_map.setdefault(int(entity_id), 1)

    assignment_rows = (
        db.query(PersohubEventRoundPanelAssignment)
        .filter(
            PersohubEventRoundPanelAssignment.event_id == event.id,
            PersohubEventRoundPanelAssignment.round_id == round_row.id,
            PersohubEventRoundPanelAssignment.entity_type == (
                PersohubEventEntityType.USER if entity_type_key == "user" else PersohubEventEntityType.TEAM
            ),
        )
        .all()
    )
    existing_assignment_map = {}
    for row in assignment_rows:
        if entity_type_key == "user" and row.user_id is not None:
            existing_assignment_map[int(row.user_id)] = row
        elif entity_type_key == "team" and row.team_id is not None:
            existing_assignment_map[int(row.team_id)] = row

    candidate_rows = []
    for entity in entities:
        entity_id = int(entity.get("entity_id"))
        if payload.include_unassigned_only and entity_id in existing_assignment_map:
            continue
        candidate_rows.append(
            {
                "entity_id": entity_id,
                "score": float(score_map.get(entity_id, 0.0)),
                "members_count": int(weighted_member_map.get(entity_id, 1)),
            }
        )
    if not candidate_rows:
        return {
            "assigned_count": 0,
            "panel_count": len(panels),
            "distribution_mode": _normalize_panel_distribution_mode(round_row.panel_team_distribution_mode),
        }

    distribution_mode = _normalize_panel_distribution_mode(round_row.panel_team_distribution_mode)
    weighted_mode = (
        event.participant_mode == PersohubEventParticipantMode.TEAM
        and distribution_mode == "member_count_weighted"
    )

    panel_state = {
        int(panel.id): {"score_sum": 0.0, "entity_count": 0, "members_sum": 0}
        for panel in panels
    }
    panel_ids = [int(panel.id) for panel in panels]
    candidate_signature = "|".join(
        f"{int(item['entity_id'])}:{round(float(item['score']), 6):.6f}:{int(item['members_count'])}"
        for item in sorted(candidate_rows, key=lambda value: int(value["entity_id"]))
    )
    seed_material = (
        f"event:{int(event.id)}|round:{int(round_row.id)}|entity:{entity_type_key}"
        f"|mode:{distribution_mode}|weighted:{int(weighted_mode)}"
        f"|only_unassigned:{int(bool(payload.include_unassigned_only))}"
        f"|panels:{','.join(str(value) for value in panel_ids)}"
        f"|candidates:{candidate_signature}"
    )
    seed_int = int(hashlib.sha256(seed_material.encode("utf-8")).hexdigest()[:16], 16)
    rng = random.Random(seed_int)

    bucket_map: Dict[float, List[dict]] = {}
    for item in candidate_rows:
        bucket_key = round(float(item["score"]), 6)
        bucket_map.setdefault(bucket_key, []).append(item)
    sorted_buckets = sorted(bucket_map.keys(), reverse=True)

    assignment_targets = []
    for bucket_score in sorted_buckets:
        bucket_items = sorted(bucket_map[bucket_score], key=lambda value: int(value["entity_id"]))
        rng.shuffle(bucket_items)
        for item in bucket_items:
            scoring_keys = {}
            for panel_id in panel_ids:
                state = panel_state[panel_id]
                load_value = state["members_sum"] if weighted_mode else state["entity_count"]
                scoring_keys[panel_id] = (float(state["score_sum"]), int(load_value))
            min_key = min(scoring_keys.values())
            candidate_panel_ids = [panel_id for panel_id in panel_ids if scoring_keys[panel_id] == min_key]
            selected_panel_id = int(rng.choice(candidate_panel_ids))
            state = panel_state[selected_panel_id]
            state["score_sum"] += float(item["score"])
            state["entity_count"] += 1
            state["members_sum"] += int(item["members_count"])
            assignment_targets.append((int(item["entity_id"]), selected_panel_id))

    created = 0
    updated = 0
    for entity_id, panel_id in assignment_targets:
        existing_row = existing_assignment_map.get(entity_id)
        if existing_row:
            if int(existing_row.panel_id or 0) != panel_id:
                existing_row.panel_id = panel_id
                existing_row.assigned_by_user_id = admin.id
                updated += 1
            continue
        db.add(
            PersohubEventRoundPanelAssignment(
                event_id=event.id,
                round_id=round_row.id,
                panel_id=panel_id,
                entity_type=PersohubEventEntityType.USER if entity_type_key == "user" else PersohubEventEntityType.TEAM,
                user_id=entity_id if entity_type_key == "user" else None,
                team_id=entity_id if entity_type_key == "team" else None,
                assigned_by_user_id=admin.id,
            )
        )
        created += 1

    if bool(round_row.panel_mode_enabled):
        _recompute_round_normalized_scores(db, event, round_row)
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "auto_assign_persohub_event_round_panels",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels/auto-assign",
        meta={
            "round_id": round_id,
            "assigned_count": len(assignment_targets),
            "created": created,
            "updated": updated,
            "distribution_mode": distribution_mode,
            "include_unassigned_only": bool(payload.include_unassigned_only),
        },
    )
    return {
        "assigned_count": len(assignment_targets),
        "created": created,
        "updated": updated,
        "panel_count": len(panels),
        "distribution_mode": distribution_mode,
    }


@router.put("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels/assignments")
def update_round_panel_assignments(
    slug: str,
    round_id: int,
    payload: PersohubRoundPanelAssignmentsUpdateRequest,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    panel_ids = {
        int(value.id)
        for value in db.query(PersohubEventRoundPanel.id).filter(
            PersohubEventRoundPanel.event_id == event.id,
            PersohubEventRoundPanel.round_id == round_row.id,
        ).all()
    }
    if not panel_ids and payload.assignments:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Configure panels before assignments")

    entities = _round_scoring_entities(db, event, round_row)
    valid_entities = {
        _panel_entity_key(item.get("entity_type"), int(item.get("entity_id")))
        for item in entities
        if item.get("entity_id") is not None
    }
    existing_rows = (
        db.query(PersohubEventRoundPanelAssignment)
        .filter(
            PersohubEventRoundPanelAssignment.event_id == event.id,
            PersohubEventRoundPanelAssignment.round_id == round_row.id,
        )
        .all()
    )
    existing_map: Dict[Tuple[str, int], PersohubEventRoundPanelAssignment] = {}
    for row in existing_rows:
        if row.entity_type == PersohubEventEntityType.USER and row.user_id is not None:
            existing_map[_panel_entity_key("user", int(row.user_id))] = row
        if row.entity_type == PersohubEventEntityType.TEAM and row.team_id is not None:
            existing_map[_panel_entity_key("team", int(row.team_id))] = row

    updated = 0
    removed = 0
    created = 0
    for item in payload.assignments or []:
        entity_type = str(item.entity_type.value if hasattr(item.entity_type, "value") else item.entity_type).strip().lower()
        entity_id = int(item.entity_id)
        key = _panel_entity_key(entity_type, entity_id)
        if key not in valid_entities:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid scoring entity for assignment: {entity_type} {entity_id}",
            )
        existing_row = existing_map.get(key)
        panel_id = int(item.panel_id) if item.panel_id is not None else None
        if panel_id is None:
            if existing_row is not None:
                db.delete(existing_row)
                removed += 1
            continue
        if panel_id not in panel_ids:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid panel_id: {panel_id}")
        if existing_row is not None:
            if int(existing_row.panel_id or 0) != panel_id:
                existing_row.panel_id = panel_id
                existing_row.assigned_by_user_id = admin.id
                updated += 1
            continue
        db.add(
            PersohubEventRoundPanelAssignment(
                event_id=event.id,
                round_id=round_row.id,
                panel_id=panel_id,
                entity_type=PersohubEventEntityType.USER if entity_type == "user" else PersohubEventEntityType.TEAM,
                user_id=entity_id if entity_type == "user" else None,
                team_id=entity_id if entity_type == "team" else None,
                assigned_by_user_id=admin.id,
            )
        )
        created += 1
    if bool(round_row.panel_mode_enabled):
        _recompute_round_normalized_scores(db, event, round_row)
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "update_persohub_event_round_panel_assignments",
        method="PUT",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels/assignments",
        meta={"round_id": round_id, "created": created, "updated": updated, "removed": removed},
    )
    return {"created": created, "updated": updated, "removed": removed}


@router.post("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels/email")
def send_round_panel_email(
    slug: str,
    round_id: int,
    payload: PersohubRoundPanelEmailRequest,
    background_tasks: BackgroundTasks,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
    request: Request = None,
):
    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    subject = str(payload.subject or "").strip()
    html = str(payload.html or "").strip()
    if not subject or not html:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Subject and HTML are required")

    panel_query = db.query(PersohubEventRoundPanel).filter(
        PersohubEventRoundPanel.event_id == event.id,
        PersohubEventRoundPanel.round_id == round_row.id,
    )
    if payload.panel_ids:
        unique_panel_ids = sorted({int(value) for value in payload.panel_ids})
        panel_query = panel_query.filter(PersohubEventRoundPanel.id.in_(unique_panel_ids))
    panels = panel_query.order_by(PersohubEventRoundPanel.panel_no.asc()).all()
    if not panels:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No panel rows found for email")
    panel_ids = [int(panel.id) for panel in panels]
    panel_map = {int(panel.id): panel for panel in panels}

    member_rows = (
        db.query(PersohubEventRoundPanelMember, PdaUser)
        .join(PdaUser, PersohubEventRoundPanelMember.admin_user_id == PdaUser.id)
        .filter(
            PersohubEventRoundPanelMember.event_id == event.id,
            PersohubEventRoundPanelMember.round_id == round_row.id,
            PersohubEventRoundPanelMember.panel_id.in_(panel_ids),
        )
        .all()
    )
    recipients: List[Tuple[str, Dict[str, object]]] = []
    for member_row, user in member_rows:
        email_value = str(user.email or "").strip().lower()
        if not email_value:
            continue
        panel = panel_map.get(int(member_row.panel_id))
        if not panel:
            continue
        context = {
            "name": user.name,
            "profile_name": user.profile_name,
            "regno": user.regno,
            "email": user.email,
            "dept": user.dept,
            "gender": user.gender,
            "phno": user.phno,
            "dob": user.dob,
            "photo_url": user.image_url,
            "is_member": bool(user.is_member),
            "email_verified": bool(user.email_verified_at),
            "created_at": user.created_at,
            "updated_at": user.updated_at,
            "batch": extract_batch(user.regno),
            "event_title": event.title,
            "event_code": event.event_code,
            "round_name": round_row.name,
            "round_no": round_row.round_no,
            "panel_no": panel.panel_no,
            "panel_name": panel.name or f"Panel {panel.panel_no}",
            "panel_link": panel.panel_link,
            "panel_time": panel.panel_time.isoformat() if panel.panel_time else "",
            "panel_instructions": panel.instructions or "",
        }
        recipients.append((email_value, context))
    if not recipients:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No panel members with valid email found")

    background_tasks.add_task(
        _send_round_panel_email_background,
        recipients,
        subject,
        html,
        payload.text,
        admin.id,
        event.id,
        event.slug,
        round_row.id,
        request.method if request else "POST",
        request.url.path if request else f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/panels/email",
    )
    return {
        "queued": len(recipients),
        "panel_count": len(panel_ids),
    }


@router.get("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/stats")
def round_stats(
    slug: str,
    round_id: int,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    entity_type = PersohubEventEntityType.USER if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else PersohubEventEntityType.TEAM

    total_count = db.query(PersohubEventRegistration).filter(
        PersohubEventRegistration.event_id == event.id,
        PersohubEventRegistration.entity_type == entity_type,
    ).count()
    score_rows = (
        db.query(PersohubEventScore)
        .filter(
            PersohubEventScore.event_id == event.id,
            PersohubEventScore.round_id == round_id,
            PersohubEventScore.entity_type == entity_type,
        )
        .all()
    )
    present_rows = [row for row in score_rows if bool(row.is_present)]
    present_count = len(present_rows)
    absent_count = max(total_count - present_count, 0)
    present_scores = [float(row.normalized_score or 0.0) for row in present_rows]

    if entity_type == PersohubEventEntityType.USER:
        entity_ids = [int(row.user_id) for row in present_rows if row.user_id is not None]
        names = {
            int(user.id): user.name
            for user in db.query(PdaUser).filter(PdaUser.id.in_(entity_ids)).all()
        } if entity_ids else {}
        sortable_rows = sorted(
            present_rows,
            key=lambda row: (-float(row.normalized_score or 0.0), int(row.user_id or 0)),
        )
        top10 = [
            {
                "entity_id": int(row.user_id),
                "name": names.get(int(row.user_id), f"User {int(row.user_id)}"),
                "normalized_score": float(row.normalized_score or 0.0),
            }
            for row in sortable_rows[:10]
            if row.user_id is not None
        ]
    else:
        entity_ids = [int(row.team_id) for row in present_rows if row.team_id is not None]
        names = {
            int(team.id): team.team_name
            for team in db.query(PersohubEventTeam).filter(PersohubEventTeam.id.in_(entity_ids)).all()
        } if entity_ids else {}
        sortable_rows = sorted(
            present_rows,
            key=lambda row: (-float(row.normalized_score or 0.0), int(row.team_id or 0)),
        )
        top10 = [
            {
                "entity_id": int(row.team_id),
                "name": names.get(int(row.team_id), f"Team {int(row.team_id)}"),
                "normalized_score": float(row.normalized_score or 0.0),
            }
            for row in sortable_rows[:10]
            if row.team_id is not None
        ]

    return {
        "total_count": total_count,
        "present_count": present_count,
        "absent_count": absent_count,
        "min_score": min(present_scores) if present_scores else None,
        "max_score": max(present_scores) if present_scores else None,
        "avg_score": (sum(present_scores) / len(present_scores)) if present_scores else None,
        "top10": top10,
    }


@router.get("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/participants")
def round_participants(
    slug: str,
    round_id: int,
    search: Optional[str] = None,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    scope = _round_admin_panel_scope(db, round_row, admin)
    panel_map, assignment_map = _round_panel_maps(db, round_row) if bool(round_row.panel_mode_enabled) else ({}, {})
    entities = _round_scoring_entities(db, event, round_row)
    if search:
        needle = str(search).strip().lower()
        entities = [
            item
            for item in entities
            if needle in " ".join(
                [
                    str(item.get("name") or ""),
                    str(item.get("regno_or_code") or ""),
                    str(item.get("email") or ""),
                ]
            ).lower()
        ]
    score_rows = db.query(PersohubEventScore).filter(PersohubEventScore.event_id == event.id, PersohubEventScore.round_id == round_id).all()
    score_map = {}
    for row in score_rows:
        key = ("user", row.user_id) if row.user_id else ("team", row.team_id)
        score_map[key] = row
    attendance_rows = db.query(PersohubEventAttendance).filter(
        PersohubEventAttendance.event_id == event.id,
        PersohubEventAttendance.round_id == round_id,
    ).all()
    attendance_map = {}
    for row in attendance_rows:
        key = ("user", row.user_id) if row.user_id else ("team", row.team_id)
        attendance_map[key] = row
    submission_rows = db.query(PersohubEventRoundSubmission).filter(
        PersohubEventRoundSubmission.event_id == event.id,
        PersohubEventRoundSubmission.round_id == round_id,
    ).all()
    submission_map = {}
    for row in submission_rows:
        key = ("user", row.user_id) if row.user_id else ("team", row.team_id)
        submission_map[key] = row
    result = []
    for entity in entities:
        key = (entity["entity_type"], entity["entity_id"])
        row = score_map.get(key)
        attendance_row = attendance_map.get(key)
        submission_row = submission_map.get(key)
        panel_assignment = assignment_map.get(_panel_entity_key(entity.get("entity_type"), int(entity.get("entity_id"))))
        panel_row = panel_map.get(int(panel_assignment.panel_id)) if panel_assignment and panel_assignment.panel_id is not None else None
        panel_id = int(panel_assignment.panel_id) if panel_assignment and panel_assignment.panel_id is not None else None
        is_present = bool(attendance_row.is_present) if attendance_row else (bool(row.is_present) if row else False)
        payload = {
            **entity,
            "score_id": row.id if row else None,
            "criteria_scores": row.criteria_scores if row else {},
            "total_score": float(row.total_score or 0.0) if row else 0.0,
            "normalized_score": float(row.normalized_score or 0.0) if row else 0.0,
            "is_present": is_present,
            "submission_id": submission_row.id if submission_row else None,
            "submission_type": submission_row.submission_type if submission_row else None,
            "submission_file_url": submission_row.file_url if submission_row else None,
            "submission_link_url": submission_row.link_url if submission_row else None,
            "submission_is_locked": bool(submission_row.is_locked) if submission_row else False,
            "panel_id": panel_id,
            "panel_no": int(panel_row.panel_no) if panel_row and panel_row.panel_no is not None else None,
            "panel_name": (str(panel_row.name or "").strip() or None) if panel_row else None,
            "is_score_editable_by_current_admin": _is_entity_editable_by_admin(
                scope,
                str(entity.get("entity_type")),
                int(entity.get("entity_id")),
                panel_id,
            ),
        }
        if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
            payload.setdefault("participant_id", entity["entity_id"])
            payload.setdefault("participant_name", entity.get("name"))
            payload.setdefault("participant_register_number", entity.get("regno_or_code"))
            payload.setdefault("participant_status", entity.get("status"))
        result.append(payload)
    return result


@router.get("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/submissions", response_model=List[PersohubRoundSubmissionAdminListItem])
def round_submissions(
    slug: str,
    round_id: int,
    search: Optional[str] = None,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    if not bool(round_row.requires_submission):
        return []

    entities = _round_scoring_entities(db, event, round_row)
    if search:
        needle = str(search).strip().lower()
        entities = [
            item for item in entities
            if needle in " ".join(
                [
                    str(item.get("name") or ""),
                    str(item.get("regno_or_code") or ""),
                    str(item.get("email") or ""),
                ]
            ).lower()
        ]

    submissions = db.query(PersohubEventRoundSubmission).filter(
        PersohubEventRoundSubmission.event_id == event.id,
        PersohubEventRoundSubmission.round_id == round_row.id,
    ).all()
    submission_map = {}
    for row in submissions:
        key = ("user", row.user_id) if row.user_id else ("team", row.team_id)
        submission_map[key] = row

    result: List[PersohubRoundSubmissionAdminListItem] = []
    for entity in entities:
        key = (entity["entity_type"], entity["entity_id"])
        result.append(_round_submission_payload_for_admin(round_row, event, entity, submission_map.get(key)))
    return result


@router.put("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/submissions/{submission_id}", response_model=PersohubRoundSubmissionAdminListItem)
def update_round_submission_as_admin(
    slug: str,
    round_id: int,
    submission_id: int,
    payload: PersohubRoundSubmissionAdminUpdate,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    submission = db.query(PersohubEventRoundSubmission).filter(
        PersohubEventRoundSubmission.id == submission_id,
        PersohubEventRoundSubmission.event_id == event.id,
        PersohubEventRoundSubmission.round_id == round_row.id,
    ).first()
    if not submission:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Submission not found")

    updates = payload.model_dump(exclude_unset=True)
    if "submission_type" in updates:
        updates["submission_type"] = (
            payload.submission_type.value
            if hasattr(payload.submission_type, "value")
            else str(payload.submission_type)
        ).lower()
    if "submission_type" in updates:
        if updates["submission_type"] not in {"file", "link"}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="submission_type must be file or link")

    allowed_mime_types = list(round_row.allowed_mime_types or _default_round_allowed_mime_types())
    max_file_size_mb = int(round_row.max_file_size_mb or 25)
    max_bytes = max_file_size_mb * 1024 * 1024

    next_type = str(updates.get("submission_type") or submission.submission_type or "").lower()
    next_file_url = str(updates.get("file_url") if "file_url" in updates else (submission.file_url or "")).strip()
    next_link_url = str(updates.get("link_url") if "link_url" in updates else (submission.link_url or "")).strip()
    next_mime_type = str(updates.get("mime_type") if "mime_type" in updates else (submission.mime_type or "")).strip().lower()
    next_file_size_bytes = updates.get("file_size_bytes", submission.file_size_bytes)

    if next_type == "file":
        if not next_file_url:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="file_url is required for file submissions")
        if not next_mime_type:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="mime_type is required for file submissions")
        if next_mime_type not in allowed_mime_types:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid file type")
        if next_file_size_bytes is None or int(next_file_size_bytes) <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="file_size_bytes is required for file submissions")
        if int(next_file_size_bytes) > max_bytes:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"File size exceeds {max_file_size_mb} MB limit")
        updates.pop("link_url", None)
        submission.link_url = None
    elif next_type == "link":
        if not next_link_url:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="link_url is required for link submissions")
        updates.pop("file_url", None)
        updates.pop("file_name", None)
        updates.pop("file_size_bytes", None)
        updates.pop("mime_type", None)
        submission.file_url = None
        submission.file_name = None
        submission.file_size_bytes = None
        submission.mime_type = None

    for field, value in updates.items():
        setattr(submission, field, value)

    submission.version = int(submission.version or 0) + 1
    submission.updated_by_user_id = admin.id
    db.commit()
    db.refresh(submission)

    entity = {
        "entity_type": "team" if submission.entity_type == PersohubEventEntityType.TEAM else "user",
        "entity_id": submission.team_id if submission.entity_type == PersohubEventEntityType.TEAM else submission.user_id,
        "name": "Unknown",
        "regno_or_code": "-",
        "status": "Active",
    }
    if submission.entity_type == PersohubEventEntityType.USER and submission.user_id:
        joined = db.query(PersohubEventRegistration, PdaUser).join(
            PdaUser, PersohubEventRegistration.user_id == PdaUser.id
        ).filter(
            PersohubEventRegistration.event_id == event.id,
            PersohubEventRegistration.user_id == submission.user_id,
            PersohubEventRegistration.entity_type == PersohubEventEntityType.USER,
        ).first()
        if joined:
            reg, user_row = joined
            entity["name"] = user_row.name
            entity["regno_or_code"] = user_row.regno
            entity["status"] = _registration_status_label(reg.status)
    elif submission.entity_type == PersohubEventEntityType.TEAM and submission.team_id:
        joined = db.query(PersohubEventRegistration, PersohubEventTeam).join(
            PersohubEventTeam, PersohubEventRegistration.team_id == PersohubEventTeam.id
        ).filter(
            PersohubEventRegistration.event_id == event.id,
            PersohubEventRegistration.team_id == submission.team_id,
            PersohubEventRegistration.entity_type == PersohubEventEntityType.TEAM,
        ).first()
        if joined:
            reg, team_row = joined
            entity["name"] = team_row.team_name
            entity["regno_or_code"] = team_row.team_code
            entity["status"] = _registration_status_label(reg.status)
    return _round_submission_payload_for_admin(round_row, event, entity, submission)


@router.post("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/scores")
def save_scores(
    slug: str,
    round_id: int,
    entries: List[PersohubManagedScoreEntry],
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    if round_row.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is frozen")

    criteria = _criteria_def(round_row)
    criteria_max = {c["name"]: float(c.get("max_marks", 0) or 0) for c in criteria}
    max_total = sum(criteria_max.values()) if criteria_max else 100
    panel_assignment_map: Dict[Tuple[str, int], PersohubEventRoundPanelAssignment] = {}
    if bool(round_row.panel_mode_enabled):
        _, panel_assignment_map = _round_panel_maps(db, round_row)

    parsed_entries = []
    user_ids: Set[int] = set()
    team_ids: Set[int] = set()
    for entry in entries:
        payload = entry.model_dump()
        entity_type, user_id, team_id = _entity_from_payload(event, payload)
        entity_type_key = "user" if entity_type == PersohubEventEntityType.USER else "team"
        entity_id_value = int(user_id) if entity_type == PersohubEventEntityType.USER else int(team_id)
        if bool(round_row.panel_mode_enabled) and bool(entry.is_present):
            assignment_row = panel_assignment_map.get(_panel_entity_key(entity_type_key, entity_id_value))
            if assignment_row is None or assignment_row.panel_id is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Panel assignment required for present scoring in panel mode: {entity_type_key} {entity_id_value}",
                )
        parsed_entries.append((entry, entity_type, user_id, team_id))
        if entity_type == PersohubEventEntityType.USER and user_id is not None:
            user_ids.add(int(user_id))
        if entity_type == PersohubEventEntityType.TEAM and team_id is not None:
            team_ids.add(int(team_id))

    reg_user_map: Dict[int, PersohubEventRegistration] = {}
    reg_team_map: Dict[int, PersohubEventRegistration] = {}
    score_user_map: Dict[int, PersohubEventScore] = {}
    score_team_map: Dict[int, PersohubEventScore] = {}
    attendance_user_map: Dict[int, PersohubEventAttendance] = {}
    attendance_team_map: Dict[int, PersohubEventAttendance] = {}

    if user_ids:
        reg_user_map = {
            int(row.user_id): row
            for row in db.query(PersohubEventRegistration).filter(
                PersohubEventRegistration.event_id == event.id,
                PersohubEventRegistration.entity_type == PersohubEventEntityType.USER,
                PersohubEventRegistration.user_id.in_(user_ids),
            ).all()
            if row.user_id is not None
        }
        score_user_map = {
            int(row.user_id): row
            for row in db.query(PersohubEventScore).filter(
                PersohubEventScore.event_id == event.id,
                PersohubEventScore.round_id == round_id,
                PersohubEventScore.entity_type == PersohubEventEntityType.USER,
                PersohubEventScore.user_id.in_(user_ids),
            ).all()
            if row.user_id is not None
        }
        attendance_user_map = {
            int(row.user_id): row
            for row in db.query(PersohubEventAttendance).filter(
                PersohubEventAttendance.event_id == event.id,
                PersohubEventAttendance.round_id == round_id,
                PersohubEventAttendance.entity_type == PersohubEventEntityType.USER,
                PersohubEventAttendance.user_id.in_(user_ids),
            ).all()
            if row.user_id is not None
        }

    if team_ids:
        reg_team_map = {
            int(row.team_id): row
            for row in db.query(PersohubEventRegistration).filter(
                PersohubEventRegistration.event_id == event.id,
                PersohubEventRegistration.entity_type == PersohubEventEntityType.TEAM,
                PersohubEventRegistration.team_id.in_(team_ids),
            ).all()
            if row.team_id is not None
        }
        score_team_map = {
            int(row.team_id): row
            for row in db.query(PersohubEventScore).filter(
                PersohubEventScore.event_id == event.id,
                PersohubEventScore.round_id == round_id,
                PersohubEventScore.entity_type == PersohubEventEntityType.TEAM,
                PersohubEventScore.team_id.in_(team_ids),
            ).all()
            if row.team_id is not None
        }
        attendance_team_map = {
            int(row.team_id): row
            for row in db.query(PersohubEventAttendance).filter(
                PersohubEventAttendance.event_id == event.id,
                PersohubEventAttendance.round_id == round_id,
                PersohubEventAttendance.entity_type == PersohubEventEntityType.TEAM,
                PersohubEventAttendance.team_id.in_(team_ids),
            ).all()
            if row.team_id is not None
        }

    for entry, entity_type, user_id, team_id in parsed_entries:
        is_user = entity_type == PersohubEventEntityType.USER
        entity_id_value = int(user_id) if is_user else int(team_id)
        reg_row = reg_user_map.get(entity_id_value) if is_user else reg_team_map.get(entity_id_value)
        if not reg_row:
            label = "user_id" if is_user else "team_id"
            value = user_id if is_user else team_id
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Registration not found for {label}={value}")
        if reg_row.status == PersohubEventRegistrationStatus.ELIMINATED:
            label = "User" if is_user else "Team"
            value = user_id if is_user else team_id
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{label} {value} is eliminated")

        score_row = score_user_map.get(entity_id_value) if is_user else score_team_map.get(entity_id_value)
        attendance_row = attendance_user_map.get(entity_id_value) if is_user else attendance_team_map.get(entity_id_value)

        if not entry.is_present:
            safe_scores = {name: 0.0 for name in criteria_max.keys()}
            total = 0.0
            normalized = 0.0
        else:
            safe_scores = {}
            for name, max_marks in criteria_max.items():
                value = float((entry.criteria_scores or {}).get(name, 0.0))
                if value < 0 or value > max_marks:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Score for {name} must be between 0 and {max_marks}")
                safe_scores[name] = value
            total = float(sum(safe_scores.values()))
            normalized = float((total / max_total * 100) if max_total > 0 else 0.0)

        if score_row:
            score_row.criteria_scores = safe_scores
            score_row.total_score = total
            score_row.normalized_score = normalized
            score_row.is_present = bool(entry.is_present)
        else:
            score_row = PersohubEventScore(
                event_id=event.id,
                round_id=round_id,
                entity_type=entity_type,
                user_id=user_id,
                team_id=team_id,
                criteria_scores=safe_scores,
                total_score=total,
                normalized_score=normalized,
                is_present=bool(entry.is_present),
            )
            db.add(score_row)
            if is_user:
                score_user_map[entity_id_value] = score_row
            else:
                score_team_map[entity_id_value] = score_row

        if attendance_row:
            attendance_row.is_present = bool(entry.is_present)
            attendance_row.marked_by_user_id = admin.id
        else:
            attendance_row = PersohubEventAttendance(
                event_id=event.id,
                round_id=round_id,
                entity_type=entity_type,
                user_id=user_id,
                team_id=team_id,
                is_present=bool(entry.is_present),
                marked_by_user_id=admin.id,
            )
            db.add(attendance_row)
            if is_user:
                attendance_user_map[entity_id_value] = attendance_row
            else:
                attendance_team_map[entity_id_value] = attendance_row
    _recompute_round_normalized_scores(db, event, round_row)
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "save_persohub_event_scores",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/scores",
        meta={"count": len(entries)},
    )
    return {"message": "Scores saved"}


@router.post("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/import-scores")
def import_scores(
    slug: str,
    round_id: int,
    file: UploadFile = File(...),
    preview: bool = Query(False),
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    if round_row.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is frozen")
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx is supported")

    wb = load_workbook(filename=io.BytesIO(file.file.read()))
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    headers_norm = {h.lower(): idx for idx, h in enumerate(headers)}
    id_col_name = "register number" if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else "team code"
    name_col_name = "name" if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else "team name"
    if id_col_name not in headers_norm:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing '{id_col_name}' column")

    criteria = _criteria_def(round_row)
    criteria_max = {c["name"]: float(c.get("max_marks", 0) or 0) for c in criteria}
    missing_criteria_headers = [name for name in criteria_max.keys() if name.lower() not in headers_norm]
    if missing_criteria_headers:
        missing = ", ".join(missing_criteria_headers)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing criteria columns: {missing}")
    max_total = sum(criteria_max.values()) if criteria_max else 100
    panel_assignment_map: Dict[Tuple[str, int], PersohubEventRoundPanelAssignment] = {}
    if bool(round_row.panel_mode_enabled):
        _, panel_assignment_map = _round_panel_maps(db, round_row)

    entities = _round_scoring_entities(db, event, round_row)
    entity_by_identifier = {}
    for entity in entities:
        identifier_key = str(entity.get("regno_or_code") or "").strip().upper()
        if identifier_key:
            entity_by_identifier[identifier_key] = entity

    id_idx = headers_norm.get(id_col_name)
    name_idx = headers_norm.get(name_col_name)
    present_idx = headers_norm.get("present")
    criteria_indices = {name: headers_norm.get(name.lower()) for name in criteria_max.keys()}

    truthy_values = {"yes", "y", "1", "true", "present"}
    falsy_values = {"no", "n", "0", "false", "absent"}

    total_rows = 0
    valid_rows = []
    identified_rows = []
    mismatched_rows = []
    unidentified_rows = []
    other_required_rows = []
    errors = []

    def _append_error(message: str) -> None:
        if len(errors) < 50:
            errors.append(message)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [row[idx] if idx < len(row) else None for idx in range(len(headers))]

        has_data = False
        for value in values:
            if value is None:
                continue
            if isinstance(value, str):
                if value.strip():
                    has_data = True
                    break
            else:
                has_data = True
                break
        if not has_data:
            continue

        total_rows += 1
        raw_identifier = values[id_idx] if id_idx is not None else None
        identifier = str(raw_identifier or "").strip().upper()
        provided_name = str(values[name_idx] or "").strip() if name_idx is not None else ""

        if not identifier:
            reason = f"Missing {id_col_name.title()}"
            other_required_rows.append({
                "row": row_idx,
                "identifier": "",
                "name": provided_name,
                "reason": reason,
            })
            _append_error(f"Row {row_idx}: {reason}")
            continue

        entity = entity_by_identifier.get(identifier)
        if not entity:
            reason = f"{id_col_name.title()} {identifier} not found in current round participants"
            unidentified_rows.append({
                "row": row_idx,
                "identifier": identifier,
                "name": provided_name,
                "reason": reason,
            })
            _append_error(f"Row {row_idx}: {reason}")
            continue

        entity_type = PersohubEventEntityType.USER if entity.get("entity_type") == "user" else PersohubEventEntityType.TEAM
        user_id = int(entity["entity_id"]) if entity_type == PersohubEventEntityType.USER else None
        team_id = int(entity["entity_id"]) if entity_type == PersohubEventEntityType.TEAM else None
        entity_type_key = "user" if entity_type == PersohubEventEntityType.USER else "team"
        entity_id_value = int(user_id) if entity_type == PersohubEventEntityType.USER else int(team_id)
        has_any_score_input = False
        for name in criteria_max.keys():
            idx = criteria_indices.get(name)
            raw = values[idx] if idx is not None else None
            if raw is not None and str(raw).strip() != "":
                has_any_score_input = True
                break

        is_present = has_any_score_input
        if present_idx is not None:
            present_raw = values[present_idx]
            present_text = str(present_raw or "").strip().lower()
            if not present_text:
                is_present = has_any_score_input
            elif present_text in truthy_values:
                is_present = True
            elif present_text in falsy_values:
                is_present = has_any_score_input
            else:
                reason = "Invalid Present value (use Yes/No)"
                other_required_rows.append({
                    "row": row_idx,
                    "identifier": identifier,
                    "name": provided_name,
                    "reason": reason,
                })
                _append_error(f"Row {row_idx}: {reason}")
                continue

        if bool(round_row.panel_mode_enabled) and is_present:
            assignment_row = panel_assignment_map.get(_panel_entity_key(entity_type_key, entity_id_value))
            if assignment_row is None or assignment_row.panel_id is None:
                reason = "Panel assignment required for present scoring in panel mode"
                other_required_rows.append({
                    "row": row_idx,
                    "identifier": identifier,
                    "name": provided_name or str(entity.get("name") or ""),
                    "reason": reason,
                })
                _append_error(f"Row {row_idx}: {reason}")
                continue

        row_errors = []
        scores = {}
        if is_present:
            for name, max_marks in criteria_max.items():
                idx = criteria_indices.get(name)
                raw = values[idx] if idx is not None else None
                if raw is None or str(raw).strip() == "":
                    row_errors.append(f"{name} is required")
                    continue
                try:
                    score = _parse_import_score_value(raw, max_marks)
                except ValueError as exc:
                    if str(exc) == "invalid_denominator":
                        row_errors.append(f"Invalid score for {name} (denominator must be > 0)")
                    else:
                        row_errors.append(f"Invalid score for {name}")
                    continue
                except Exception:
                    row_errors.append(f"Invalid score for {name}")
                    continue
                if score < 0 or score > max_marks:
                    row_errors.append(f"{name} must be between 0 and {max_marks}")
                    continue
                scores[name] = score
        else:
            scores = {name: 0.0 for name in criteria_max.keys()}

        if row_errors:
            reason = "; ".join(row_errors)
            other_required_rows.append({
                "row": row_idx,
                "identifier": identifier,
                "name": provided_name,
                "reason": reason,
            })
            _append_error(f"Row {row_idx}: {reason}")
            continue

        total = float(sum(scores.values())) if is_present else 0.0
        normalized = float((total / max_total * 100) if max_total > 0 and is_present else 0.0)
        canonical_name = str(entity.get("name") or "").strip()
        mismatch = bool(
            provided_name
            and canonical_name
            and _normalize_compare_text(provided_name) != _normalize_compare_text(canonical_name)
        )

        row_payload = {
            "row": row_idx,
            "identifier": identifier,
            "provided_name": provided_name,
            "expected_name": canonical_name,
            "is_present": is_present,
            "entity_type": entity_type,
            "user_id": user_id,
            "team_id": team_id,
            "scores": scores,
            "total": total,
            "normalized": normalized,
        }
        valid_rows.append(row_payload)

        if mismatch:
            mismatched_rows.append({
                "row": row_idx,
                "identifier": identifier,
                "provided_name": provided_name,
                "expected_name": canonical_name,
                "reason": "Name does not match canonical record",
            })
        else:
            identified_rows.append({
                "row": row_idx,
                "identifier": identifier,
                "name": canonical_name or provided_name,
            })

    ready_to_import = len(valid_rows)

    def _response_payload(imported_count: int):
        return {
            "preview": bool(preview),
            "total_rows": total_rows,
            "identified_count": len(identified_rows),
            "mismatched_count": len(mismatched_rows),
            "unidentified_count": len(unidentified_rows),
            "other_required_count": len(other_required_rows),
            "ready_to_import": ready_to_import,
            "identified_rows": identified_rows[:200],
            "mismatched_rows": mismatched_rows[:200],
            "unidentified_rows": unidentified_rows[:200],
            "other_required_rows": other_required_rows[:200],
            "imported": imported_count,
            "errors": errors[:50],
        }

    if preview:
        return _response_payload(imported_count=0)

    for item in valid_rows:
        entity_type = item["entity_type"]
        user_id = item["user_id"]
        team_id = item["team_id"]
        existing = db.query(PersohubEventScore).filter(
            PersohubEventScore.event_id == event.id,
            PersohubEventScore.round_id == round_id,
            PersohubEventScore.entity_type == entity_type,
            PersohubEventScore.user_id == user_id,
            PersohubEventScore.team_id == team_id,
        ).first()
        if existing:
            existing.criteria_scores = item["scores"]
            existing.total_score = item["total"]
            existing.normalized_score = item["normalized"]
            existing.is_present = item["is_present"]
        else:
            db.add(
                PersohubEventScore(
                    event_id=event.id,
                    round_id=round_id,
                    entity_type=entity_type,
                    user_id=user_id,
                    team_id=team_id,
                    criteria_scores=item["scores"],
                    total_score=item["total"],
                    normalized_score=item["normalized"],
                    is_present=item["is_present"],
                )
            )

        attendance_row = db.query(PersohubEventAttendance).filter(
            PersohubEventAttendance.event_id == event.id,
            PersohubEventAttendance.round_id == round_id,
            PersohubEventAttendance.entity_type == entity_type,
            PersohubEventAttendance.user_id == user_id,
            PersohubEventAttendance.team_id == team_id,
        ).first()
        if attendance_row:
            attendance_row.is_present = item["is_present"]
            attendance_row.marked_by_user_id = admin.id
        else:
            db.add(
                PersohubEventAttendance(
                    event_id=event.id,
                    round_id=round_id,
                    entity_type=entity_type,
                    user_id=user_id,
                    team_id=team_id,
                    is_present=item["is_present"],
                    marked_by_user_id=admin.id,
                )
            )

    _recompute_round_normalized_scores(db, event, round_row)
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "import_persohub_event_scores",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/import-scores",
        meta={
            "preview": False,
            "total_rows": total_rows,
            "ready_to_import": ready_to_import,
            "imported": ready_to_import,
            "unidentified": len(unidentified_rows),
            "other_required": len(other_required_rows),
            "mismatched": len(mismatched_rows),
        },
    )
    return _response_payload(imported_count=ready_to_import)


@router.get("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/score-template")
def score_template(
    slug: str,
    round_id: int,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    criteria = _criteria_def(round_row)
    criteria_names = [c["name"] for c in criteria]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{event.event_code}-R{round_row.round_no}"
    id_col = "Register Number" if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else "Team Code"
    name_col = "Name" if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL else "Team Name"
    ws.append([id_col, name_col] + criteria_names)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{event.event_code}_round_{round_row.round_no}_template.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/freeze")
def freeze_round(
    slug: str,
    round_id: int,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    criteria = _criteria_def(round_row)
    zero_scores = {c["name"]: 0.0 for c in criteria}
    entities = _registered_entities(db, event)
    entities = [item for item in entities if _status_is_active(item.get("status"))]
    for entity in entities:
        entity_type = PersohubEventEntityType.USER if entity["entity_type"] == "user" else PersohubEventEntityType.TEAM
        user_id = entity["entity_id"] if entity_type == PersohubEventEntityType.USER else None
        team_id = entity["entity_id"] if entity_type == PersohubEventEntityType.TEAM else None
        existing = db.query(PersohubEventScore).filter(
            PersohubEventScore.event_id == event.id,
            PersohubEventScore.round_id == round_id,
            PersohubEventScore.entity_type == entity_type,
            PersohubEventScore.user_id == user_id,
            PersohubEventScore.team_id == team_id,
        ).first()
        if not existing:
            db.add(
                PersohubEventScore(
                    event_id=event.id,
                    round_id=round_id,
                    entity_type=entity_type,
                    user_id=user_id,
                    team_id=team_id,
                    criteria_scores=zero_scores,
                    total_score=0.0,
                    normalized_score=0.0,
                    is_present=False,
                )
            )
    round_row.is_frozen = True
    if bool(round_row.panel_mode_enabled):
        _recompute_round_normalized_scores(db, event, round_row)
    db.commit()
    db.refresh(round_row)
    freeze_audit_meta = _upload_round_audit_snapshot(
        db=db,
        event=event,
        round_row=round_row,
        admin=admin,
        audit_type="freeze_snapshot",
        folder="freeze",
    )
    _log_event_admin_action(
        db,
        admin,
        event,
        "freeze_persohub_event_round",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/freeze",
        meta={"round_id": round_id, **freeze_audit_meta},
    )
    return {"message": "Round frozen", **freeze_audit_meta}


@router.post("/persohub/admin/persohub-events/{slug}/rounds/{round_id}/unfreeze")
def unfreeze_round(
    slug: str,
    round_id: int,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    round_row = db.query(PersohubEventRound).filter(PersohubEventRound.id == round_id, PersohubEventRound.event_id == event.id).first()
    if not round_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    round_row.is_frozen = False
    round_row.state = PersohubEventRoundState.ACTIVE
    db.commit()
    _log_event_admin_action(
        db,
        admin,
        event,
        "unfreeze_persohub_event_round",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/rounds/{round_id}/unfreeze",
        meta={"round_id": round_id},
    )
    return {"message": "Round unfrozen"}


@router.get("/persohub/admin/persohub-events/{slug}/leaderboard")
def event_leaderboard(
    slug: str,
    department: Optional[str] = None,
    gender: Optional[str] = None,
    batch: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = None,
    round_ids: Optional[List[int]] = Query(None),
    sort: Optional[str] = Query("rank"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=500),
    response: Response = None,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    sort_option = _normalize_leaderboard_sort(sort)
    rows = []
    round_rows = (
        db.query(PersohubEventRound.id, PersohubEventRound.state, PersohubEventRound.is_frozen)
        .filter(PersohubEventRound.event_id == event.id)
        .all()
    )
    event_round_ids = {int(row.id) for row in round_rows}
    eligible_round_ids = {
        int(row.id)
        for row in round_rows
        if bool(row.is_frozen) or row.state == PersohubEventRoundState.COMPLETED
    }
    requested_round_ids: List[int] = []
    if round_ids:
        seen = set()
        for value in round_ids:
            round_id = int(value)
            if round_id not in seen:
                requested_round_ids.append(round_id)
                seen.add(round_id)
    if requested_round_ids:
        invalid_rounds = sorted([round_id for round_id in requested_round_ids if round_id not in event_round_ids])
        if invalid_rounds:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid round_ids for this event: {invalid_rounds}",
            )
        ineligible_rounds = sorted([round_id for round_id in requested_round_ids if round_id not in eligible_round_ids])
        if ineligible_rounds:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Only completed or frozen rounds are allowed in round_ids: {ineligible_rounds}",
            )
    effective_round_ids = requested_round_ids if requested_round_ids else sorted(eligible_round_ids)

    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        entities = _registered_entities(db, event)
        if department:
            entities = [item for item in entities if str(item.get("department") or "") == str(department)]
        if gender:
            entities = [item for item in entities if str(item.get("gender") or "") == str(gender)]
        if batch:
            entities = [item for item in entities if str(item.get("batch") or "") == str(batch)]
        if status_filter:
            normalized = str(status_filter or "").strip().lower()
            entities = [item for item in entities if str(item.get("status") or "").strip().lower() == normalized]
        if search:
            needle = str(search).lower()
            entities = [
                item for item in entities
                if needle in " ".join(
                    [
                        str(item.get("name") or ""),
                        str(item.get("regno_or_code") or ""),
                        str(item.get("email") or ""),
                        str(item.get("department") or ""),
                        str(item.get("gender") or ""),
                        str(item.get("batch") or ""),
                    ]
                ).lower()
            ]

        entity_ids = [int(item["entity_id"]) for item in entities]
        if entity_ids and effective_round_ids:
            score_rows = (
                db.query(
                    PersohubEventScore.user_id.label("entity_id"),
                    func.coalesce(func.sum(PersohubEventScore.normalized_score), 0.0).label("cumulative_score"),
                    func.coalesce(func.count(func.distinct(PersohubEventScore.round_id)).filter(PersohubEventScore.is_present == True), 0).label("rounds_participated"),  # noqa: E712
                )
                .filter(
                    PersohubEventScore.event_id == event.id,
                    PersohubEventScore.entity_type == PersohubEventEntityType.USER,
                    PersohubEventScore.user_id.in_(entity_ids),
                    PersohubEventScore.round_id.in_(effective_round_ids),
                )
                .group_by(PersohubEventScore.user_id)
                .all()
            )
            attendance_rows = (
                db.query(
                    PersohubEventAttendance.user_id.label("entity_id"),
                    func.count(PersohubEventAttendance.id).label("attendance_count"),
                )
                .filter(
                    PersohubEventAttendance.event_id == event.id,
                    PersohubEventAttendance.entity_type == PersohubEventEntityType.USER,
                    PersohubEventAttendance.is_present == True,  # noqa: E712
                    PersohubEventAttendance.user_id.in_(entity_ids),
                    PersohubEventAttendance.round_id.in_(effective_round_ids),
                )
                .group_by(PersohubEventAttendance.user_id)
                .all()
            )
        else:
            score_rows = []
            attendance_rows = []
        score_map = {
            int(row.entity_id): {
                "cumulative_score": float(row.cumulative_score or 0.0),
                "rounds_participated": int(row.rounds_participated or 0),
            }
            for row in score_rows
            if row.entity_id is not None
        }
        attendance_map = {
            int(row.entity_id): int(row.attendance_count or 0)
            for row in attendance_rows
            if row.entity_id is not None
        }

        for entity in entities:
            user_id = int(entity["entity_id"])
            score_info = score_map.get(user_id, {"cumulative_score": 0.0, "rounds_participated": 0})
            rows.append(
                {
                    **entity,
                    "participant_id": user_id,
                    "register_number": entity.get("regno_or_code"),
                    "cumulative_score": float(score_info["cumulative_score"]),
                    "attendance_count": int(attendance_map.get(user_id, 0)),
                    "rounds_participated": int(score_info["rounds_participated"]),
                }
            )

        rows.sort(
            key=lambda item: (
                0 if str(item.get("status") or "").lower() == "active" else 1,
                -float(item.get("cumulative_score") or 0),
                str(item.get("name") or "").lower(),
            )
        )
        active_rank = 0
        prev_score = None
        for row in rows:
            if str(row.get("status") or "").lower() == "active":
                score = float(row.get("cumulative_score") or 0.0)
                if prev_score is None or score != prev_score:
                    active_rank += 1
                row["rank"] = active_rank
                prev_score = score
            else:
                row["rank"] = None
    else:
        entities = _registered_entities(db, event)
        if status_filter:
            normalized = str(status_filter or "").strip().lower()
            entities = [item for item in entities if str(item.get("status") or "").strip().lower() == normalized]
        if search:
            needle = str(search).lower()
            entities = [
                item
                for item in entities
                if needle in " ".join(
                    [
                        str(item.get("name") or ""),
                        str(item.get("regno_or_code") or ""),
                        str(item.get("status") or ""),
                    ]
                ).lower()
            ]
        entity_ids = [int(item["entity_id"]) for item in entities]
        if entity_ids and effective_round_ids:
            score_rows = (
                db.query(
                    PersohubEventScore.team_id.label("entity_id"),
                    func.coalesce(func.sum(PersohubEventScore.total_score), 0.0).label("cumulative_score"),
                    func.coalesce(func.count(func.distinct(PersohubEventScore.round_id)).filter(PersohubEventScore.is_present == True), 0).label("rounds_participated"),  # noqa: E712
                )
                .filter(
                    PersohubEventScore.event_id == event.id,
                    PersohubEventScore.entity_type == PersohubEventEntityType.TEAM,
                    PersohubEventScore.team_id.in_(entity_ids),
                    PersohubEventScore.round_id.in_(effective_round_ids),
                )
                .group_by(PersohubEventScore.team_id)
                .all()
            )
            attendance_rows = (
                db.query(
                    PersohubEventAttendance.team_id.label("entity_id"),
                    func.count(PersohubEventAttendance.id).label("attendance_count"),
                )
                .filter(
                    PersohubEventAttendance.event_id == event.id,
                    PersohubEventAttendance.entity_type == PersohubEventEntityType.TEAM,
                    PersohubEventAttendance.is_present == True,  # noqa: E712
                    PersohubEventAttendance.team_id.in_(entity_ids),
                    PersohubEventAttendance.round_id.in_(effective_round_ids),
                )
                .group_by(PersohubEventAttendance.team_id)
                .all()
            )
        else:
            score_rows = []
            attendance_rows = []
        score_map = {
            int(row.entity_id): {
                "cumulative_score": float(row.cumulative_score or 0.0),
                "rounds_participated": int(row.rounds_participated or 0),
            }
            for row in score_rows
            if row.entity_id is not None
        }
        attendance_map = {
            int(row.entity_id): int(row.attendance_count or 0)
            for row in attendance_rows
            if row.entity_id is not None
        }

        for entity in entities:
            team_id = int(entity["entity_id"])
            score_info = score_map.get(team_id, {"cumulative_score": 0.0, "rounds_participated": 0})
            rows.append(
                {
                    **entity,
                    "cumulative_score": float(score_info["cumulative_score"]),
                    "attendance_count": int(attendance_map.get(team_id, 0)),
                    "rounds_participated": int(score_info["rounds_participated"]),
                }
            )
        rows.sort(
            key=lambda item: (
                0 if _status_is_active(item.get("status")) else 1,
                -float(item.get("cumulative_score") or 0),
                str(item.get("name") or "").lower(),
            )
        )
        active_rank = 0
        prev_score = None
        for row in rows:
            if _status_is_active(row.get("status")):
                score = float(row.get("cumulative_score") or 0.0)
                if prev_score is None or score != prev_score:
                    active_rank += 1
                row["rank"] = active_rank
                prev_score = score
            else:
                row["rank"] = None

    _apply_leaderboard_sort(rows, sort_option)

    total = len(rows)
    start = (page - 1) * page_size
    end = start + page_size
    paged = rows[start:end]
    if response is not None:
        response.headers["X-Total-Count"] = str(total)
        response.headers["X-Page"] = str(page)
        response.headers["X-Page-Size"] = str(page_size)
    return paged


@router.post("/persohub/admin/persohub-events/{slug}/email/bulk")
def send_bulk_event_email(
    slug: str,
    payload: EventBulkEmailRequest,
    background_tasks: BackgroundTasks,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
    request: Request = None,
):
    mode = str(payload.recipient_mode or "").strip().lower()
    allowed_modes = {"registered", "active", "eliminated", "top_k", "selected", "unregistered"}
    if mode not in allowed_modes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid recipient_mode")

    subject = str(payload.subject or "").strip()
    html = str(payload.html or "").strip()
    if not subject or not html:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Subject and HTML are required")

    event = _get_event_or_404(db, slug)
    selected_source = str(payload.selected_source or "").strip().lower()
    if mode == "selected" and selected_source == "unregistered":
        items = _unregistered_entities(db, event)
    else:
        items = _unregistered_entities(db, event) if mode == "unregistered" else _registered_entities(db, event)

    if mode != "selected":
        if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
            if payload.department:
                items = [item for item in items if str(item.get("department") or "") == str(payload.department)]
            if payload.gender:
                items = [item for item in items if str(item.get("gender") or "") == str(payload.gender)]
            if payload.batch:
                items = [item for item in items if str(item.get("batch") or "") == str(payload.batch)]
        if payload.status and mode != "unregistered":
            normalized = str(payload.status or "").strip().lower()
            items = [item for item in items if str(item.get("status") or "").strip().lower() == normalized]
        if payload.search:
            needle = str(payload.search).lower()
            filtered = []
            for item in items:
                haystack = " ".join(
                    [
                        str(item.get("name") or ""),
                        str(item.get("regno_or_code") or ""),
                        str(item.get("email") or ""),
                        str(item.get("department") or ""),
                        str(item.get("gender") or ""),
                        str(item.get("batch") or ""),
                    ]
                ).lower()
                if needle in haystack:
                    filtered.append(item)
            items = filtered

    if mode == "active":
        items = [item for item in items if _status_is_active(item.get("status"))]
    elif mode == "eliminated":
        items = [item for item in items if not _status_is_active(item.get("status"))]
    elif mode == "selected":
        if not payload.entity_ids:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="entity_ids required for selected mode")
        target_set = {int(value) for value in payload.entity_ids}
        items = [item for item in items if int(item.get("entity_id")) in target_set]

    leaderboard_rows = event_leaderboard(
        slug=slug,
        department=payload.department,
        gender=payload.gender,
        batch=payload.batch,
        status_filter=payload.status,
        search=payload.search,
        page=1,
        page_size=10000,
        response=None,
        _=admin,
        db=db,
    )
    leaderboard_map: Dict[int, dict] = {}
    for row in leaderboard_rows:
        try:
            entity_id = int(row.get("entity_id"))
            leaderboard_map[entity_id] = row
        except Exception:
            continue

    if mode == "top_k":
        limit = int(payload.top_k or 0)
        if limit <= 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="top_k must be positive")
        source = leaderboard_rows if leaderboard_rows else items
        items = source[:limit]

    entity_ids = [int(item.get("entity_id")) for item in items if item.get("entity_id") is not None]

    user_map: Dict[int, PdaUser] = {}
    team_map: Dict[int, PersohubEventTeam] = {}
    leader_map: Dict[int, PdaUser] = {}
    home_team_map: Dict[int, PdaTeam] = {}
    treat_as_individual = (
        event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL
        or mode == "unregistered"
        or (mode == "selected" and selected_source == "unregistered")
    )
    if treat_as_individual:
        users = db.query(PdaUser).filter(PdaUser.id.in_(entity_ids)).all() if entity_ids else []
        user_map = {user.id: user for user in users}
        if users:
            team_rows = (
                db.query(PdaTeam)
                .filter(PdaTeam.user_id.in_([user.id for user in users]))
                .order_by(PdaTeam.id.asc())
                .all()
            )
            for member in team_rows:
                if member.user_id not in home_team_map:
                    home_team_map[member.user_id] = member
    else:
        teams = db.query(PersohubEventTeam).filter(PersohubEventTeam.id.in_(entity_ids)).all() if entity_ids else []
        team_map = {team.id: team for team in teams}
        leader_ids = [team.team_lead_user_id for team in teams if team.team_lead_user_id]
        leaders = db.query(PdaUser).filter(PdaUser.id.in_(leader_ids)).all() if leader_ids else []
        leader_map = {leader.id: leader for leader in leaders}
        leader_team_map: Dict[int, PdaTeam] = {}
        if leaders:
            leader_team_rows = (
                db.query(PdaTeam)
                .filter(PdaTeam.user_id.in_([leader.id for leader in leaders]))
                .order_by(PdaTeam.id.asc())
                .all()
            )
            for member in leader_team_rows:
                if member.user_id not in leader_team_map:
                    leader_team_map[member.user_id] = member

    skipped_no_email = 0
    skipped_duplicate = 0
    seen = set()
    unique_recipients: List[Tuple[str, Dict[str, object]]] = []
    for item in items:
        entity_id = int(item.get("entity_id"))
        leaderboard = leaderboard_map.get(entity_id, {})
        if treat_as_individual:
            user = user_map.get(entity_id)
            if not user:
                skipped_no_email += 1
                continue
            email_value = str(user.email or "").strip().lower()
            member = home_team_map.get(user.id)
            context = {
                "name": user.name,
                "profile_name": user.profile_name,
                "regno": user.regno,
                "email": user.email,
                "dept": user.dept,
                "gender": user.gender,
                "phno": user.phno,
                "dob": user.dob,
                "team": member.team if member else None,
                "designation": member.designation if member else None,
                "instagram_url": user.instagram_url,
                "linkedin_url": user.linkedin_url,
                "github_url": user.github_url,
                "photo_url": user.image_url,
                "is_member": bool(user.is_member),
                "email_verified": bool(user.email_verified_at),
                "created_at": user.created_at,
                "updated_at": user.updated_at,
                "batch": extract_batch(user.regno),
            }
        else:
            team = team_map.get(entity_id)
            leader = leader_map.get(team.team_lead_user_id) if team else None
            leader_member = leader_team_map.get(leader.id) if leader else None
            if not leader:
                skipped_no_email += 1
                continue
            email_value = str(leader.email or "").strip().lower()
            context = {
                "name": leader.name,
                "profile_name": leader.profile_name,
                "regno": leader.regno,
                "email": leader.email,
                "dept": leader.dept,
                "gender": leader.gender,
                "phno": leader.phno,
                "dob": leader.dob,
                "team": leader_member.team if leader_member else None,
                "designation": leader_member.designation if leader_member else None,
                "instagram_url": leader.instagram_url,
                "linkedin_url": leader.linkedin_url,
                "github_url": leader.github_url,
                "photo_url": leader.image_url,
                "is_member": bool(leader.is_member),
                "email_verified": bool(leader.email_verified_at),
                "created_at": leader.created_at,
                "updated_at": leader.updated_at,
                "batch": extract_batch(leader.regno),
                "team_name": team.team_name if team else None,
                "team_code": team.team_code if team else None,
                "members_count": item.get("members_count"),
                "leader_name": leader.name,
                "leader_regno": leader.regno,
                "leader_email": leader.email,
                "leader_profile_name": leader.profile_name,
                "leader_dept": leader.dept,
                "leader_phno": leader.phno,
                "leader_gender": leader.gender,
                "leader_batch": extract_batch(leader.regno),
            }

        context.update(
            {
                "status": item.get("status"),
                "batch": item.get("batch") or context.get("batch"),
                "regno_or_code": item.get("regno_or_code"),
                "referral_code": item.get("referral_code"),
                "referred_by": item.get("referred_by"),
                "referral_count": item.get("referral_count"),
                "entity_id": item.get("entity_id"),
                "participant_id": item.get("participant_id") or item.get("entity_id"),
                "entity_type": item.get("entity_type"),
                "event_title": event.title,
                "event_code": event.event_code,
                "rank": leaderboard.get("rank"),
                "cumulative_score": leaderboard.get("cumulative_score"),
                "attendance_count": leaderboard.get("attendance_count"),
                "rounds_participated": leaderboard.get("rounds_participated"),
            }
        )
        if not email_value:
            skipped_no_email += 1
            continue
        if email_value in seen:
            skipped_duplicate += 1
            continue
        seen.add(email_value)
        unique_recipients.append((email_value, context))

    background_tasks.add_task(
        _send_bulk_event_email_background,
        unique_recipients,
        subject,
        html,
        payload.text,
        admin.id,
        event.id,
        event.slug,
        mode,
        request.method if request else "POST",
        request.url.path if request else f"/persohub/admin/persohub-events/{slug}/email/bulk",
        skipped_no_email,
        skipped_duplicate,
    )
    return {
        "requested": len(items),
        "queued": len(unique_recipients),
        "skipped_no_email": skipped_no_email,
        "skipped_duplicate": skipped_duplicate,
    }


@router.get("/persohub/admin/persohub-events/{slug}/logs", response_model=List[PersohubEventLogResponse])
def event_logs(
    slug: str,
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    action: Optional[str] = None,
    method: Optional[str] = None,
    path_contains: Optional[str] = None,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    query = db.query(PersohubEventLog).filter(PersohubEventLog.event_slug == event.slug)
    if action:
        query = query.filter(PersohubEventLog.action == str(action).strip())
    if method:
        query = query.filter(func.lower(PersohubEventLog.method) == str(method).strip().lower())
    if path_contains:
        query = query.filter(PersohubEventLog.path.ilike(f"%{str(path_contains).strip()}%"))
    logs = query.order_by(PersohubEventLog.created_at.desc()).offset(offset).limit(limit).all()
    return [PersohubEventLogResponse.model_validate(row) for row in logs]


def _export_to_csv(headers: List[str], rows: List[List[object]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


def _export_to_xlsx(headers: List[str], rows: List[List[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _normalize_excel_sheet_title(value: Optional[str], fallback: str = "Sheet") -> str:
    raw = str(value or "").strip()
    if raw:
        raw = re.sub(r"[\[\]\:\*\?\/\\]", " ", raw)
        raw = re.sub(r"\s+", " ", raw).strip()
    candidate = raw or fallback
    return candidate[:31] if len(candidate) > 31 else candidate


def _unique_excel_sheet_title(value: Optional[str], used_titles: Set[str], fallback: str = "Sheet") -> str:
    base = _normalize_excel_sheet_title(value, fallback=fallback)
    if not base:
        base = fallback
    index = 1
    while True:
        suffix = f" ({index})" if index > 1 else ""
        max_base_len = max(31 - len(suffix), 1)
        candidate = f"{base[:max_base_len]}{suffix}"
        key = candidate.lower()
        if key not in used_titles:
            used_titles.add(key)
            return candidate
        index += 1


def _export_to_multi_sheet_xlsx(headers: List[str], sheets: List[Tuple[str, List[List[object]]]]) -> bytes:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    used_titles: Set[str] = set()
    safe_sheets = sheets if sheets else [("Participants", [])]
    for sheet_name, rows in safe_sheets:
        ws = wb.create_sheet(title=_unique_excel_sheet_title(sheet_name, used_titles, fallback="Participants"))
        ws.append(headers)
        for row in rows:
            ws.append(row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _extract_round_state_text(value) -> str:
    if hasattr(value, "value"):
        return str(value.value or "").strip().lower()
    return str(value or "").strip().lower()


def _official_shortlist_round_number(db: Session, event: PersohubEvent) -> int:
    round_rows = (
        db.query(PersohubEventRound.round_no, PersohubEventRound.state)
        .filter(PersohubEventRound.event_id == event.id)
        .order_by(PersohubEventRound.round_no.asc())
        .all()
    )
    latest_completed_round_no: Optional[int] = None
    for row in round_rows:
        if _extract_round_state_text(row.state) == "completed":
            latest_completed_round_no = int(row.round_no)
    return latest_completed_round_no if latest_completed_round_no is not None else 1


def _official_shortlist_heading(db: Session, event: PersohubEvent) -> str:
    return f"ROUND {_official_shortlist_round_number(db, event)} SHORTLISTED"


def _normalize_leaderboard_sort(sort_value: Optional[str]) -> str:
    candidate = str(sort_value or "rank").strip().lower()
    allowed = {"rank", "score_desc", "score_asc", "name_asc", "name_desc", "rounds_desc", "rounds_asc"}
    if candidate not in allowed:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid sort option: {candidate}")
    return candidate


def _apply_leaderboard_sort(rows: List[dict], sort_option: str) -> None:
    def _num(value) -> float:
        try:
            return float(value or 0)
        except Exception:
            return 0.0

    def _name(item: dict) -> str:
        return str(item.get("name") or "").strip().lower()

    if sort_option == "score_desc":
        rows.sort(key=lambda item: (-_num(item.get("cumulative_score")), _name(item)))
        return
    if sort_option == "score_asc":
        rows.sort(key=lambda item: (_num(item.get("cumulative_score")), _name(item)))
        return
    if sort_option == "name_asc":
        rows.sort(key=lambda item: _name(item))
        return
    if sort_option == "name_desc":
        rows.sort(key=lambda item: _name(item), reverse=True)
        return
    if sort_option == "rounds_desc":
        rows.sort(key=lambda item: (-_num(item.get("rounds_participated")), _name(item)))
        return
    if sort_option == "rounds_asc":
        rows.sort(key=lambda item: (_num(item.get("rounds_participated")), _name(item)))
        return

    rows.sort(
        key=lambda item: (
            int(item.get("rank")) if item.get("rank") is not None else 10**9,
            _name(item),
        )
    )


def _official_logo_url(db: Session, event: PersohubEvent) -> Optional[str]:
    candidate_ids: List[int] = []
    community = db.query(PersohubCommunity).filter(PersohubCommunity.id == event.community_id).first()
    if community and community.club_id:
        try:
            club_id = int(community.club_id)
            if club_id > 0:
                candidate_ids.append(club_id)
        except Exception:
            pass
    if 1 not in candidate_ids:
        candidate_ids.append(1)

    clubs = db.query(PersohubClub).filter(PersohubClub.id.in_(candidate_ids)).all() if candidate_ids else []
    club_map = {int(club.id): club for club in clubs}
    for club_id in candidate_ids:
        club = club_map.get(club_id)
        logo_url = str(getattr(club, "club_logo_url", "") or "").strip() if club else ""
        if logo_url:
            return logo_url
    return None


def _load_remote_image_data_uri(
    image_url: Optional[str],
) -> Optional[str]:
    if not image_url:
        return None
    try:
        request = UrlRequest(image_url, headers={"User-Agent": "Mozilla/5.0"})
        try:
            with urlopen(request, timeout=8) as response:
                image_bytes = response.read()
                content_type = str(response.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        except Exception:
            with urlopen(request, timeout=8, context=ssl._create_unverified_context()) as response:
                image_bytes = response.read()
                content_type = str(response.headers.get("Content-Type") or "").split(";")[0].strip().lower()
        if not image_bytes:
            return None
        if not content_type.startswith("image/"):
            if image_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
                content_type = "image/png"
            elif image_bytes[:3] == b"\xff\xd8\xff":
                content_type = "image/jpeg"
            elif image_bytes[:4] == b"RIFF" and image_bytes[8:12] == b"WEBP":
                content_type = "image/webp"
            else:
                content_type = "image/png"
        encoded = base64.b64encode(image_bytes).decode("ascii")
        return f"data:{content_type};base64,{encoded}"
    except Exception:
        return None


def _render_leaderboard_template_html(
    *,
    event: PersohubEvent,
    round_number: int,
    table_headers: List[dict],
    table_rows: List[dict],
    is_team_mode: bool,
    left_logo_data_uri: Optional[str],
    right_logo_data_uri: Optional[str],
    watermark_logo_data_uri: Optional[str],
) -> str:
    try:
        from jinja2 import Environment, FileSystemLoader, select_autoescape
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Template dependency is unavailable") from exc

    template_dir = Path(__file__).resolve().parents[1] / "templates"
    template_name = "event_leaderboard_official.html"
    template_path = template_dir / template_name
    if not template_path.exists():
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Leaderboard PDF template is missing")

    environment = Environment(
        loader=FileSystemLoader(str(template_dir)),
        autoescape=select_autoescape(enabled_extensions=("html", "xml")),
    )
    template = environment.get_template(template_name)
    return template.render(
        event_name=str(event.title or "").upper(),
        round_number=round_number,
        table_headers=table_headers,
        table_rows=table_rows,
        is_team_mode=is_team_mode,
        column_count=len(table_headers),
        min_rows=6,
        left_logo=left_logo_data_uri,
        right_logo=right_logo_data_uri,
        watermark_logo=watermark_logo_data_uri,
        footer_text="The leaderboard was autogenerated using PERSOHUB version 1.0",
    )


def _render_html_to_pdf(html_content: str) -> bytes:
    try:
        from xhtml2pdf import pisa
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="PDF rendering dependency is unavailable") from exc

    output = io.BytesIO()
    result = pisa.CreatePDF(src=html_content, dest=output, encoding="utf-8")
    if getattr(result, "err", 0):
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to render leaderboard PDF")
    output.seek(0)
    return output.read()


def _decode_data_uri_bytes(data_uri: Optional[str]) -> Optional[bytes]:
    if not data_uri or not str(data_uri).startswith("data:"):
        return None
    try:
        _, payload = str(data_uri).split(",", 1)
        return base64.b64decode(payload)
    except Exception:
        return None


def _apply_pdf_background_and_footer(pdf_bytes: bytes, watermark_data_uri: Optional[str]) -> bytes:
    try:
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas as pdf_canvas
        from pypdf import PdfReader, PdfWriter
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="PDF decoration dependencies are unavailable") from exc

    base_reader = PdfReader(io.BytesIO(pdf_bytes))
    writer = PdfWriter()
    watermark_bytes = _decode_data_uri_bytes(watermark_data_uri)
    watermark_reader = None
    if watermark_bytes:
        try:
            watermark_reader = ImageReader(io.BytesIO(watermark_bytes))
        except Exception:
            watermark_reader = None

    for index, base_page in enumerate(base_reader.pages, start=1):
        width = float(base_page.mediabox.width)
        height = float(base_page.mediabox.height)
        stamp_buffer = io.BytesIO()
        stamp_canvas = pdf_canvas.Canvas(stamp_buffer, pagesize=(width, height))

        if watermark_reader is not None:
            mark_size = min(400.0, width * 0.72)
            stamp_canvas.saveState()
            if hasattr(stamp_canvas, "setFillAlpha"):
                stamp_canvas.setFillAlpha(0.1)
            stamp_canvas.drawImage(
                watermark_reader,
                (width - mark_size) / 2.0,
                (height - mark_size) / 2.0,
                width=mark_size,
                height=mark_size,
                preserveAspectRatio=True,
                mask="auto",
            )
            stamp_canvas.restoreState()

        stamp_canvas.setFont("Times-Roman", 10)
        stamp_canvas.drawCentredString(width / 2.0, 10 * mm, f"Page {index}")
        stamp_canvas.save()

        stamp_buffer.seek(0)
        stamp_page = PdfReader(stamp_buffer).pages[0]
        base_page.merge_page(stamp_page)
        writer.add_page(base_page)

    out = io.BytesIO()
    writer.write(out)
    out.seek(0)
    return out.read()


def _export_leaderboard_to_pdf(db: Session, event: PersohubEvent, leaderboard: List[dict]) -> bytes:
    round_number = _official_shortlist_round_number(db, event)
    is_team_mode = event.participant_mode == PersohubEventParticipantMode.TEAM
    if is_team_mode:
        headers = [
            {"label": "SI.NO", "class_name": "si-no-column"},
            {"label": "TEAM CODE", "class_name": "team-code-column"},
            {"label": "TEAM NAME", "class_name": "team-name-column"},
        ]
        rows = [{
            "cells": [
                str(index),
                str(row.get("regno_or_code") or row.get("register_number") or ""),
                str(row.get("name") or ""),
            ]
        } for index, row in enumerate(leaderboard, start=1)]
    else:
        headers = [
            {"label": "SI.NO", "class_name": "si-no-column"},
            {"label": "REGISTER NO", "class_name": "register-no-column"},
            {"label": "NAME", "class_name": "name-column"},
            {"label": "DEPARTMENT", "class_name": "department-column"},
        ]
        rows = [{
            "cells": [
                str(index),
                str(row.get("register_number") or row.get("regno_or_code") or ""),
                str(row.get("name") or ""),
                str(row.get("department") or "-"),
            ]
        } for index, row in enumerate(leaderboard, start=1)]

    club_logo_data_uri = _load_remote_image_data_uri(_official_logo_url(db, event))
    left_logo_data_uri = _load_remote_image_data_uri(OFFICIAL_LETTERHEAD_LEFT_LOGO_URL) or club_logo_data_uri
    html_content = _render_leaderboard_template_html(
        event=event,
        round_number=round_number,
        table_headers=headers,
        table_rows=rows,
        is_team_mode=is_team_mode,
        left_logo_data_uri=left_logo_data_uri,
        right_logo_data_uri=club_logo_data_uri,
        watermark_logo_data_uri=club_logo_data_uri,
    )
    rendered_pdf = _render_html_to_pdf(html_content)
    return _apply_pdf_background_and_footer(rendered_pdf, club_logo_data_uri)


@router.get("/persohub/admin/persohub-events/{slug}/export/participants")
def export_participants(
    slug: str,
    format: str = Query("csv"),
    department: Optional[str] = None,
    gender: Optional[str] = None,
    batch: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = None,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    entities = event_participants(
        slug=slug,
        search=search,
        department=department,
        gender=gender,
        batch=batch,
        status_filter=status_filter,
        page=1,
        page_size=100000,
        response=None,
        _=None,
        db=db,
    )
    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        headers = ["Register Number", "Name", "Email", "Department", "Gender", "Batch", "Status", "Referral Code", "Referred By", "Referral Count"]
        rows = [
            [
                e.get("register_number"),
                e.get("name"),
                e.get("email"),
                e.get("department"),
                e.get("gender"),
                e.get("batch"),
                e.get("status"),
                e.get("referral_code"),
                e.get("referred_by"),
                e.get("referral_count", 0),
            ]
            for e in entities
        ]
    else:
        headers = ["Entity Type", "Name", "Register/Team Code", "Members Count", "Status"]
        rows = [[e["entity_type"], e["name"], e["regno_or_code"], e.get("members_count", 1), e.get("status")] for e in entities]
    if format == "xlsx":
        content = _export_to_xlsx(headers, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{event.event_code}_participants.xlsx"
    else:
        content = _export_to_csv(headers, rows)
        media_type = "text/csv"
        filename = f"{event.event_code}_participants.csv"
    return StreamingResponse(io.BytesIO(content), media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/persohub/admin/persohub-events/{slug}/export/leaderboard")
def export_leaderboard(
    slug: str,
    format: str = Query("csv"),
    department: Optional[str] = None,
    gender: Optional[str] = None,
    batch: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    search: Optional[str] = None,
    round_ids: Optional[List[int]] = Query(None),
    sort: Optional[str] = Query("rank"),
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    leaderboard = event_leaderboard(
        slug=slug,
        department=department,
        gender=gender,
        batch=batch,
        status_filter=status_filter,
        search=search,
        round_ids=round_ids,
        sort=sort,
        page=1,
        page_size=10000,
        response=None,
        _=None,
        db=db,
    )
    event = _get_event_or_404(db, slug)
    if event.participant_mode == PersohubEventParticipantMode.INDIVIDUAL:
        headers = ["Rank", "Register Number", "Name", "Department", "Gender", "Batch", "Status", "Rounds", "Attendance", "Score", "Referral Count"]
        rows = [
            [
                row.get("rank"),
                row.get("register_number"),
                row.get("name"),
                row.get("department"),
                row.get("gender"),
                row.get("batch"),
                row.get("status"),
                row.get("rounds_participated", 0),
                row.get("attendance_count", 0),
                row.get("cumulative_score", 0),
                row.get("referral_count", 0),
            ]
            for row in leaderboard
        ]
    else:
        headers = ["Rank", "Entity Type", "Name", "Register/Team Code", "Status", "Rounds", "Attendance", "Score"]
        rows = [
            [
                row["rank"],
                row["entity_type"],
                row["name"],
                row["regno_or_code"],
                row.get("status"),
                row.get("rounds_participated", 0),
                row["attendance_count"],
                row["cumulative_score"],
            ]
            for row in leaderboard
        ]
    if format == "pdf":
        content = _export_leaderboard_to_pdf(db=db, event=event, leaderboard=leaderboard)
        media_type = "application/pdf"
        filename = f"{event.event_code}_leaderboard_official.pdf"
    elif format == "xlsx":
        content = _export_to_xlsx(headers, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{event.event_code}_leaderboard.xlsx"
    else:
        content = _export_to_csv(headers, rows)
        media_type = "text/csv"
        filename = f"{event.event_code}_leaderboard.csv"
    return StreamingResponse(io.BytesIO(content), media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/persohub/admin/persohub-events/{slug}/export/round/{round_id}")
def export_round(
    slug: str,
    round_id: int,
    format: str = Query("csv"),
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    result = round_participants(slug=slug, round_id=round_id, admin=admin, db=db)
    headers = ["Entity Type", "Name", "Register/Team Code", "Total Score", "Normalized Score", "Present"]
    rows = [
        [
            row["entity_type"],
            row["name"],
            row["regno_or_code"],
            row["total_score"],
            row["normalized_score"],
            row["is_present"],
        ]
        for row in result
    ]
    event = _get_event_or_404(db, slug)
    if format == "xlsx":
        content = _export_to_xlsx(headers, rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"{event.event_code}_round_{round_id}.xlsx"
    else:
        content = _export_to_csv(headers, rows)
        media_type = "text/csv"
        filename = f"{event.event_code}_round_{round_id}.csv"
    return StreamingResponse(io.BytesIO(content), media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/persohub/admin/persohub-events/{slug}/export/round/{round_id}/panel-wise")
def export_round_panel_wise(
    slug: str,
    round_id: int,
    format: str = Query("xlsx"),
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    normalized_format = str(format or "xlsx").strip().lower()
    if normalized_format != "xlsx":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only xlsx export is supported for panel-wise export")

    event = _get_event_or_404(db, slug)
    round_row = _get_event_round_or_404(db, event.id, round_id)
    result = round_participants(slug=slug, round_id=round_id, search=None, admin=admin, db=db)

    headers = ["Entity Type", "Name", "Register/Team Code", "Total Score", "Normalized Score", "Present"]

    def _to_export_row(row: dict) -> List[object]:
        return [
            row.get("entity_type"),
            row.get("name"),
            row.get("regno_or_code"),
            row.get("total_score"),
            row.get("normalized_score"),
            row.get("is_present"),
        ]

    sheets: List[Tuple[str, List[List[object]]]] = []
    if bool(round_row.panel_mode_enabled):
        panel_rows = (
            db.query(PersohubEventRoundPanel)
            .filter(
                PersohubEventRoundPanel.event_id == event.id,
                PersohubEventRoundPanel.round_id == round_row.id,
            )
            .order_by(PersohubEventRoundPanel.panel_no.asc(), PersohubEventRoundPanel.id.asc())
            .all()
        )
        rows_by_panel_id: Dict[int, List[List[object]]] = {}
        unassigned_rows: List[List[object]] = []
        for participant in result:
            panel_id = participant.get("panel_id")
            export_row = _to_export_row(participant)
            if panel_id is None:
                unassigned_rows.append(export_row)
                continue
            try:
                safe_panel_id = int(panel_id)
            except Exception:
                unassigned_rows.append(export_row)
                continue
            rows_by_panel_id.setdefault(safe_panel_id, []).append(export_row)

        known_panel_ids: Set[int] = set()
        for panel in panel_rows:
            panel_id = int(panel.id)
            known_panel_ids.add(panel_id)
            panel_no = int(panel.panel_no) if panel.panel_no is not None else panel_id
            panel_name = str(panel.name or "").strip()
            label = f"Panel {panel_no}"
            if panel_name:
                label = f"{label} - {panel_name}"
            sheets.append((label, rows_by_panel_id.get(panel_id, [])))

        orphan_panel_ids = sorted([panel_id for panel_id in rows_by_panel_id.keys() if panel_id not in known_panel_ids])
        for panel_id in orphan_panel_ids:
            sheets.append((f"Panel {panel_id}", rows_by_panel_id.get(panel_id, [])))

        if unassigned_rows or not sheets:
            sheets.append(("Unassigned", unassigned_rows))
    else:
        sheets.append(("All Participants", [_to_export_row(row) for row in result]))

    content = _export_to_multi_sheet_xlsx(headers, sheets)
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = f"{event.event_code}_round_{round_row.round_no}_panel_wise.xlsx"
    return StreamingResponse(io.BytesIO(content), media_type=media_type, headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/persohub/admin/persohub-events/{slug}/badges", response_model=PersohubManagedBadgeResponse)
def create_badge(
    slug: str,
    payload: PersohubManagedBadgeCreate,
    admin: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    if payload.user_id and payload.team_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only one of user_id or team_id is allowed")
    badge = PersohubEventBadge(
        event_id=event.id,
        title=payload.title,
        image_url=payload.image_url,
        place=PersohubEventBadgePlace[payload.place.name],
        score=payload.score,
        user_id=payload.user_id,
        team_id=payload.team_id,
    )
    db.add(badge)
    db.commit()
    db.refresh(badge)
    _log_event_admin_action(
        db,
        admin,
        event,
        "create_persohub_event_badge",
        method="POST",
        path=f"/persohub/admin/persohub-events/{slug}/badges",
        meta={"badge_id": badge.id},
    )
    return PersohubManagedBadgeResponse.model_validate(badge)


@router.get("/persohub/admin/persohub-events/{slug}/badges", response_model=List[PersohubManagedBadgeResponse])
def list_badges(
    slug: str,
    _: PdaUser = Depends(require_persohub_event_admin),
    db: Session = Depends(get_db),
):
    event = _get_event_or_404(db, slug)
    badges = db.query(PersohubEventBadge).filter(PersohubEventBadge.event_id == event.id).order_by(PersohubEventBadge.created_at.desc()).all()
    return [PersohubManagedBadgeResponse.model_validate(badge) for badge in badges]
