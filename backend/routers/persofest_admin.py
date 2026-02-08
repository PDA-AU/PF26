from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Request, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, case, and_, or_
from typing import List, Optional
import io
import csv
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from database import get_db
from models import (
    Participant, Round, Score, SystemConfig, RoundState, ParticipantStatus,
    Department, YearOfStudy, Gender, UserRole, RoundMode
)
from schemas import (
    DashboardStats, ParticipantListResponse, ParticipantStatusEnum, AdminParticipantRoundStat,
    ParticipantLeaderboardSummary, RoundCreate, RoundUpdate, RoundResponse,
    RoundStatsResponse, RoundStatsTopEntry, ScoreEntry, ScoreResponse, LeaderboardEntry,
    AdminLogResponse, DepartmentEnum, YearOfStudyEnum, GenderEnum,
    PresignRequest, PresignResponse
)
from security import require_pda_pf_admin, require_superadmin
from utils import log_admin_action, _upload_to_s3, _generate_presigned_put_url
from models import AdminLog

router = APIRouter()


def _build_leaderboard_grouped_subquery(
    db: Session,
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    search: Optional[str] = None,
    gender: Optional[GenderEnum] = None,
):
    """Shared leaderboard dataset: all participants, scores from completed or frozen rounds."""
    base_query = db.query(
        Participant.id.label("participant_id"),
        Participant.register_number,
        Participant.name,
        Participant.email,
        Participant.department,
        Participant.year_of_study,
        Participant.gender,
        Participant.status,
        Participant.referral_count,
        Participant.profile_picture,
        func.count(
            func.distinct(
                case(
                    (
                        and_(Round.id.isnot(None), Score.is_present == True),  # noqa: E712
                        Round.id
                    ),
                    else_=None
                )
            )
        ).label("rounds_participated"),
        func.coalesce(
            func.sum(
                case(
                    (
                        and_(Round.id.isnot(None), Score.is_present == True),  # noqa: E712
                        Score.normalized_score
                    ),
                    else_=0.0
                )
            ),
            0.0
        ).label("cumulative_score")
    ).outerjoin(
        Score, Score.participant_id == Participant.id
    ).outerjoin(
        Round,
        and_(
            Round.id == Score.round_id,
            or_(Round.state == RoundState.COMPLETED, Round.is_frozen == True)  # noqa: E712
        )
    ).filter(
        Participant.role == UserRole.PARTICIPANT
    )

    if department:
        base_query = base_query.filter(Participant.department == Department[department.name])
    if year:
        base_query = base_query.filter(Participant.year_of_study == YearOfStudy[year.name])
    if gender:
        base_query = base_query.filter(Participant.gender == Gender[gender.name])
    if search:
        base_query = base_query.filter(
            (Participant.name.ilike(f"%{search}%")) |
            (Participant.register_number.ilike(f"%{search}%"))
        )

    return base_query.group_by(
        Participant.id,
        Participant.register_number,
        Participant.name,
        Participant.email,
        Participant.department,
        Participant.year_of_study,
        Participant.gender,
        Participant.status,
        Participant.referral_count,
        Participant.profile_picture,
    ).subquery()


@router.get("/persofest/admin/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    total = db.query(Participant).filter(Participant.role == UserRole.PARTICIPANT).count()
    active = db.query(Participant).filter(Participant.role == UserRole.PARTICIPANT, Participant.status == ParticipantStatus.ACTIVE).count()
    eliminated = total - active

    reg_config = db.query(SystemConfig).filter(SystemConfig.key == "registration_open").first()
    registration_open = reg_config.value == "true" if reg_config else True

    rounds_completed = db.query(Round).filter(Round.state == RoundState.COMPLETED).count()
    active_round = db.query(Round).filter(Round.state == RoundState.ACTIVE).first()

    gender_dist = db.query(Participant.gender, func.count(Participant.id)).filter(
        Participant.role == UserRole.PARTICIPANT
    ).group_by(Participant.gender).all()
    gender_distribution = {g.value: c for g, c in gender_dist}

    dept_dist = db.query(Participant.department, func.count(Participant.id)).filter(
        Participant.role == UserRole.PARTICIPANT
    ).group_by(Participant.department).all()
    department_distribution = {d.value: c for d, c in dept_dist}

    year_dist = db.query(Participant.year_of_study, func.count(Participant.id)).filter(
        Participant.role == UserRole.PARTICIPANT
    ).group_by(Participant.year_of_study).all()
    year_distribution = {y.value: c for y, c in year_dist}

    grouped_subquery = _build_leaderboard_grouped_subquery(db=db)
    active_scores = db.query(
        grouped_subquery.c.cumulative_score
    ).filter(
        grouped_subquery.c.status == ParticipantStatus.ACTIVE
    ).all()
    active_score_values = [float(row.cumulative_score or 0.0) for row in active_scores]
    if active_score_values:
        min_score = min(active_score_values)
        max_score = max(active_score_values)
        avg_score = sum(active_score_values) / len(active_score_values)
    else:
        min_score = max_score = avg_score = None

    round_min_score = round_max_score = round_avg_score = None
    if active_round:
        round_scores = db.query(Score).filter(
            Score.round_id == active_round.id,
            Score.is_present == True  # noqa: E712
        ).all()
        if round_scores:
            totals = [float(s.total_score or 0.0) for s in round_scores]
            round_min_score = min(totals)
            round_max_score = max(totals)
            round_avg_score = sum(totals) / len(totals)

    return DashboardStats(
        total_participants=total,
        registration_open=registration_open,
        rounds_completed=rounds_completed,
        current_active_round=active_round.name if active_round else None,
        active_count=active,
        eliminated_count=eliminated,
        gender_distribution=gender_distribution,
        department_distribution=department_distribution,
        year_distribution=year_distribution,
        leaderboard_min_score=min_score,
        leaderboard_max_score=max_score,
        leaderboard_avg_score=avg_score,
        round_min_score=round_min_score,
        round_max_score=round_max_score,
        round_avg_score=round_avg_score
    )


@router.post("/persofest/admin/toggle-registration")
async def toggle_registration(admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    reg_config = db.query(SystemConfig).filter(SystemConfig.key == "registration_open").first()
    if not reg_config:
        reg_config = SystemConfig(key="registration_open", value="false")
        db.add(reg_config)
    else:
        reg_config.value = "false" if reg_config.value == "true" else "true"
    db.commit()
    log_admin_action(db, admin, "toggle_registration", method="POST", path="/persofest/admin/toggle-registration", meta={"registration_open": reg_config.value})
    return {"registration_open": reg_config.value == "true"}


@router.get("/persofest/admin/participants", response_model=List[ParticipantListResponse])
async def get_participants(
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    gender: Optional[GenderEnum] = None,
    status: Optional[ParticipantStatusEnum] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    response: Response = None,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    query = db.query(Participant).filter(Participant.role == UserRole.PARTICIPANT)

    if department:
        query = query.filter(Participant.department == Department[department.name])
    if year:
        query = query.filter(Participant.year_of_study == YearOfStudy[year.name])
    if gender:
        query = query.filter(Participant.gender == Gender[gender.name])
    if status:
        query = query.filter(Participant.status == ParticipantStatus[status.name])
    if search:
        query = query.filter(
            (Participant.name.ilike(f"%{search}%")) |
            (Participant.register_number.ilike(f"%{search}%")) |
            (Participant.email.ilike(f"%{search}%"))
        )

    total_count = query.count()
    offset = (page - 1) * page_size
    participants = query.order_by(Participant.name).offset(offset).limit(page_size).all()
    if response is not None:
        response.headers["X-Total-Count"] = str(total_count)
        response.headers["X-Page"] = str(page)
        response.headers["X-Page-Size"] = str(page_size)
    return [ParticipantListResponse.model_validate(p) for p in participants]


@router.put("/persofest/admin/participants/{participant_id}/status")
async def update_participant_status(
    participant_id: int,
    new_status: ParticipantStatusEnum = Query(...),
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    participant = db.query(Participant).filter(Participant.id == participant_id, Participant.role == UserRole.PARTICIPANT).first()
    if not participant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Participant not found")

    participant.status = ParticipantStatus[new_status.name]
    db.commit()
    log_admin_action(db, admin, "update_participant_status", method="PUT", path=f"/persofest/admin/participants/{participant_id}/status", meta={"participant_id": participant_id, "new_status": new_status.value})
    return {"message": "Status updated successfully"}


@router.get("/persofest/admin/participants/{participant_id}/rounds", response_model=List[AdminParticipantRoundStat])
async def get_participant_round_stats(
    participant_id: int,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    participant = db.query(Participant).filter(Participant.id == participant_id, Participant.role == UserRole.PARTICIPANT).first()
    if not participant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Participant not found")

    rounds = db.query(Round).order_by(Round.id).all()
    stats = []
    round_rank_maps = {}

    for round in rounds:
        if round.id not in round_rank_maps:
            scores = (
                db.query(Score)
                .filter(Score.round_id == round.id, Score.is_present == True)
                .order_by(Score.normalized_score.desc())
                .all()
            )
            rank_map = {}
            for idx, score in enumerate(scores):
                rank_map[score.participant_id] = idx + 1
            round_rank_maps[round.id] = rank_map

        score = db.query(Score).filter(Score.participant_id == participant.id, Score.round_id == round.id).first()
        if score is None:
            status_label = "Eliminated" if round.state == RoundState.COMPLETED else "Pending"
        elif not score.is_present:
            status_label = "Absent"
        else:
            status_label = "Active"

        stats.append(AdminParticipantRoundStat(
            round_id=round.id,
            round_no=round.round_no,
            round_name=round.name,
            round_state=round.state,
            status=status_label,
            is_present=score.is_present if score else None,
            total_score=score.total_score if score else None,
            normalized_score=score.normalized_score if score else None,
            round_rank=round_rank_maps.get(round.id, {}).get(participant.id)
        ))
    return stats


@router.get("/persofest/admin/participants/{participant_id}/summary", response_model=ParticipantLeaderboardSummary)
async def get_participant_summary(participant_id: int, admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    participant = db.query(Participant).filter(Participant.id == participant_id, Participant.role == UserRole.PARTICIPANT).first()
    if not participant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Participant not found")

    grouped_subquery = _build_leaderboard_grouped_subquery(db)
    active_rank_subquery = db.query(
        grouped_subquery.c.participant_id.label("participant_id"),
        func.dense_rank().over(
            order_by=grouped_subquery.c.cumulative_score.desc()
        ).label("active_rank")
    ).filter(
        grouped_subquery.c.status == ParticipantStatus.ACTIVE
    ).subquery()

    summary_row = db.query(
        grouped_subquery.c.cumulative_score,
        grouped_subquery.c.status,
        active_rank_subquery.c.active_rank,
    ).outerjoin(
        active_rank_subquery,
        active_rank_subquery.c.participant_id == grouped_subquery.c.participant_id
    ).filter(
        grouped_subquery.c.participant_id == participant_id
    ).first()

    overall_points = float(summary_row.cumulative_score) if summary_row else 0.0
    overall_rank = int(summary_row.active_rank) if (summary_row and summary_row.status == ParticipantStatus.ACTIVE and summary_row.active_rank is not None) else None

    return ParticipantLeaderboardSummary(
        participant_id=participant.id,
        overall_rank=overall_rank,
        overall_points=overall_points,
    )


@router.get("/persofest/admin/rounds", response_model=List[RoundResponse])
async def get_all_rounds(admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    rounds = db.query(Round).order_by(Round.id).all()
    return [RoundResponse.model_validate(r) for r in rounds]


@router.post("/persofest/admin/rounds", response_model=RoundResponse)
async def create_round(round_data: RoundCreate, admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    latest_round = db.query(Round).order_by(Round.id.desc()).first()
    next_round_no = f"PF{(latest_round.id + 1) if latest_round else 1:02d}"

    new_round = Round(
        round_no=next_round_no,
        name=round_data.name,
        description=round_data.description,
        tags=round_data.tags,
        date=round_data.date,
        mode=round_data.mode.name,
        conducted_by=round_data.conducted_by,
        evaluation_criteria=[c.model_dump() for c in round_data.evaluation_criteria] if round_data.evaluation_criteria else None
    )
    db.add(new_round)
    db.commit()
    db.refresh(new_round)
    log_admin_action(db, admin, "create_round", method="POST", path="/persofest/admin/rounds", meta={"round_id": new_round.id})
    return RoundResponse.model_validate(new_round)


@router.put("/persofest/admin/rounds/{round_id}", response_model=RoundResponse)
async def update_round(round_id: int, round_data: RoundUpdate, admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    update_data = round_data.model_dump(exclude_unset=True)
    eliminate_absent = update_data.pop("eliminate_absent", None)
    if "mode" in update_data and round_data.mode is not None:
        update_data["mode"] = round_data.mode.name
    if "state" in update_data and round_data.state is not None:
        update_data["state"] = RoundState[round_data.state.name]

    for field, value in update_data.items():
        setattr(round, field, value)
    if round_data.evaluation_criteria is not None:
        round.evaluation_criteria = [c.model_dump() for c in round_data.evaluation_criteria]

    should_shortlist = (
        round.is_frozen
        and round.state != RoundState.COMPLETED
        and "elimination_type" in update_data
        and "elimination_value" in update_data
        and round.elimination_type
        and round.elimination_value is not None
    )

    if should_shortlist:
        apply_absent_elimination = True if eliminate_absent is None else bool(eliminate_absent)
        grouped_subquery = _build_leaderboard_grouped_subquery(db=db)
        leaderboard_query = db.query(
            grouped_subquery.c.participant_id,
            grouped_subquery.c.cumulative_score,
            grouped_subquery.c.rounds_participated,
            grouped_subquery.c.status,
        ).filter(
            grouped_subquery.c.status == ParticipantStatus.ACTIVE
        )

        active_participants = db.query(Participant).filter(
            Participant.role == UserRole.PARTICIPANT,
            Participant.status == ParticipantStatus.ACTIVE
        ).all()
        active_ids = [p.id for p in active_participants]
        round_scores = db.query(Score).filter(Score.round_id == round_id).all()
        score_map = {s.participant_id: s for s in round_scores}
        absent_ids = [
            pid for pid in active_ids
            if (pid not in score_map) or (score_map[pid].is_present is False)
        ]

        if apply_absent_elimination and absent_ids:
            absent_participants = db.query(Participant).filter(
                Participant.id.in_(absent_ids),
                Participant.role == UserRole.PARTICIPANT
            ).all()
            for participant in absent_participants:
                participant.status = ParticipantStatus.ELIMINATED

        if apply_absent_elimination:
            eligible_ids = [pid for pid in active_ids if pid not in set(absent_ids)]
        else:
            eligible_ids = active_ids
        leaderboard_query = leaderboard_query.filter(grouped_subquery.c.participant_id.in_(eligible_ids))

        leaderboard_rows = leaderboard_query.order_by(
            grouped_subquery.c.cumulative_score.desc(),
            grouped_subquery.c.participant_id.asc(),
        ).all()

        participant_ids = [row.participant_id for row in leaderboard_rows]
        participants = db.query(Participant).filter(
            Participant.id.in_(participant_ids),
            Participant.role == UserRole.PARTICIPANT
        ).all()
        participant_map = {p.id: p for p in participants}

        if round.elimination_type == "top_k":
            cutoff = max(0, int(round.elimination_value))
            for idx, row in enumerate(leaderboard_rows):
                participant = participant_map.get(row.participant_id)
                if not participant:
                    continue
                participant.status = ParticipantStatus.ACTIVE if idx < cutoff else ParticipantStatus.ELIMINATED
        elif round.elimination_type == "min_score":
            threshold = float(round.elimination_value)
            for row in leaderboard_rows:
                participant = participant_map.get(row.participant_id)
                if not participant:
                    continue
                score = float(row.cumulative_score or 0.0)
                participant.status = ParticipantStatus.ACTIVE if score >= threshold else ParticipantStatus.ELIMINATED
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid elimination type")

        round.state = RoundState.COMPLETED

    db.commit()
    db.refresh(round)
    log_admin_action(db, admin, "update_round", method="PUT", path=f"/persofest/admin/rounds/{round_id}", meta={"round_id": round_id})
    if should_shortlist:
        log_admin_action(
            db,
            admin,
            "shortlist_round",
            method="PUT",
            path=f"/persofest/admin/rounds/{round_id}",
            meta={
                "round_id": round_id,
                "elimination_type": round.elimination_type,
                "elimination_value": round.elimination_value
            }
        )
    return RoundResponse.model_validate(round)


@router.post("/persofest/admin/rounds/{round_id}/description-pdf", response_model=RoundResponse)
async def upload_round_pdf(round_id: int, file: UploadFile = File(...), admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Missing filename")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File size exceeds 5MB limit")
    file.file = io.BytesIO(contents)
    round.description_pdf = _upload_to_s3(file, "persofest/round-pdfs", allowed_types=["application/pdf"])
    db.commit()
    db.refresh(round)
    log_admin_action(db, admin, "upload_round_pdf", method="POST", path=f"/persofest/admin/rounds/{round_id}/description-pdf", meta={"round_id": round_id})
    return RoundResponse.model_validate(round)


@router.post("/persofest/admin/rounds/{round_id}/description-pdf/presign", response_model=PresignResponse)
async def presign_round_pdf(
    round_id: int,
    payload: PresignRequest,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    return _generate_presigned_put_url(
        "persofest/round-pdfs",
        payload.filename,
        payload.content_type,
        allowed_types=["application/pdf"]
    )


@router.delete("/persofest/admin/rounds/{round_id}")
async def delete_round(round_id: int, admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")
    db.delete(round)
    db.commit()
    log_admin_action(db, admin, "delete_round", method="DELETE", path=f"/persofest/admin/rounds/{round_id}", meta={"round_id": round_id})
    return {"message": "Round deleted"}


@router.get("/persofest/admin/rounds/{round_id}/participants", response_model=List[ScoreResponse])
async def get_round_participants(
    round_id: int,
    search: Optional[str] = None,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    scores = db.query(Score).filter(Score.round_id == round_id).all()
    score_map = {s.participant_id: s for s in scores}
    has_round_result = round.is_frozen or round.state == RoundState.COMPLETED

    query = db.query(Participant).filter(Participant.role == UserRole.PARTICIPANT)
    if has_round_result:
        # For completed/frozen rounds, show all participants that were evaluated in that round,
        # including absent or eliminated ones.
        participant_ids = list(score_map.keys())
        if not participant_ids:
            return []
        query = query.filter(Participant.id.in_(participant_ids))
    else:
        # For active editable rounds, keep list restricted to currently active participants.
        query = query.filter(Participant.status == ParticipantStatus.ACTIVE)

    if search:
        query = query.filter(
            (Participant.name.ilike(f"%{search}%")) |
            (Participant.register_number.ilike(f"%{search}%")) |
            (Participant.email.ilike(f"%{search}%"))
        )
    participants = query.order_by(Participant.name).all()

    results = []
    for participant in participants:
        score = score_map.get(participant.id)
        if score:
            results.append(ScoreResponse(
                id=score.id,
                participant_id=participant.id,
                round_id=round_id,
                criteria_scores=score.criteria_scores or {},
                total_score=float(score.total_score or 0),
                normalized_score=float(score.normalized_score or 0),
                is_present=bool(score.is_present),
                participant_name=participant.name,
                participant_register_number=participant.register_number,
                participant_status=participant.status,
            ))
        else:
            results.append(ScoreResponse(
                id=0,
                participant_id=participant.id,
                round_id=round_id,
                criteria_scores={},
                total_score=0,
                normalized_score=0,
                is_present=True,
                participant_name=participant.name,
                participant_register_number=participant.register_number,
                participant_status=participant.status,
            ))
    return results


@router.get("/persofest/admin/rounds/{round_id}/stats", response_model=RoundStatsResponse)
async def get_round_stats(round_id: int, admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    scores = db.query(Score).filter(Score.round_id == round_id).all()
    total_count = len(scores)
    present_scores = [s for s in scores if s.is_present]
    absent_scores = [s for s in scores if not s.is_present]

    if present_scores:
        min_score = min(s.total_score for s in present_scores)
        max_score = max(s.total_score for s in present_scores)
        avg_score = sum(s.total_score for s in present_scores) / len(present_scores)
    else:
        min_score = max_score = avg_score = None

    top10 = (
        db.query(Participant.id, Participant.name, Participant.register_number, Score.normalized_score)
        .join(Score, Score.participant_id == Participant.id)
        .filter(Score.round_id == round_id, Score.is_present == True)  # noqa: E712
        .order_by(Score.normalized_score.desc())
        .limit(10)
        .all()
    )

    return RoundStatsResponse(
        round_id=round_id,
        total_count=total_count,
        present_count=len(present_scores),
        absent_count=len(absent_scores),
        min_score=min_score,
        max_score=max_score,
        avg_score=avg_score,
        top10=[RoundStatsTopEntry(participant_id=i, name=n, register_number=r, normalized_score=s) for i, n, r, s in top10]
    )


@router.post("/persofest/admin/rounds/{round_id}/scores")
async def enter_scores(
    round_id: int,
    scores: List[ScoreEntry],
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    if round.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is frozen")

    criteria_def = round.evaluation_criteria or [{"name": "Score", "max_marks": 100}]
    criteria_max = {c["name"]: float(c["max_marks"]) for c in criteria_def if c.get("name") is not None}
    max_score = sum(criteria_max.values()) if criteria_max else 100

    for entry in scores:
        existing = db.query(Score).filter(
            Score.participant_id == entry.participant_id,
            Score.round_id == round_id
        ).first()

        is_present = bool(entry.is_present)
        if not is_present:
            # Absent participants must have wiped scores.
            validated_scores = {cname: 0.0 for cname in criteria_max.keys()}
            total = 0.0
            normalized = 0.0
        else:
            validated_scores = {}
            for cname, cmax in criteria_max.items():
                raw = (entry.criteria_scores or {}).get(cname, 0)
                try:
                    value = float(raw)
                except (TypeError, ValueError):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Invalid score for '{cname}' (participant_id={entry.participant_id})"
                    )
                if value < 0 or value > cmax:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Score for '{cname}' must be between 0 and {cmax} (participant_id={entry.participant_id})"
                    )
                validated_scores[cname] = value

            total = sum(validated_scores.values()) if validated_scores else 0
            normalized = (total / max_score * 100) if max_score > 0 else 0

        if existing:
            existing.criteria_scores = validated_scores
            existing.total_score = total
            existing.normalized_score = normalized
            existing.is_present = is_present
        else:
            new_score = Score(
                participant_id=entry.participant_id,
                round_id=round_id,
                criteria_scores=validated_scores,
                total_score=total,
                normalized_score=normalized,
                is_present=is_present
            )
            db.add(new_score)

    db.commit()
    log_admin_action(db, admin, "save_scores", method="POST", path=f"/persofest/admin/rounds/{round_id}/scores", meta={"round_id": round_id, "count": len(scores)})
    return {"message": "Scores saved successfully"}


@router.post("/persofest/admin/rounds/{round_id}/import-scores")
async def import_scores_from_excel(
    round_id: int,
    file: UploadFile = File(...),
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    if round.is_frozen:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Round is frozen")

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only Excel .xlsx files are supported")

    contents = await file.read()
    wb = load_workbook(filename=io.BytesIO(contents))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if not headers:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel header row is empty")

    criteria_def = round.evaluation_criteria or [{"name": "Score", "max_marks": 100}]
    criteria_names = [c["name"] for c in criteria_def]
    criteria_max = {c["name"]: float(c["max_marks"]) for c in criteria_def if c.get("name") is not None}
    max_score = sum(criteria_max.values()) if criteria_max else 100

    def norm_header(value):
        return str(value or "").strip().lower().replace("_", " ").replace("-", " ")

    canonical_header_aliases = {
        "register_number": {"register number", "register no", "reg no", "reg number", "register_number", "regno"},
        "present": {"present", "attendance", "is present", "status"},
    }
    for cname in criteria_names:
        canonical_header_aliases[cname] = {norm_header(cname)}

    header_to_column = {}
    for header in headers:
        normalized = norm_header(header)
        matched_column = None
        for canonical, aliases in canonical_header_aliases.items():
            if normalized in aliases:
                matched_column = canonical
                break
        if matched_column:
            header_to_column[header] = matched_column

    if "register_number" not in header_to_column.values():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel must have a Register Number column")

    imported_count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        raw_row_dict = dict(zip(headers, row))
        row_dict = {}
        for raw_key, raw_val in raw_row_dict.items():
            canonical = header_to_column.get(raw_key)
            if canonical:
                row_dict[canonical] = raw_val
        raw_reg = row_dict.get("register_number", "")
        if isinstance(raw_reg, float):
            register_number = str(int(raw_reg)) if raw_reg.is_integer() else str(raw_reg).strip()
        elif isinstance(raw_reg, int):
            register_number = str(raw_reg)
        else:
            register_number = str(raw_reg or "").strip()
        if register_number.endswith(".0"):
            register_number = register_number[:-2]
        register_number = register_number.replace(" ", "")
        if not register_number:
            continue

        participant = db.query(Participant).filter(
            Participant.register_number == register_number,
            Participant.role == UserRole.PARTICIPANT,
            Participant.status == ParticipantStatus.ACTIVE
        ).first()
        if not participant:
            errors.append(f"Row {row_idx}: Register number {register_number} not found")
            continue

        present_val = str(row_dict.get("present", "Yes")).strip().lower()
        is_present = present_val in ("yes", "y", "1", "true", "present")

        if not is_present:
            criteria_scores = {cname: 0.0 for cname in criteria_names}
            total = 0.0
            normalized = 0.0
        else:
            criteria_scores = {}
            invalid_row = False
            for cname in criteria_names:
                if cname in row_dict and row_dict[cname] is not None:
                    try:
                        score_value = float(row_dict[cname])
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_idx}: Invalid score for '{cname}'")
                        invalid_row = True
                        break
                else:
                    score_value = 0.0

                cmax = criteria_max.get(cname, 100.0)
                if score_value < 0 or score_value > cmax:
                    errors.append(f"Row {row_idx}: '{cname}' score must be between 0 and {cmax}")
                    invalid_row = True
                    break
                criteria_scores[cname] = score_value

            if invalid_row:
                continue

            total = sum(criteria_scores.values())
            normalized = (total / max_score * 100) if max_score > 0 else 0

        existing = db.query(Score).filter(
            Score.participant_id == participant.id,
            Score.round_id == round_id
        ).first()

        if existing:
            existing.criteria_scores = criteria_scores
            existing.total_score = total
            existing.normalized_score = normalized
            existing.is_present = is_present
        else:
            new_score = Score(
                participant_id=participant.id,
                round_id=round_id,
                criteria_scores=criteria_scores,
                total_score=total,
                normalized_score=normalized,
                is_present=is_present
            )
            db.add(new_score)

        imported_count += 1

    db.commit()

    log_admin_action(db, admin, "import_scores", method="POST", path=f"/persofest/admin/rounds/{round_id}/import-scores", meta={"round_id": round_id, "imported": imported_count, "errors": len(errors)})

    return {
        "message": f"Successfully imported {imported_count} scores",
        "imported": imported_count,
        "errors": errors[:10] if errors else []
    }


@router.get("/persofest/admin/rounds/{round_id}/score-template")
async def download_score_template(
    round_id: int,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    participants = db.query(Participant).filter(
        Participant.role == UserRole.PARTICIPANT,
        Participant.status == ParticipantStatus.ACTIVE
    ).order_by(Participant.name).all()

    criteria_names = [c["name"] for c in round.evaluation_criteria] if round.evaluation_criteria else ["Score"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{round.round_no} Scores"

    headers = ["Register Number", "Name", "Present"] + criteria_names
    ws.append(headers)

    for p in participants:
        row = [p.register_number, p.name, "Yes"] + [0] * len(criteria_names)
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    log_admin_action(db, admin, "download_score_template", method="GET", path=f"/persofest/admin/rounds/{round_id}/score-template", meta={"round_id": round_id, "round_no": round.round_no})
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={round.round_no}_score_template.xlsx"}
    )


@router.post("/persofest/admin/rounds/{round_id}/freeze")
async def freeze_round(round_id: int, admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    criteria_def = round.evaluation_criteria or [{"name": "Score", "max_marks": 100}]
    zero_scores = {c["name"]: 0.0 for c in criteria_def if c.get("name") is not None}

    active_participants = db.query(Participant).filter(
        Participant.role == UserRole.PARTICIPANT,
        Participant.status == ParticipantStatus.ACTIVE
    ).all()
    score_rows = db.query(Score).filter(Score.round_id == round_id).all()
    score_map = {row.participant_id: row for row in score_rows}

    for participant in active_participants:
        score = score_map.get(participant.id)
        if score:
            continue
        else:
            # Persist an explicit absent+zero row so completed round views can show the participant.
            score = Score(
                participant_id=participant.id,
                round_id=round_id,
                criteria_scores=zero_scores,
                total_score=0.0,
                normalized_score=0.0,
                is_present=False
            )
            db.add(score)
            score_map[participant.id] = score

    round.is_frozen = True
    db.commit()
    log_admin_action(db, admin, "freeze_round", method="POST", path=f"/persofest/admin/rounds/{round_id}/freeze", meta={"round_id": round_id})
    return {"message": "Round frozen"}


@router.post("/persofest/admin/rounds/{round_id}/unfreeze")
async def unfreeze_round(round_id: int, admin=Depends(require_pda_pf_admin), db: Session = Depends(get_db)):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    round.is_frozen = False
    round.state = RoundState.ACTIVE
    db.commit()
    log_admin_action(db, admin, "unfreeze_round", method="POST", path=f"/persofest/admin/rounds/{round_id}/unfreeze", meta={"round_id": round_id})
    return {"message": "Round unfrozen"}


@router.get("/persofest/admin/leaderboard", response_model=List[LeaderboardEntry])
async def get_leaderboard(
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    gender: Optional[GenderEnum] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    response: Response = None,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    grouped_subquery = _build_leaderboard_grouped_subquery(
        db=db,
        department=department,
        year=year,
        gender=gender,
        search=search,
    )

    active_rank_subquery = db.query(
        grouped_subquery.c.participant_id.label("participant_id"),
        func.dense_rank().over(
            order_by=grouped_subquery.c.cumulative_score.desc()
        ).label("active_rank")
    ).filter(
        grouped_subquery.c.status == ParticipantStatus.ACTIVE
    ).subquery()

    total_count = db.query(func.count()).select_from(grouped_subquery).scalar() or 0
    offset = (page - 1) * page_size
    leaderboard_rows = db.query(
        grouped_subquery.c.participant_id,
        grouped_subquery.c.register_number,
        grouped_subquery.c.name,
        grouped_subquery.c.email,
        grouped_subquery.c.department,
        grouped_subquery.c.year_of_study,
        grouped_subquery.c.gender,
        grouped_subquery.c.status,
        grouped_subquery.c.referral_count,
        grouped_subquery.c.profile_picture,
        grouped_subquery.c.rounds_participated,
        grouped_subquery.c.cumulative_score,
        active_rank_subquery.c.active_rank.label("rank"),
    ).outerjoin(
        active_rank_subquery,
        active_rank_subquery.c.participant_id == grouped_subquery.c.participant_id
    ).order_by(
        case((grouped_subquery.c.status == ParticipantStatus.ACTIVE, 0), else_=1),
        grouped_subquery.c.cumulative_score.desc(),
        grouped_subquery.c.participant_id.asc(),
    ).offset(offset).limit(page_size).all()

    if response is not None:
        response.headers["X-Total-Count"] = str(total_count)
        response.headers["X-Page"] = str(page)
        response.headers["X-Page-Size"] = str(page_size)

    return [
        LeaderboardEntry(
            rank=int(row.rank) if (row.status == ParticipantStatus.ACTIVE and row.rank is not None) else None,
            participant_id=row.participant_id,
            register_number=row.register_number,
            name=row.name,
            email=row.email,
            department=row.department,
            year_of_study=row.year_of_study,
            gender=row.gender,
            cumulative_score=float(row.cumulative_score),
            rounds_participated=int(row.rounds_participated),
            status=row.status,
            referral_count=int(row.referral_count or 0),
            profile_picture=row.profile_picture,
        )
        for row in leaderboard_rows
    ]


@router.get("/persofest/admin/logs", response_model=List[AdminLogResponse])
async def get_admin_logs(
    admin=Depends(require_superadmin),
    db: Session = Depends(get_db),
    limit: int = 50,
    offset: int = 0
):
    logs = db.query(AdminLog).filter(
        or_(
            AdminLog.path.like("/persofest/admin%"),
            AdminLog.path.like("/api/persofest/admin%"),
        )
    ).order_by(AdminLog.id.desc()).offset(offset).limit(limit).all()
    return [AdminLogResponse.model_validate(l) for l in logs]


@router.get("/persofest/admin/export/participants")
async def export_participants(
    format: str = "csv",
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    gender: Optional[GenderEnum] = None,
    status: Optional[ParticipantStatusEnum] = None,
    search: Optional[str] = None,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    query = db.query(Participant).filter(Participant.role == UserRole.PARTICIPANT)
    if department:
        query = query.filter(Participant.department == Department[department.name])
    if year:
        query = query.filter(Participant.year_of_study == YearOfStudy[year.name])
    if gender:
        query = query.filter(Participant.gender == Gender[gender.name])
    if status:
        query = query.filter(Participant.status == ParticipantStatus[status.name])
    if search:
        query = query.filter(
            (Participant.name.ilike(f"%{search}%")) |
            (Participant.register_number.ilike(f"%{search}%")) |
            (Participant.email.ilike(f"%{search}%"))
        )

    participants = query.order_by(Participant.name).all()
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(["Register Number", "Name", "Email", "Phone", "Department", "Year", "Status"])
        for p in participants:
            ws.append([p.register_number, p.name, p.email, p.phone, p.department.value, p.year_of_study.value, p.status.value])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        headers = {"Content-Disposition": "attachment; filename=participants.xlsx"}
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Register Number", "Name", "Email", "Phone", "Department", "Year", "Status"])
    for p in participants:
        writer.writerow([p.register_number, p.name, p.email, p.phone, p.department.value, p.year_of_study.value, p.status.value])
    headers = {"Content-Disposition": "attachment; filename=participants.csv"}
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers=headers)


@router.get("/persofest/admin/export/round/{round_id}")
async def export_round_scores(
    round_id: int,
    format: str = "csv",
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    round = db.query(Round).filter(Round.id == round_id).first()
    if not round:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Round not found")

    criteria_def = round.evaluation_criteria or []
    criteria_names = [c.get("name") for c in criteria_def if c.get("name")]
    criteria_headers = [
        f"{c.get('name')} (/{c.get('max_marks')})"
        for c in criteria_def
        if c.get("name") is not None
    ]

    scores = (
        db.query(Score, Participant)
        .join(Participant, Participant.id == Score.participant_id)
        .filter(Score.round_id == round_id)
        .all()
    )
    base_headers = ["Register Number", "Name", "Present", "Total Score", "Normalized Score"]
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(base_headers + criteria_headers)
        for score, participant in scores:
            criteria_scores = score.criteria_scores or {}
            criteria_values = [criteria_scores.get(name, 0) for name in criteria_names]
            ws.append([
                participant.register_number,
                participant.name,
                score.is_present,
                score.total_score,
                score.normalized_score,
                *criteria_values
            ])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        headers = {"Content-Disposition": f"attachment; filename=round_{round_id}.xlsx"}
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(base_headers + criteria_headers)
    for score, participant in scores:
        criteria_scores = score.criteria_scores or {}
        criteria_values = [criteria_scores.get(name, 0) for name in criteria_names]
        writer.writerow([
            participant.register_number,
            participant.name,
            score.is_present,
            score.total_score,
            score.normalized_score,
            *criteria_values
        ])
    headers = {"Content-Disposition": f"attachment; filename=round_{round_id}.csv"}
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers=headers)


@router.get("/persofest/admin/export/leaderboard")
async def export_leaderboard(
    format: str = "csv",
    department: Optional[DepartmentEnum] = None,
    year: Optional[YearOfStudyEnum] = None,
    gender: Optional[GenderEnum] = None,
    search: Optional[str] = None,
    admin=Depends(require_pda_pf_admin),
    db: Session = Depends(get_db)
):
    grouped_subquery = _build_leaderboard_grouped_subquery(
        db=db,
        department=department,
        year=year,
        gender=gender,
        search=search
    )
    active_rank_subquery = db.query(
        grouped_subquery.c.participant_id.label("participant_id"),
        func.dense_rank().over(
            order_by=grouped_subquery.c.cumulative_score.desc()
        ).label("active_rank")
    ).filter(
        grouped_subquery.c.status == ParticipantStatus.ACTIVE
    ).subquery()

    leaderboard_rows = db.query(
        grouped_subquery.c.register_number,
        grouped_subquery.c.name,
        grouped_subquery.c.department,
        grouped_subquery.c.year_of_study,
        grouped_subquery.c.status,
        grouped_subquery.c.rounds_participated,
        grouped_subquery.c.cumulative_score,
        active_rank_subquery.c.active_rank.label("rank"),
    ).outerjoin(
        active_rank_subquery,
        active_rank_subquery.c.participant_id == grouped_subquery.c.participant_id
    ).order_by(
        case((grouped_subquery.c.status == ParticipantStatus.ACTIVE, 0), else_=1),
        grouped_subquery.c.cumulative_score.desc(),
        grouped_subquery.c.participant_id.asc(),
    ).all()

    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(["Rank", "Register Number", "Name", "Department", "Year", "Status", "Rounds", "Total Score"])
        for row in leaderboard_rows:
            rank = int(row.rank) if (row.status == ParticipantStatus.ACTIVE and row.rank is not None) else ""
            ws.append([
                rank,
                row.register_number,
                row.name,
                row.department.value if row.department else "",
                row.year_of_study.value if row.year_of_study else "",
                row.status.value if row.status else "",
                int(row.rounds_participated or 0),
                float(row.cumulative_score or 0.0),
            ])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        headers = {"Content-Disposition": "attachment; filename=leaderboard.xlsx"}
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Rank", "Register Number", "Name", "Department", "Year", "Status", "Rounds", "Total Score"])
    for row in leaderboard_rows:
        rank = int(row.rank) if (row.status == ParticipantStatus.ACTIVE and row.rank is not None) else ""
        writer.writerow([
            rank,
            row.register_number,
            row.name,
            row.department.value if row.department else "",
            row.year_of_study.value if row.year_of_study else "",
            row.status.value if row.status else "",
            int(row.rounds_participated or 0),
            float(row.cumulative_score or 0.0),
        ])
    headers = {"Content-Disposition": "attachment; filename=leaderboard.csv"}
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers=headers)
